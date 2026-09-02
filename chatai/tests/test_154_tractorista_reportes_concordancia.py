"""Screen, Excel, PDF and /reportes bulk must share tractorista report builders."""

import asyncio
import datetime
import io
import os
import sys
from collections import defaultdict

import openpyxl
import psycopg2
import pytest
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"), interpolate=False)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))
import controllers.reports_controller as rc  # noqa: E402
import controllers.tarjas_controller as tc  # noqa: E402

AUG_INICIO = "2026-08-01"
AUG_TERMINO = "2026-09-01"


@pytest.fixture(scope="module")
def conn():
    c = psycopg2.connect(
        host=os.environ["DB_HOST"],
        port=int(os.environ.get("DB_PORT", "5432")),
        dbname=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
    )
    yield c
    c.rollback()
    c.close()


def run(coro):
    return asyncio.run(coro)


async def _body_bytes(resp):
    if getattr(resp, "body", None):
        return resp.body
    return b"".join([c async for c in resp.body_iterator])


def test_154_ranking_totals_match_resumen_by_worker_contratista():
    gt = run(
        tc.get_tarjas_general_tractorista(
            fecha_inicio=AUG_INICIO,
            fecha_termino=AUG_TERMINO,
            centro_costo=None,
            labor=None,
            maquina=None,
            contratista=None,
            empresa=None,
        )
    )
    rp = run(
        tc.get_tarjas_resumen_persona_tractorista(
            fecha_inicio=AUG_INICIO,
            fecha_termino=AUG_TERMINO,
            trabajador=None,
            tipo_pago=None,
            maquina=None,
            contratista=None,
            empresa=None,
        )
    )
    gt_sum = sum(float(r["total"] or 0) for r in gt["person_ranking"])
    labor_sum = sum(float(r["total"] or 0) for r in gt["labor_summary"])
    rp_sum = sum(float(r["total_tractor"] or 0) for r in rp["rows"])
    assert abs(gt_sum - rp_sum) < 0.01
    assert abs(labor_sum - rp_sum) < 0.01

    rp_by = defaultdict(float)
    for r in rp["rows"]:
        rp_by[(r.get("trabajador"), r.get("contratista"))] += float(
            r.get("total_tractor") or 0
        )
    for row in gt["person_ranking"]:
        key = (row.get("trabajador"), row.get("contratista"))
        assert abs(float(row["total"] or 0) - rp_by[key]) < 0.01, key


def test_154_general_pdf_html_matches_screen_tables(conn):
    with conn.cursor() as cur:
        where, params = tc._general_tractorista_where(cur, AUG_INICIO, AUG_TERMINO)
        labor_summary, person_ranking = tc._fetch_general_tractorista_tables(
            cur, where, params
        )
        html = tc._build_general_tractorista_html(cur, AUG_INICIO, AUG_TERMINO)
    assert "Ganancia promedio por labor" in html
    assert "Ranking por persona" in html
    assert ">Promedio<" in html
    assert "Tipo de pago" not in html
    labor_sum = sum(float(r["total"] or 0) for r in labor_summary)
    rank_sum = sum(float(r["total"] or 0) for r in person_ranking)
    assert tc._fmt_clp(labor_sum) in html
    assert tc._fmt_clp(rank_sum) in html
    for r in labor_summary:
        assert (r["labor"] or "") in html
    for r in person_ranking[:5]:
        assert (r["trabajador"] or "") in html


def test_154_resumen_pdf_html_matches_screen_columns(conn):
    with conn.cursor() as cur:
        where, params = tc._resumen_persona_tractorista_where(
            cur, AUG_INICIO, AUG_TERMINO
        )
        rows = tc._fetch_resumen_persona_tractorista_rows(cur, where, params)
        html = tc._build_resumen_persona_tractorista_html(cur, AUG_INICIO, AUG_TERMINO)
    assert ">Máquina<" in html
    assert ">Horas extra<" in html
    assert ">Contratista<" in html
    assert "Tipo pago" not in html
    assert "Suma total" in html
    total = sum(float(r["total_tractor"] or 0) for r in rows)
    assert tc._fmt_clp(total) in html
    dates, groups = tc._pivot_resumen_persona_tractorista(rows)
    assert dates
    for d in dates:
        assert tc._fmt_date_slash(d) in html
    for g in groups[:3]:
        assert g["trabajador"] in html


def test_154_bulk_pdf_uses_same_html_as_standalone(conn):
    with conn.cursor() as cur:
        expected = tc._build_general_tractorista_html(cur, AUG_INICIO, AUG_TERMINO)
    with conn.cursor() as cur:
        actual = rc._REPORT_GENERATORS["general-tractorista"](
            cur, AUG_INICIO, AUG_TERMINO, None, None
        )
    assert actual == expected
    with conn.cursor() as cur:
        expected = tc._build_resumen_persona_tractorista_html(
            cur, AUG_INICIO, AUG_TERMINO
        )
    with conn.cursor() as cur:
        actual = rc._REPORT_GENERATORS["resumen-tractorista"](
            cur, AUG_INICIO, AUG_TERMINO, None, None
        )
    assert actual == expected


def test_154_general_excel_two_sheets_match_screen(conn, monkeypatch):
    class _KeepOpen:
        def __init__(self, inner):
            self._inner = inner

        def cursor(self, *a, **kw):
            return self._inner.cursor(*a, **kw)

        def close(self):
            return None

        def __getattr__(self, name):
            return getattr(self._inner, name)

    monkeypatch.setattr(tc, "get_connection", lambda: _KeepOpen(conn))
    with conn.cursor() as cur:
        where, params = tc._general_tractorista_where(cur, AUG_INICIO, AUG_TERMINO)
        labor_summary, person_ranking = tc._fetch_general_tractorista_tables(
            cur, where, params
        )
    resp = run(
        tc.download_tarjas_general_tractorista_excel(
            fecha_inicio=AUG_INICIO,
            fecha_termino=AUG_TERMINO,
            centro_costo=None,
            labor=None,
            maquina=None,
            contratista=None,
            empresa=None,
        )
    )
    body = run(_body_bytes(resp))
    wb = openpyxl.load_workbook(io.BytesIO(body))
    assert wb.sheetnames == ["Por labor", "Ranking"]
    ws = wb["Por labor"]
    assert [c.value for c in ws[1]] == ["Labor", "Promedio", "Total"]
    labor_sum = sum(float(r["total"] or 0) for r in labor_summary)
    last = ws.max_row
    assert ws.cell(last, 1).value == "Suma total"
    assert abs(float(ws.cell(last, 3).value or 0) - labor_sum) < 0.01
    ws2 = wb["Ranking"]
    assert [c.value for c in ws2[1]] == [
        "Trabajador",
        "Contratista",
        "Promedio",
        "Total",
    ]
    rank_sum = sum(float(r["total"] or 0) for r in person_ranking)
    last = ws2.max_row
    assert ws2.cell(last, 1).value == "Suma total"
    assert abs(float(ws2.cell(last, 4).value or 0) - rank_sum) < 0.01


def test_154_resumen_excel_columns_match_screen(conn, monkeypatch):
    class _KeepOpen:
        def __init__(self, inner):
            self._inner = inner

        def cursor(self, *a, **kw):
            return self._inner.cursor(*a, **kw)

        def close(self):
            return None

        def __getattr__(self, name):
            return getattr(self._inner, name)

    monkeypatch.setattr(tc, "get_connection", lambda: _KeepOpen(conn))
    with conn.cursor() as cur:
        where, params = tc._resumen_persona_tractorista_where(
            cur, AUG_INICIO, AUG_TERMINO
        )
        rows = tc._fetch_resumen_persona_tractorista_rows(cur, where, params)
    resp = run(
        tc.download_tarjas_resumen_persona_tractorista_excel(
            fecha_inicio=AUG_INICIO,
            fecha_termino=AUG_TERMINO,
            trabajador=None,
            tipo_pago=None,
            maquina=None,
            contratista=None,
            empresa=None,
        )
    )
    body = run(_body_bytes(resp))
    wb = openpyxl.load_workbook(io.BytesIO(body))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[:4] == ["Trabajador", "Contratista", "Máquina", "Horas extra"]
    assert headers[-1] == "Total"
    assert "Tipo de pago" not in headers
    dates, _groups = tc._pivot_resumen_persona_tractorista(rows)
    excel_dates = headers[4:-1]
    expected_dates = [
        datetime.date.fromisoformat(d).strftime("%d/%m/%Y") for d in dates
    ]
    assert excel_dates == expected_dates
    last = ws.max_row
    assert ws.cell(last, 1).value == "Suma total"
    rp_sum = sum(float(r["total_tractor"] or 0) for r in rows)
    assert abs(float(ws.cell(last, len(headers)).value or 0) - rp_sum) < 0.01

