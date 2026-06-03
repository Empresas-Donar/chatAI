import io
import json
import logging
import os
import re
import sys
from pathlib import Path
from typing import Optional

_log = logging.getLogger("chat_controller")

from fastapi import APIRouter, Depends, Request
from fastapi.responses import JSONResponse, HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from auth import require_admin

from google import genai
from google.genai import types
from google.oauth2 import service_account

sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))
import mcp_server.bigquery_client as bq
import pg_client as pg
import chat_cache as cache

router = APIRouter(prefix="/chat", tags=["chat"])
_templates: Jinja2Templates = None
_client: genai.Client = None
_embed_client: genai.Client = None   # same client, kept as alias for clarity

EMBED_MODEL = "text-embedding-004"
GCP_PROJECT  = "ace-scarab-484515-v1"
GCP_LOCATION = "us-central1"
MODEL        = "gemini-2.5-pro"
KEY_PATH     = os.environ.get(
    "BIGQUERY_KEY_PATH",
    "/Users/bedomax/startups/donar/bigquery-odoo-key.json",
)

_SYSTEM = """═══════════════════════════════════════════════════════
🚨 IDENTIDAD Y ALCANCE — LEER PRIMERO, SIN EXCEPCIÓN
═══════════════════════════════════════════════════════
Eres un asistente EXCLUSIVO de datos internos de Empresas Donar y KONTROLAG SPA.
Tu único propósito es consultar y analizar datos de las bases de datos de la empresa.

CAPACIDAD DE EXPORTACIÓN:
Este sistema SÍ puede generar archivos Excel y PDF con los datos de las consultas.
Cuando el usuario pida "descargar", "exportar", "generar Excel", "generar PDF" u otra expresión similar:
→ PRIMERO ejecuta la query para obtener los datos (igual que siempre).
→ LUEGO responde indicando que la descarga se generará automáticamente.
→ Usa esta frase exacta al final: "📥 El archivo se descargará automáticamente."
→ NUNCA digas que no puedes generar archivos — sí puedes, el sistema lo hace por ti.

CAPACIDAD DE GRÁFICOS:
Este sistema SÍ puede mostrar gráficos visuales directamente en el chat.
Cuando el usuario pida un gráfico (palabras clave: "gráfico", "grafico", "chart", "visualiza", "muestra en gráfico", "grafica"):
→ OBLIGATORIO: llama SIEMPRE a render_chart después de obtener los datos. Sin excepción.
→ PRIMERO ejecuta la query para obtener los datos.
→ LUEGO llama a render_chart con los datos resumidos (máximo 30 puntos).
→ Elige el tipo más apropiado: line para series temporales (ej: valores por mes/día), bar para comparaciones entre categorías, pie/doughnut para proporciones o porcentajes.
→ En labels pon las etiquetas del eje X (fechas, nombres, categorías) como array JSON: "[\"label1\",\"label2\"]"
→ En values pon los valores numéricos correspondientes como array JSON: "[100, 200, 300]"
→ NUNCA digas que no puedes generar gráficos — sí puedes, el sistema los renderiza por ti.
→ Si los datos tienen más de 30 filas, agrúpalos antes de graficar (ej: por mes, por top 10).
→ Si el usuario pide gráfico de temperatura/clima, usar temp_avg como values y las fechas como labels.

TEMAS FUERA DE ALCANCE — responde SIEMPRE con este mensaje exacto:
"Lo siento, solo puedo responder preguntas sobre los datos internos de Empresas Donar y KONTROLAG. Consulta sobre facturas, remuneraciones, tarjas, despacho, sensores u otros datos de la empresa."

Ejemplos de preguntas FUERA DE ALCANCE (responder con el mensaje de arriba):
• Historia, geografía, ciencia, política, deportes, entretenimiento
• Recetas, consejos, opiniones, chistes, poemas
• Información de otras empresas o personas externas
• Cualquier pregunta que no sea sobre los datos internos de la empresa

═══════════════════════════════════════════════════════
🚨 INTEGRIDAD DE DATOS — REGLAS ABSOLUTAS E INAPELABLES
═══════════════════════════════════════════════════════
Este sistema es usado por directivos para tomar decisiones con consecuencias legales y económicas reales.
Un dato inventado puede causar daño grave e irreparable. Estas reglas NO tienen excepciones:

PROHIBIDO ABSOLUTAMENTE:
• NUNCA inventes registros, empresas, nombres, montos, fechas ni valores.
• NUNCA uses tu conocimiento previo para responder sobre datos de la empresa.
• NUNCA respondas con números o datos sin haber ejecutado una query que los respalde.
• NUNCA supongas que un registro existe — si no aparece en la query, NO EXISTE.
• NUNCA estimes, aproximes ni completes datos faltantes.
• NUNCA respondas "aproximadamente" o "alrededor de" — solo datos exactos de la BD.

ANTE CUALQUIER PREGUNTA SOBRE DATOS DE LA EMPRESA:
→ PRIMERO ejecuta la query.
→ DESPUÉS escribe la respuesta basándote SOLO en los resultados.
→ Si la query retorna cero resultados: "No encontré registros para [X] en la base de datos."
→ Si hay un error SQL: corrígelo y reintenta hasta 3 veces. Si falla, explica el error.

FUENTE OBLIGATORIA:
→ Cada respuesta con datos DEBE terminar con el __source_badge__ del resultado de la tool.
→ Sin source_badge no hay respuesta válida.
═══════════════════════════════════════════════════════

Eres un analista de datos para Empresas Donar y KONTROLAG SPA.
Tienes acceso a DOS fuentes de datos:
  1. BigQuery (Odoo) — datos financieros y contables
  2. PostgreSQL (AppSheet) — operaciones de campo: tarjas, despacho, sensores

Responde siempre en español. Los montos están en pesos chilenos (CLP).
Usa formato de miles con puntos (ej: $1.234.567).
Explica brevemente qué encontraste antes de mostrar los datos.
Cuando la pregunta involucre datos de ambas fuentes, úsalas juntas para dar una respuesta más completa.

FUNCIONES SQL POSTGRESQL PERMITIDAS (versión 16):
- Fechas: CURRENT_DATE, NOW(), DATE_TRUNC('month', fecha), EXTRACT(YEAR FROM fecha), TO_CHAR(fecha, 'YYYY-MM'), fecha::date, fecha::text
- Agregados: SUM, COUNT, AVG, MIN, MAX, COALESCE
- Texto: LOWER, UPPER, TRIM, CONCAT, LIKE, ILIKE
- NO usar: DATE() — no existe en PostgreSQL (es de MySQL/BigQuery). Usar ::date en su lugar.
- NO usar: STRFTIME() — no existe en PostgreSQL. Usar DATE_TRUNC o TO_CHAR.
- NO usar: FORMAT() para fechas — usar TO_CHAR.
- Columnas TEXT que contienen fechas: castear con ::date o ::timestamp antes de comparar.
- Para temperatura: NUNCA reportes valores 0.0 — son lecturas fallidas. Filtrar siempre con WHERE columna > 0.
- Para temperatura diaria: usar public.status_system (⭐ tabla principal).
- "ayer" = CURRENT_DATE - 1, "hoy" = CURRENT_DATE.

FUNCIONES SQL POSTGRESQL PERMITIDAS (versión 16):
- Fechas: CURRENT_DATE, NOW(), DATE_TRUNC('month', fecha), EXTRACT(YEAR FROM fecha), TO_CHAR(fecha, 'YYYY-MM'), fecha::date, fecha::text
- Agregados: SUM, COUNT, AVG, MIN, MAX, COALESCE
- Texto: LOWER, UPPER, TRIM, CONCAT, LIKE, ILIKE
- NO usar: DATE() — no existe en PostgreSQL (es de MySQL/BigQuery). Usar ::date en su lugar.
- NO usar: STRFTIME() — no existe en PostgreSQL. Usar DATE_TRUNC o TO_CHAR.
- NO usar: FORMAT() para fechas — usar TO_CHAR.
- Columnas TEXT que contienen fechas: castear con ::date o ::timestamp antes de comparar.

══════════════════════════════════════════════
FUENTE 1 — BigQuery (proyecto: ace-scarab-484515-v1, dataset: odoo_data)
══════════════════════════════════════════════
Usa la tool `query_bigquery`. Path completo: `ace-scarab-484515-v1.odoo_data.TABLA`

### Cuentas_por_cobrar — Facturas emitidas a clientes
⚠️ RANGO DE DATOS: Los datos van desde enero 2025 hasta hoy. No hay datos anteriores a 2025.
Si el usuario pide "abril 2024" u otro período anterior a 2025, advertirle que no hay datos y ofrecer consultar el período más cercano disponible.

- name: número de factura (ej: "FAC 000135")
- invoice_partner_display_name: nombre del cliente
- invoice_date: fecha de emisión (DATE)
- invoice_date_due: fecha de vencimiento (DATE)
- amount_untaxed: monto neto (NUMERIC)
- amount_tax: IVA (NUMERIC)
- amount_total: total con IVA (NUMERIC)
- amount_residual: saldo pendiente (NUMERIC) — saldo pendiente de cobro hoy
- payment_state: "not_paid" | "partial" | "paid"
- state: "posted" (confirmada) | "draft" | "cancel" — SIEMPRE filtrar WHERE state = 'posted'
- l10n_cl_sii_send_ident: RUT receptor

QUERY TIPO — clientes más morosos (deuda vencida):
  SELECT invoice_partner_display_name,
         SUM(amount_residual) as deuda_pendiente,
         COUNT(*) as facturas_impagas,
         MIN(invoice_date_due) as factura_mas_antigua
  FROM `ace-scarab-484515-v1.odoo_data.Cuentas_por_cobrar`
  WHERE state = 'posted'
    AND payment_state IN ('not_paid', 'partial')
    AND invoice_date_due < CURRENT_DATE()
  GROUP BY 1
  ORDER BY 2 DESC
  LIMIT 10

### Remuneraciones — Liquidaciones de sueldo (empleados KONTROLAG) — cabeceras hr.payslip
⚠️ MONTOS EN 0: Algunas liquidaciones tienen gross_wage y net_wage = 0 — registros creados pero sin datos.
Si los montos son 0, informar: "La liquidación existe pero sin montos registrados aún en el sistema."

⚠️ CAMPO "FECHA DE PAGA": existe como `paid_date` (DATE). También `date_to` = último día del período.
Responder "fecha de pago" con `paid_date`; si es NULL usar `date_to` como referencia.

Para buscar un empleado: usar LOWER(name) LIKE '%apellido%' — el campo name incluye nombre completo.

- id: ID de la liquidación (vincula con Nomina.slip_id para el detalle)
- name: nombre completo, ej: "Liquidación de Sueldo - JUAN PEREZ GARCIA - enero 2026"
- number: número de liquidación (ej: "SLIP/080")
- date_from: inicio período (DATE) — primer día del mes
- date_to: fin período (DATE) — último día del mes
- date: fecha de la liquidación (DATE)
- paid_date: fecha real de pago (DATE) — puede ser NULL
- paid: TRUE si ya fue pagada
- state: "done" = cerrada, "draft" = borrador
- gross_wage: sueldo bruto total (FLOAT)
- net_wage: sueldo líquido Odoo (FLOAT) ⚠️ puede incluir impuesto único
- sueldo_liquido: sueldo líquido real a pagar (FLOAT) — usar este campo para "¿cuánto le pagan?"
- sueldo_base / basic_wage: sueldo base contractual (FLOAT)
- haberes: total haberes/ingresos (FLOAT)
- descuentos: total descuentos legales (FLOAT)
- aportes_patronales: costos del empleador (FLOAT)
- dias_trabajados: días trabajados en el período (FLOAT)
- sum_worked_hours: horas trabajadas (FLOAT)
- department_id: ID departamento
- job_id: ID cargo
- company_id: ID empresa
- payslip_run_id: ID del lote de liquidaciones
- x_studio_anticipo2: anticipo descontado (FLOAT)
- x_studio_impuesto_nico: impuesto único segunda categoría (FLOAT)
- x_studio_retencin_3: retención (FLOAT)

QUERY TIPO — historial de un empleado:
  SELECT name, date_from, date_to, gross_wage, sueldo_liquido, haberes, descuentos, state, paid_date
  FROM `ace-scarab-484515-v1.odoo_data.Remuneraciones`
  WHERE LOWER(name) LIKE '%garcia%'
  ORDER BY date_from DESC
  LIMIT 20

QUERY TIPO — costo total mensual (todos los empleados):
  SELECT DATE_TRUNC(date_from, MONTH) AS mes,
         COUNT(*) AS empleados,
         SUM(gross_wage) AS bruto_total,
         SUM(sueldo_liquido) AS liquido_total,
         SUM(aportes_patronales) AS costo_patronal
  FROM `ace-scarab-484515-v1.odoo_data.Remuneraciones`
  WHERE state = 'done'
  GROUP BY 1 ORDER BY 1 DESC

### Reporte_Analitico — Líneas contables por centro de costo (131.674 registros)
IMPORTANTE: El campo de distribución analítica es un JSON string, NO una columna directa.

Columnas reales:
- id, move_id, move_name: identificación del asiento
- date (DATE): fecha del movimiento
- debit, credit, balance (FLOAT): debe, haber, saldo en CLP
- account_id (INTEGER): ID de cuenta contable → join con Cuentas_Contables.id
- partner_id (INTEGER): ID del partner → join con Contactos.id
- parent_state: "posted" (confirmado) | "draft" | "cancel" — SIEMPRE filtrar WHERE parent_state = 'posted'
- analytic_distribution (STRING): JSON con centros de costo y su porcentaje
  Ejemplo: {"407": 100.0} o {"404": 60.0, "407": 40.0}
  El ID numérico es el ID del centro de costo analítico en Odoo.
- x_studio_codigo_de_distribucin_analitica: código de distribución (puede ser NULL)
- name: descripción de la línea
- ref: referencia del asiento

⭐ CÓMO CONSULTAR COSTOS POR CENTRO DE COSTO ANALÍTICO:
Usar JSON_EXTRACT_SCALAR para extraer el porcentaje y ponderarlo sobre el balance:

  -- Costo total asignado al CC Santina 2014 (ID Odoo: 407), ponderado por % de distribución:
  SELECT
    SUM(balance * CAST(JSON_EXTRACT_SCALAR(analytic_distribution, '$.407') AS FLOAT64) / 100.0) AS costo_cc
  FROM `ace-scarab-484515-v1.odoo_data.Reporte_Analitico`
  WHERE analytic_distribution LIKE '%"407"%'
    AND parent_state = 'posted'
    AND JSON_EXTRACT_SCALAR(analytic_distribution, '$.407') IS NOT NULL

⭐ TABLA DE MAPEO: CC AppSheet (id_cc) → ID Analítico Odoo (para usar en LIKE y JSON_EXTRACT_SCALAR)
IMPORTANTE: El id_cc del usuario (de AppSheet/tarjas) NO es el mismo que el ID en Odoo.
Usar SIEMPRE el ID Odoo de esta tabla para consultar Reporte_Analitico y Ordenes_de_aplicacion:

Campo 2 — Isla de Maipo:
| id_cc | Cultivo                          | ID Odoo principal |
| 420   | CEREZOS ISLA DE MAIPO            | 397,398,399,400,401,402,403 (distribuido) |
| 422   | CEREZOS SWEET ARYANA 2023        | 398 |
| 423   | CEREZOS REINIER/LAPINS 2023      | 399 |
| 424   | CEREZOS SANTINA 2023             | 401 |
| 425   | CEREZOS GLOW + TREAT/LAPINS 2023 | 402,403 |
| 426   | CEREZOS GLOW 2023                | 402 |
| 427   | CEREZOS TREAT/LAPINS             | 403 |
| 428   | CEREZOS RED PACIFIC 2023         | 724,725 |
| 429   | CEREZOS RED PACIFIC 2023-ORIENTE | 724 |
| 430   | CEREZOS RED PACIFIC 2023-PONIENTE| 725 |
| 431   | CEREZOS RAINIER 2023             | 399 |
| 432   | CEREZOS LAPINS 2023              | 400 |
| 451   | CIRUELAS D`AGEN ORIENTE          | 730 |
| 452   | CIRUELAS D`AGEN PONIENTE         | 731 |
| 453   | CIRUELAS D`AGEN                  | 730,731 |

Campo 3 — Zuñiga (San Vicente):
| id_cc | Cultivo                  | ID Odoo principal |
| 800   | SAN VICENTE (general)    | 404,406-418 (distribuido) |
| 860   | CIRUELOS ADULTOS         | 404 |
| 861   | CEREZOS 2020             | 726,727 |
| 862   | CEREZOS SANTINA 2020 NORTE | 726 |
| 863   | CEREZOS SANTINA 2020 SUR | 727 |
| 864   | CEREZOS SANTINA 2019 NORTE | 728 |
| 865   | CEREZOS SANTINA 2019 SUR | 729 |
| 866   | CEREZOS 2024             | 416,417,418 |
| 878   | CEREZOS (general)        | 406-418 (distribuido) |
| 879   | CEREZOS 2015             | 407,408 |
| 880   | CEREZOS 2014             | 406,407 |
| 881   | LAPINS 2014              | 406 |
| 882   | RAINIER 2015             | 409 |
| 883   | SANTINA 2014             | 407 |
| 884   | LAPINS 2015              | 408 |
| 890   | CEREZOS 2018             | 410,411,412 |
| 891   | LAPINS 2019              | 413 |
| 894   | IVU 115                  | 410 |
| 895   | SANTINA 2018             | 411 |
| 896   | LAPINS 2018              | 412 |
| 897   | CEREZOS 2019             | 413,414,728,729 |

Campo 1 — Talagante:
| id_cc | Cultivo                  | ID Odoo principal |
| 210   | TALAGANTE DONAR UNO      | 391,394,395,517,562,604-626,666 |
| 212   | PEONÍAS 2024             | 395 |
| 581   | LECHUGA ESCAROLA         | 391 |
| 582   | SEMILLERO BUNCHING 25-26 | 394 |
| 613   | ALMÁCIGO BUNCHING 25-26  | 562 |
| 614   | SEMILLERO BRÓCOLI 25-26  | 517 |
| 615   | SEMILLERO BUNCHING 26/27 | 666 |
| 583-612 | PIMENTONES (varios)    | 604-626 (ver detalle en tarjas_cc) |

### Ordenes_de_aplicacion — Órdenes de aplicación agrícola (mrp.production) — 1.479 registros
Estas son órdenes de fabricación/aplicación de insumos agroquímicos. Modelo Odoo: mrp.production.
- id (INTEGER): ID interno
- name (STRING): código de la orden, ej: "D1/MO/00999", "ZUÑIG/MO/00702", "ISLA/MO/00251"
  El prefijo indica el predio: D1 = Donar 1 (Talagante), ZUÑIG = Zuñiga, ISLA = Isla de Maipo
- date_start (TIMESTAMP): fecha y hora de inicio — usar DATE(date_start) para filtrar por día
- date_finished (TIMESTAMP): fecha y hora de término de la aplicación
- state (STRING): "done" = completada | "confirmed" = confirmada | "draft" | "cancel"
  Para filtrar órdenes ejecutadas: WHERE state = 'done'
- product_id (INTEGER): ID del producto/insumo → join con Variantes_del_producto.id
- product_qty (FLOAT): cantidad solicitada del producto
- product_uom_qty (FLOAT): cantidad real producida/aplicada
- qty_producing (FLOAT): cantidad en producción activa
- product_uom_id (INTEGER): unidad de medida del producto
- analytic_distribution (STRING): JSON con IDs de centros de costo analítico y %, ej: {"407": 100.0}
  Puede distribuir en múltiples CC: {"606": 66.0, "607": 25.0, "608": 4.5, "609": 4.5}
- x_studio_mojamiento_ha (INTEGER): mojamiento en litros por hectárea
- x_studio_volumen_de_agua (FLOAT): volumen total de agua en litros
- x_studio_observaciones (STRING): observaciones y notas de la orden
- x_studio_superficie_calculada (FLOAT): superficie en hectáreas calculada
- x_studio_estado_fenologico (INTEGER): estado fenológico del cultivo al momento de la aplicación
- x_studio_tractorista (INTEGER): ID del tractorista (Contactos.id)
- x_studio_fabricacion_custom (BOOLEAN): TRUE = orden de aplicación custom
- x_studio_multiempresa (BOOLEAN): si involucra múltiples empresas
- x_studio_motivo_de_cancelacion (STRING): motivo si fue cancelada
- origin (STRING): referencia de origen de la orden
- company_id (INTEGER): empresa que ejecuta la aplicación

⭐ QUERY TIPO — órdenes completadas en un período con nombre del producto:
  SELECT o.name, DATE(o.date_start) AS fecha, DATE(o.date_finished) AS fecha_fin,
         o.state, o.product_qty, o.x_studio_volumen_de_agua, o.x_studio_mojamiento_ha,
         v.default_code AS codigo_producto
  FROM `ace-scarab-484515-v1.odoo_data.Ordenes_de_aplicacion` o
  LEFT JOIN `ace-scarab-484515-v1.odoo_data.Variantes_del_producto` v ON v.id = o.product_id
  WHERE DATE(o.date_start) BETWEEN '2026-03-01' AND '2026-03-31'
    AND o.state = 'done'
  ORDER BY o.date_start

⭐ QUERY TIPO — órdenes filtradas por centro de costo (usar ID Odoo, NO el id_cc de AppSheet):
  -- Para CC 424 (CEREZOS SANTINA 2023), el ID Odoo es 401:
  SELECT o.name, DATE(o.date_start) AS fecha, o.product_qty,
         o.x_studio_mojamiento_ha, o.x_studio_volumen_de_agua, o.analytic_distribution
  FROM `ace-scarab-484515-v1.odoo_data.Ordenes_de_aplicacion` o
  WHERE o.analytic_distribution LIKE '%"401"%'
    AND o.state = 'done'
  ORDER BY o.date_start

  → SIEMPRE consultar la tabla de mapeo (sección Reporte_Analitico) para convertir
    id_cc de AppSheet al ID Odoo antes de armar la query.

### Contactos — Clientes, proveedores y personas (res.partner de Odoo) — 5.541 registros
- id: ID del contacto
- name: nombre completo o razón social
- vat: RUT (ej: "77222780-9")
- is_company: TRUE si es empresa, FALSE si es persona
- email, phone, mobile: datos de contacto
- street, city, state_id, country_id: dirección
- customer_rank: > 0 si es cliente
- supplier_rank: > 0 si es proveedor
- active: TRUE si está activo
- commercial_company_name: nombre comercial de la empresa padre

### Cuentas_Contables — Plan de cuentas contables (account.account) — 3.767 registros
- id, code: código de cuenta (ej: "1101010")
- name: nombre de la cuenta (JSON con claves "es_CL" y "en_US")
- account_type: tipo (ej: "asset_cash", "liability_current", "expense")
- internal_group: "asset" | "liability" | "equity" | "income" | "expense"
- company_id: empresa dueña de la cuenta
- deprecated: TRUE si la cuenta está obsoleta

### Diarios_Contables — Diarios contables (account.journal) — 203 registros
- id, code, name: identificación del diario
- type: "sale" | "purchase" | "bank" | "cash" | "general"
- company_id: empresa propietaria
- default_account_id: cuenta contable por defecto
- active: TRUE si está activo

### Nomina — Líneas detalle de liquidaciones de sueldo (hr.payslip.line) — 46.068 registros
Cada fila es una línea de una liquidación (haber, descuento o totalizador).
- id, name: nombre de la línea (ej: "BONO TRATO", "TOTAL DESCUENTOS", "Salario neto")
- code: código de la regla salarial (ej: "BTRATO", "TDE", "NET")
- slip_id: ID de la liquidación madre (vincula con Remuneraciones.id)
- employee_id: ID del empleado
- contract_id: ID del contrato
- date_from, date_to: período de la liquidación
- amount, total, bl_signed_total: montos en CLP (FLOAT)
- quantity, rate: cantidad y tasa aplicada
- category_id: categoría de la línea (haber=7, descuento=14, neto=5)
- tipo_auxiliar: clasificación auxiliar

### Ubicaciones — Ubicaciones de bodega (stock.location) — 129 registros
- id, name, complete_name: nombre corto y ruta completa (ej: "WH/Bodega Central")
- usage: "internal" | "customer" | "supplier" | "transit" | "view" | "production"
- company_id: empresa propietaria
- active: TRUE si está activa
- warehouse_id: bodega a la que pertenece

### Variantes_del_producto — Variantes de producto (product.product) — 6.074 registros
- id: ID de la variante
- product_tmpl_id: ID del template (vincula con Producto.id para obtener nombre e ingredientes)
- default_code: referencia interna / código de producto
- barcode: código de barras
- active: TRUE si está activo

### Producto — Templates de producto (product.template) — 6.077 registros
Tabla más completa que Variantes_del_producto. Contiene nombre, tipo y datos agroquímicos.
- id: ID del template (diferente al ID de Variantes_del_producto)
- name: nombre en JSON bilingüe {"es_CL": "...", "en_US": "..."} — extraer con JSON_VALUE(name, '$.es_CL')
- default_code: código interno
- type: "product" (almacenable) | "consu" (consumible) | "service"
- active: TRUE si está activo
- list_price: precio de venta
- bl_marca: marca comercial
- ingrediente_activo_1, ingrediente_activo_2, ingrediente_activo_3: ingredientes activos (agroquímicos)
- concentracion_ia1, concentracion_ia2, concentracion_ia3: concentración de cada ingrediente activo (FLOAT)
- x_studio_concentracin: concentración total (custom Studio)
- x_studio_utilidad: utilidad/uso del producto
- sale_ok, purchase_ok: se puede vender / comprar

### Empleados — Ficha de empleados (hr.employee) — 118 registros
- id: ID del empleado (vincula con Remuneraciones.employee_id)
- name: nombre completo
- firstname, last_name, middle_name, mothers_name: nombre desglosado
- formated_vat: RUT formateado (ej: "12.345.678-9")
- job_title: cargo
- department_id: ID departamento
- company_id: ID empresa (importante: 10 empresas distintas del grupo)
- work_email: email laboral
- mobile_phone: teléfono móvil
- contract_id: ID contrato activo
- first_contract_date: fecha primer contrato (antigüedad)
- active: TRUE si activo (todos los 118 registros son activos)
- employee_type: "employee" | "student" etc.
- bl_tramo_asignacion_familiar: tramo AF Chile
- bl_pago_efectivo: TRUE = se paga en efectivo

⚠️ IMPORTANTE: La tabla Empleados tiene 118 registros pero son de MÚLTIPLES empresas del grupo.
Para filtrar solo KONTROLAG: WHERE company_id = 7
Para buscar un empleado: LOWER(name) LIKE '%apellido%' o LOWER(formated_vat) LIKE '%rut%'

QUERY TIPO — empleado por nombre:
  SELECT id, name, formated_vat, job_title, work_email, first_contract_date, company_id
  FROM `ace-scarab-484515-v1.odoo_data.Empleados`
  WHERE LOWER(name) LIKE '%garcia%'

### Pedidos_de_Venta — Pedidos de venta (sale.order) — 951 registros
Datos desde jul 2024. Corresponden a ventas de Agricola Donar Uno SPA principalmente.
- id: ID del pedido
- name: número de pedido (ej: "S00485")
- partner_id: ID cliente (join con Contactos.id)
- state: "sale" (confirmado) | "draft" | "cancel" | "done" — filtrar WHERE state = 'sale'
- date_order: fecha de confirmación (TIMESTAMP)
- amount_untaxed: monto neto (FLOAT)
- amount_tax: IVA (FLOAT)
- amount_total: total con IVA (FLOAT)
- amount_to_invoice: monto pendiente de facturar (FLOAT)
- invoice_status: "invoiced" | "to invoice" | "nothing"
- delivery_status: "pending" | "partial" | "full" | NULL
- company_id: empresa emisora
- analytic_account_id: centro de costo analítico del pedido
- warehouse_id: bodega de despacho

QUERY TIPO — ventas por cliente este año:
  SELECT c.name AS cliente, COUNT(p.id) AS pedidos,
         SUM(p.amount_untaxed) AS neto_total
  FROM `ace-scarab-484515-v1.odoo_data.Pedidos_de_Venta` p
  JOIN `ace-scarab-484515-v1.odoo_data.Contactos` c ON p.partner_id = c.id
  WHERE p.state = 'sale'
    AND EXTRACT(YEAR FROM p.date_order) = 2026
  GROUP BY 1 ORDER BY 3 DESC

QUERY TIPO — pedidos pendientes de facturar:
  SELECT p.name, c.name AS cliente, p.amount_untaxed, p.amount_to_invoice, p.delivery_status
  FROM `ace-scarab-484515-v1.odoo_data.Pedidos_de_Venta` p
  JOIN `ace-scarab-484515-v1.odoo_data.Contactos` c ON p.partner_id = c.id
  WHERE p.state = 'sale' AND p.invoice_status = 'to invoice'
  ORDER BY p.amount_to_invoice DESC

### Lineas_del_Pedido_de_Venta — Líneas de pedidos (sale.order.line) — 3.297 registros
- id: ID de la línea
- order_id: ID pedido padre (join con Pedidos_de_Venta.id)
- order_partner_id: ID cliente del pedido
- product_id: ID variante de producto (join con Variantes_del_producto.id)
- name: descripción de la línea
- product_uom_qty: cantidad pedida (FLOAT)
- qty_delivered: cantidad entregada (FLOAT)
- qty_invoiced: cantidad facturada (FLOAT)
- qty_to_invoice: cantidad pendiente de facturar (FLOAT)
- price_unit: precio unitario (FLOAT)
- price_subtotal: subtotal neto (FLOAT)
- price_total: total con impuesto (FLOAT)
- invoice_status: "invoiced" | "to invoice" | "nothing"
- state: estado del pedido padre
- analytic_distribution: JSON de distribución analítica (mismo formato que Reporte_Analitico)
- is_service: TRUE = es un servicio
- display_type: "line_note" = línea de texto sin precio (ignorar para sumas)
- company_id: empresa

⚠️ Filtrar SIEMPRE display_type IS NULL o display_type != 'line_note' para excluir líneas de texto.

### Despachos — Guías de despacho (stock.picking) — 6.469 registros
Corresponden a Agricola Donar Uno SPA. Cada fila es una transferencia completa (no la línea de producto).
- id: ID del despacho
- name: referencia interna (ej: "D1/OUT/00421")
- state: "done" | "assigned" | "cancel" — filtrar WHERE state = 'done' para completados
- partner_id: ID cliente receptor (join con Contactos.id)
- sale_id: ID pedido de venta origen (join con Pedidos_de_Venta.id) — puede ser NULL
- date_done: fecha de completado (TIMESTAMP)
- date, scheduled_date: fecha programada (TIMESTAMP)
- l10n_latam_document_number: N° guía DTE (ej: "000421")
- l10n_cl_dte_status: estado SII: "accepted" | "rejected" | "objected"
- l10n_cl_delivery_guide_reason: razón de traslado (código SII)
- l10n_cl_sii_send_ident: RUT receptor
- origin: referencia de origen (ej: "S00485" vincula con pedido de venta)
- location_id, location_dest_id: ubicación origen/destino (join con Ubicaciones.id)
- note: notas en HTML (incluye datos de chofer y patente)

QUERY TIPO — despachos de un cliente:
  SELECT d.name, d.l10n_latam_document_number, DATE(d.date_done) AS fecha,
         d.l10n_cl_dte_status, d.origin
  FROM `ace-scarab-484515-v1.odoo_data.Despachos` d
  JOIN `ace-scarab-484515-v1.odoo_data.Contactos` c ON d.partner_id = c.id
  WHERE d.state = 'done'
    AND LOWER(c.name) LIKE '%donar%'
  ORDER BY d.date_done DESC

### Movimientos_de_Stock — Movimientos de stock nivel operación (stock.move) — 19.777 registros
⚠️ DIFERENCIA IMPORTANTE: Esta tabla es stock.move (una operación completa).
Las Ordenes_de_aplicacion son stock.move.line (línea individual dentro de la operación).
- id: ID del movimiento
- name, reference: referencia (ej: "D1/MO/00112")
- product_id: ID producto (join con Variantes_del_producto.id)
- product_uom_qty: cantidad demandada (FLOAT)
- quantity: cantidad ejecutada (FLOAT)
- location_id, location_dest_id: ubicación origen/destino
- picking_id: ID del despacho padre (join con Despachos.id) — NULL si es manufactura
- state: "done" | "assigned" | "cancel"
- date: fecha del movimiento (TIMESTAMP)
- origin: referencia de origen (pedido, orden de manufactura)
- sale_line_id: línea de pedido de venta origen
- raw_material_production_id: ID orden de manufactura (si es consumo de MP)
- is_done: TRUE si completado
- picked: TRUE si retirado físicamente
- scrapped: TRUE si es merma
- x_studio_dosis_x_hectarea, x_studio_dosis_x_100l: dosis agronómicas (FLOAT)
- x_studio_stock_isla, x_studio_stock_zuig, x_stock_bodega_lanas: stock por predio (FLOAT)

CRUCES CLAVE ENTRE TABLAS ODOO:
- Nombre de producto en movimiento: Movimientos_de_Stock.product_id ↔ Variantes_del_producto.id → Producto.id
- Detalle liquidación empleado: Remuneraciones.id ↔ Nomina.slip_id
- Empleado → liquidaciones: Empleados.id ↔ Remuneraciones.employee_id
- Nombre cliente/proveedor: Cuentas_por_cobrar.partner_id ↔ Contactos.id
- Cuenta contable: Reporte_Analitico.account_id ↔ Cuentas_Contables.id
- Ubicación bodega: Movimientos_de_Stock.location_id ↔ Ubicaciones.id
- Despacho → pedido origen: Despachos.sale_id ↔ Pedidos_de_Venta.id
- Pedido → líneas: Pedidos_de_Venta.id ↔ Lineas_del_Pedido_de_Venta.order_id
- Movimientos de un despacho: Despachos.id ↔ Movimientos_de_Stock.picking_id
- Nombre template de producto: Variantes_del_producto.product_tmpl_id ↔ Producto.id

══════════════════════════════════════════════
FUENTE 2 — PostgreSQL (esquema appsheet / public)
══════════════════════════════════════════════
Usa la tool `query_postgres`. SQL estándar PostgreSQL.
Las fechas en tarjas_pagos y otros son tipo TEXT — usar CAST(fecha AS DATE) o fecha::date.

### appsheet.tarjas_pagos — Pagos diarios a trabajadores de campo
- id_Resumen, id_supervisor, fecha (TEXT → cast a date)
- nombre_campo: campo donde se trabajó (ej: "Isla de Maipo", "Zuñiga", "Talagante")
- cuartel_cc: código del centro de costo (INTEGER)
- labor: nombre de la labor realizada
- contratista: empresa contratista
- trabajador: nombre del trabajador
- rut_trabajador: RUT del trabajador
- tipo_pago: "trato" | "Al dia" | "Al día"
- valor_jornada, valor_trato, base_trato (INTEGER)
- rendimiento, horas_extras, horas_trabajadas (TEXT)
- total_jornada, total_trato, total_trabajado (INTEGER)
- contratista_jornada, contratista_trato, total_contratista (INTEGER)
- total_pagar: total a pagar a la empresa (INTEGER)
- estado: estado del pago

### appsheet.tarjas_reporte — Vista resumida de tarjas por día/campo/labor
- contratista, nombre_campo, fecha (DATE)
- total_a_trato, total_al_dia, total_a_pagar (BIGINT)
- tipo_pago, CC (INTEGER), "Nombre Labor" (TEXT)
- jornadas (BIGINT), total_labor (BIGINT)

### appsheet.tarjas_contratistas — Contratistas registrados
- id_contratista, id_campo (INTEGER), nombre, tipo
- valor_hora_extra (INTEGER), valor_jornada (INTEGER)
- porcentaje_jornada, porcentaje_trato (TEXT)

### appsheet.tarjas_personal — Nómina de trabajadores por contratista
- id_personal, id_contratista, rut, nombre
- licencia_clase_d, certificado_sag, id_sistema

### appsheet.tarjas_campo — Campos/predios
- id_campo (INTEGER), nombre

### appsheet.tarjas_cc — Centros de costo (cuarteles)
- id_cc, cultivo, id_campo (INTEGER), valor_odoo (JSONB)

### appsheet.tarjas_labor / tarjas_labores — Catálogo de labores
- id_labor, nombre, tipo

### appsheet.despacho_ingreso — Ingresos de fruta (cosecha)
- id_conteo, cliente, kg_brutos, total_bins, total_kg_liquidos
- total_kg_liquidos_bins, chofer, rut, patente, nombre_chofer
- cantidad_de_pallet, producto, cajas, unidades_caja, total_unidades
- fecha_cosecha (TEXT), calidad, variedad, plantas, productor, destino
- mallas, fecha_limpia (TEXT)

### appsheet.despacho_guia — Guías de despacho
- id_guia, contacto, producto, cantidad, chofer
- patente_camion, fecha (TEXT), descripcion

### appsheet.despacho_venta — Órdenes de venta despacho
- id_venta, cliente, producto, descripcion
- cantidad, centro_costo, fecha (TEXT)

### appsheet.despacho_cliente — Clientes de despacho
- cliente_id, cliente, nombre_odoo

### appsheet.despacho_productos — Productos de despacho con CC
- id, producto, calidad, cc

### public.status_system — ⭐ RESUMEN DIARIO POR CAMPO (USAR PARA TEMPERATURA Y RIEGO)
Esta es la tabla principal para preguntas de temperatura ambiente y riego por campo y fecha.
- date (DATE): fecha del día
- field (TEXT): campo, valores: "ISLA DE MAIPO" | "ZUÑIGA"
- temp_avg (NUMERIC): temperatura promedio del día en °C
- temp_min (NUMERIC): temperatura mínima del día en °C
- temp_max (NUMERIC): temperatura máxima del día en °C
- irrigation_mm (NUMERIC): riego del día en mm
- status (TEXT): estado del sistema ese día
- notes (TEXT): observaciones

EJEMPLO — temperatura de ayer por campo:
  SELECT field, date, temp_avg, temp_min, temp_max
  FROM public.status_system
  WHERE date = CURRENT_DATE - 1
  ORDER BY field;

### public.sensor_inventory — Inventario de sensores
- id, source (ej: "ubibot", "wiseconn"), field, zone, orchard
- sensor_id, sensor_name, unit, last_value (NUMERIC)
- last_seen (DATE), status ("ok" | "offline"), updated_at

### public.ubi_sensor_pivot — Lecturas horarias de sensores ambiente (temperatura, humedad, luz)
USAR cuando se piden lecturas horarias o de sensores específicos, NO para resumen diario.
- channel_id, field, irrigation_sector, orchard, crop_type
- date (DATE), hour (TIME)
- temperature (NUMERIC): temperatura ambiente en °C — IGNORAR filas donde temperature = 0
- humidity (NUMERIC): humedad relativa %
- light, voltage, wind_speed (NUMERIC)
- temperatura_suelo_25cm, temperatura_suelo_50cm (NUMERIC): temperatura de suelo
- humedad_suelo_25cm, humedad_suelo_50cm (NUMERIC): humedad de suelo

### public.ubi_ambient_temperature — Resumen diario de temperatura por sensor
- channel_id, channel_name, field, irrigation_sector, orchard, crop_type
- date (DATE)
- temp_avg, temp_min, temp_max (NUMERIC)

### public.ubi_soil_sensors — Sensores de suelo (temperatura y humedad por hora)
ADVERTENCIA: muchos valores son 0.0 cuando el sensor no tiene lectura — filtrar con > 0.
- field, irrigation_sector, orchard, crop_type
- date (DATE), hour (TIME)
- temp_25cm, temp_50cm (NUMERIC): temperatura suelo en °C
- hum_10cm .. hum_50cm (NUMERIC): humedad suelo %

### public.wc_farms_irrigation — Riegos ejecutados (WiseConn)
- date (DATE), hour (TIME), inittime, endtime, delta_time
- status, irrigationtype, zone_id, farm_id
- volume_m3, precipitation_mm, theoreticalflow_m3_h (FLOAT)

### public.field_sectors — Sectores/cuarteles con sus sensores
- id, field, farm_id, irrigation_sector, wc_zone_id
- orchard, crop_type, ubibot_channel_ids (ARRAY)

### public.ubi_chill_hours — ⭐ Horas frío acumuladas por huerto (CLAVE PARA AGRONOMÍA)
Registros horarios de temperatura y acumulación de frío por modelo. Fundamental para estimar
fecha de brotación y necesidades de frío de cada variedad de cerezo.
- channel_id, channel_name, field_sector_id
- field: "ISLA DE MAIPO" | "ZUÑIGA"
- irrigation_sector, orchard, crop_type
- date (DATE), hour (TIME)
- temperature (NUMERIC): temperatura en °C en esa hora
- hf_value (INTEGER): horas frío modelo HF en esa hora (1 = sí cuenta, 0 = no)
- hf_accumulated (INTEGER): horas frío HF acumuladas en la temporada
- utah_value (NUMERIC): unidades frío modelo Utah en esa hora (puede ser negativo)
- utah_accumulated (NUMERIC): unidades frío Utah acumuladas en la temporada
- gda_value (NUMERIC): valor GDA (grados-día de acumulación) en esa hora
- gda_accumulated (NUMERIC): GDA acumulado en la temporada
- gda_season (TEXT): temporada para GDA (ej: "2025-2026")
- season (TEXT): temporada para HF/Utah (ej: "2026-2027")
- rn (INTEGER): número de hora en la temporada

MODELOS DE FRÍO:
  - HF (Horas Frío): cuenta horas con temperatura entre 0°C y 7.2°C. Simple y usado en Chile.
  - Utah: modelo más preciso, pondera positivo y negativo según rango de temperatura.
    Valores óptimos: 0-7.2°C = +1, 7.3-12.4°C = +0.5, 12.5-15.9°C = 0, 16-18°C = -0.5, >18°C = -1
  - GDA (Grados Día de Acumulación): calor acumulado desde inicio de temporada,
    indica avance fenológico post-reposo.

REQUERIMIENTOS TÍPICOS CEREZOS (referencia):
  - Bing/Lapins: 1.000-1.200 HF
  - Santina: 800-1.000 HF
  - Regina: 1.200-1.400 HF

EJEMPLO — horas frío acumuladas hoy por huerto:
  SELECT DISTINCT ON (field, orchard)
    field, orchard, date, hf_accumulated, utah_accumulated, gda_accumulated, season
  FROM public.ubi_chill_hours
  ORDER BY field, orchard, date DESC, hour DESC;

### public.ubi_chill_portions — Porciones de frío (Dynamic Model) por huerto
El Dynamic Model es el más preciso para predecir brotación. Muy usado en cerezos.
- channel_id, channel_name, field_sector_id
- field, irrigation_sector, orchard, crop_type
- date (DATE), hour (TIME), temperature (NUMERIC)
- dm_season (TEXT): temporada (ej: "2026")
- dm_state (NUMERIC): estado interno del modelo en esa hora
- dm_value (NUMERIC): porciones de frío generadas en esa hora
- dm_accumulated (NUMERIC): porciones de frío acumuladas en la temporada

REQUERIMIENTOS TÍPICOS (porciones Dynamic Model):
  - Santina: 40-50 porciones
  - Bing: 55-65 porciones
  - Regina: 65-80 porciones

### public.wc_ema — Estación meteorológica WiseConn (datos diarios por campo)
- date (DATE), farm_id, field ("ISLA DE MAIPO" | "ZUÑIGA")
- temperatura_c (NUMERIC): temperatura promedio diaria °C
- humedad_relativa_pct (NUMERIC): humedad relativa %
- presion_atmosferica_pa (NUMERIC): presión atmosférica en Pascales
- radiacion_solar_wm2 (NUMERIC): radiación solar W/m²
- pluviometria_mm (NUMERIC): lluvia en mm
- velocidad_viento_kmh (NUMERIC): velocidad viento km/h
- direccion_viento_deg (NUMERIC): dirección viento en grados

══════════════════════════════════════════════
CRUCES DE NEGOCIO
══════════════════════════════════════════════
ODOO INTERNO:
- Detalle liquidación empleado: Remuneraciones.id ↔ Nomina.slip_id
- Empleado → liquidaciones: Empleados.id ↔ Remuneraciones.employee_id
- Nombre en línea contable: Reporte_Analitico.partner_id ↔ Contactos.id
- Nombre cuenta contable: Reporte_Analitico.account_id ↔ Cuentas_Contables.id
- Nombre producto en orden aplicación: Ordenes_de_aplicacion.product_id ↔ Variantes_del_producto.id
- Nombre template/ingredientes: Variantes_del_producto.product_tmpl_id ↔ Producto.id
- Ubicación bodega: Movimientos_de_Stock.location_id ↔ Ubicaciones.id
- Despacho → pedido origen: Despachos.sale_id ↔ Pedidos_de_Venta.id
- Despacho → movimientos: Despachos.id ↔ Movimientos_de_Stock.picking_id
- Pedido → líneas: Pedidos_de_Venta.id ↔ Lineas_del_Pedido_de_Venta.order_id
- RUT de un cliente/proveedor: Cuentas_por_cobrar.partner_id → Contactos.vat
- RUT de un empleado: Empleados.formated_vat (directo)

ODOO ↔ APPSHEET:
- Costo campo vs facturación: tarjas_pagos.nombre_campo ↔ Cuentas_por_cobrar.invoice_partner_display_name
- Costo laboral mensual: tarjas_pagos.fecha::date ↔ Remuneraciones.date_from/date_to
- CC AppSheet → CC Odoo: tarjas_pagos.cuartel_cc → tabla de mapeo → clave en analytic_distribution JSON
  ⚠️ NO existe columna analytic_account_id. El CC Odoo va dentro del JSON de analytic_distribution.

APPSHEET INTERNO:
- Sensores por campo: sensor_inventory.field ↔ tarjas_campo.nombre

CAMPOS Y EMPRESA FACTURADA EN ODOO:
  - "Isla de Maipo" → Agricola Donar Dos SpA
  - "Zuñiga" → Agricola Donar Dos SpA

══════════════════════════════════════════════
CONTEXTO AGRONÓMICO — CEREZOS
══════════════════════════════════════════════
Los campos cultivan principalmente CEREZOS. El frío invernal es crítico para romper
la dormancia y garantizar una brotación uniforme en primavera.

PREGUNTAS FRECUENTES Y CÓMO RESPONDERLAS:

1. "¿Cuántas horas frío llevamos?" o "¿Cómo vamos de frío?"
   → Consultar ubi_chill_hours: último hf_accumulated, utah_accumulated por field/orchard
   → Comparar con requerimientos de la variedad si se menciona
   → Indicar si el acumulado es suficiente, bajo o alto para la fecha

2. "¿Hubo helada?" o "¿Cuándo heló?"
   → Consultar ubi_chill_hours o status_system: temp_min < 0°C
   → Una helada agronómica es temp < 0°C. Helada dañina en flores/frutos: < -2°C

3. "¿Cómo están las porciones de frío?" (Dynamic Model)
   → Consultar ubi_chill_portions: último dm_accumulated por field/orchard

4. "¿Cuánto llovió?" o "¿Hubo lluvia?"
   → Consultar wc_ema: pluviometria_mm por campo y fecha

5. "¿Cómo está el clima?" o preguntas meteorológicas generales
   → Consultar wc_ema para datos diarios completos (temp, humedad, viento, radiación)
   → Complementar con status_system para temp_min/max

GLOSARIO AGRONÓMICO:
- Dormancia: período de reposo invernal del árbol que requiere frío para completarse
- Brotación: inicio del crecimiento vegetativo en primavera tras completar el frío
- HF: Horas Frío — método más simple, cuenta horas entre 0°C y 7.2°C
- Utah: modelo ponderado más preciso que HF para zonas con inviernos irregulares
- Dynamic Model / Porciones: modelo más moderno y preciso, especialmente para climas mediterráneos
- GDA: Grados Día Acumulados — mide calor acumulado desde inicio de temporada
- Helada: temperatura menor a 0°C a nivel del suelo
- ETo: evapotranspiración de referencia (no disponible actualmente en la DB)

══════════════════════════════════════════════
CATÁLOGO DE REPORTES — LOOKER STUDIO
══════════════════════════════════════════════
REGLA ESTRICTA: Si el usuario menciona "looker studio", "looker", "reporte visual" o "dashboard visual":
→ Llama a query_postgres con la siguiente query para obtener los reportes actualizados:
  SELECT name, url, category, description FROM public.looker_reports ORDER BY category, name
→ Presenta los resultados agrupados por categoría con el formato: **[nombre](url)**
→ Si el usuario menciona una categoría o tema específico, filtra con WHERE category ILIKE '%término%' OR name ILIKE '%término%'
→ NUNCA inventes URLs — solo usa las que devuelve la query."""

# ── Tools ─────────────────────────────────────────────────────────────────────

_TOOLS = [
    types.FunctionDeclaration(
        name="query_bigquery",
        description=(
            "Ejecuta SQL en BigQuery y devuelve resultados. "
            "Usar path completo: `ace-scarab-484515-v1.odoo_data.TABLA`. "
            "Tablas disponibles: Cuentas_por_cobrar, Remuneraciones, Reporte_Analitico, "
            "Ordenes_de_aplicacion, Contactos, Cuentas_Contables, Diarios_Contables, "
            "Nomina, Ubicaciones, Variantes_del_producto, Producto, Empleados, "
            "Pedidos_de_Venta, Lineas_del_Pedido_de_Venta, Despachos, Movimientos_de_Stock. "
            "Para Reporte_Analitico y Lineas_del_Pedido_de_Venta: usar JSON_EXTRACT_SCALAR(analytic_distribution, '$.ID') "
            "para extraer porcentaje de un CC específico y ponderarlo sobre balance. "
            "Para nombre de producto en Producto: JSON_VALUE(name, '$.es_CL')."
        ),
        parameters=types.Schema(
            type=types.Type.OBJECT,
            properties={
                "sql": types.Schema(type=types.Type.STRING, description="Query SQL BigQuery"),
                "max_rows": types.Schema(type=types.Type.INTEGER, description="Máximo filas (default 100)"),
            },
            required=["sql"],
        ),
    ),
    types.FunctionDeclaration(
        name="query_postgres",
        description=(
            "Ejecuta SQL en PostgreSQL (appsheet y public schemas) y devuelve resultados. "
            "Schemas disponibles: appsheet (tarjas, despacho) y public (sensores, riego). "
            "Tablas: appsheet.tarjas_pagos, appsheet.tarjas_reporte, appsheet.tarjas_contratistas, "
            "appsheet.tarjas_personal, appsheet.tarjas_campo, appsheet.tarjas_cc, "
            "appsheet.tarjas_labor, appsheet.despacho_ingreso, appsheet.despacho_guia, "
            "appsheet.despacho_venta, appsheet.despacho_cliente, appsheet.despacho_productos, "
            "public.sensor_inventory, public.ubi_soil_sensors, public.wc_farms_irrigation, "
            "public.field_sectors, "
            "public.looker_reports (id, name, description, url, category) — catálogo de reportes Looker Studio."
        ),
        parameters=types.Schema(
            type=types.Type.OBJECT,
            properties={
                "sql": types.Schema(type=types.Type.STRING, description="Query SQL PostgreSQL estándar"),
                "max_rows": types.Schema(type=types.Type.INTEGER, description="Máximo filas (default 200)"),
            },
            required=["sql"],
        ),
    ),
    types.FunctionDeclaration(
        name="list_tables",
        description="Lista todas las tablas disponibles en BigQuery (Odoo) con cantidad de filas.",
        parameters=types.Schema(type=types.Type.OBJECT, properties={}),
    ),
    types.FunctionDeclaration(
        name="describe_table",
        description="Devuelve esquema y muestra de una tabla BigQuery.",
        parameters=types.Schema(
            type=types.Type.OBJECT,
            properties={
                "table_id": types.Schema(
                    type=types.Type.STRING,
                    description="Nombre de la tabla BigQuery, ej: 'Cuentas_por_cobrar'",
                )
            },
            required=["table_id"],
        ),
    ),
    types.FunctionDeclaration(
        name="render_chart",
        description=(
            "Renderiza un gráfico visual inline en el chat a partir de datos ya consultados. "
            "Llamar DESPUÉS de query_bigquery o query_postgres cuando el usuario pide un gráfico "
            "o cuando los datos son claramente comparativos/temporales y un gráfico los muestra mejor. "
            "Tipos disponibles: 'bar' (barras), 'line' (línea temporal), 'pie' (torta/porcentajes), 'doughnut' (dona). "
            "Elegir el tipo más apropiado: line para series temporales, bar para comparaciones, pie/doughnut para proporciones. "
            "labels: array de etiquetas del eje X o categorías. "
            "values: array de valores numéricos correspondientes. "
            "dataset_label: nombre del dataset (ej: 'Monto CLP', 'Temperatura °C'). "
            "title: título descriptivo del gráfico."
        ),
        parameters=types.Schema(
            type=types.Type.OBJECT,
            properties={
                "chart_type": types.Schema(type=types.Type.STRING, description="'bar' | 'line' | 'pie' | 'doughnut'"),
                "title": types.Schema(type=types.Type.STRING, description="Título del gráfico"),
                "labels": types.Schema(type=types.Type.STRING, description="Array JSON de etiquetas, ej: '[\"Ene\",\"Feb\"]'"),
                "values": types.Schema(type=types.Type.STRING, description="Array JSON de valores numéricos, ej: '[1200,3400]'"),
                "dataset_label": types.Schema(type=types.Type.STRING, description="Nombre del dataset"),
            },
            required=["chart_type", "title", "labels", "values", "dataset_label"],
        ),
    ),
]


def _get_embedding(text: str) -> Optional[list]:
    """Return a 768-dim embedding vector for text using Gemini text-embedding-004."""
    try:
        result = _client.models.embed_content(
            model=EMBED_MODEL,
            contents=text,
        )
        return result.embeddings[0].values
    except Exception as e:
        _log.warning("embedding error: %s", e)
        return None


_WRITE_RE = re.compile(
    r"^\s*(insert|update|delete|drop|truncate|alter|create|grant|revoke)\b",
    re.IGNORECASE,
)

_PK_CANDIDATES = ["id", "name", "id_Resumen", "id_resumen", "id_guia", "id_venta",
                  "id_conteo", "id_contratista", "id_personal", "id_labor",
                  "channel_id", "sensor_id", "zone_id"]


def _extract_table_name(sql: str) -> str:
    """Extract the main table name from a SQL query."""
    # Include hyphens to handle BQ paths like `ace-scarab-484515-v1.odoo_data.Tabla`
    table_match = re.search(r'\bFROM\s+([`"\w\.\-]+)', sql, re.IGNORECASE)
    if not table_match:
        return "tabla desconocida"
    table = table_match.group(1).strip('`"')
    # Remove BQ project prefix for display (keep only last segment)
    if "." in table:
        parts = table.split(".")
        table = parts[-1] if len(parts) >= 3 else table
    return table


def _build_trace(sql: str, rows: list[dict], system: str) -> dict:
    """Build a structured trace object for data provenance."""
    table = _extract_table_name(sql)
    n = len(rows)
    preview = []
    if rows:
        # Take first 3 rows, convert all values to strings for safe serialization
        for r in rows[:3]:
            preview.append({k: str(v) for k, v in r.items()})
    return {
        "table": table,
        "system": system,
        "row_count": n,
        "sql": sql.strip(),
        "preview": preview,
    }


def _build_source_badge(trace: dict) -> str:
    """Generate a text source attribution line from a trace object."""
    if not trace or trace["row_count"] == 0:
        return ""
    n = trace["row_count"]
    return (
        f"\n\n📊 Fuente: {trace['table']} · {trace['system']} · "
        f"{n} registro{'s' if n != 1 else ''}"
    )


def _call_tool(name: str, args: dict) -> str:
    try:
        if name == "query_bigquery":
            sql = args["sql"]
            if _WRITE_RE.match(sql):
                return json.dumps({"error": "Solo se permiten consultas SELECT"})
            rows = bq.run_query(sql, int(args.get("max_rows", 100)))
            trace = _build_trace(sql, rows, "BigQuery (Odoo)")
            badge = _build_source_badge(trace)
            result = json.loads(json.dumps(rows, ensure_ascii=False, default=str))
            return json.dumps({"rows": result, "__source_badge__": badge, "__trace__": trace}, ensure_ascii=False)
        if name == "query_postgres":
            sql = args["sql"]
            rows = pg.run_pg_query(sql, int(args.get("max_rows", 200)))
            trace = _build_trace(sql, rows, "PostgreSQL (AppSheet)")
            badge = _build_source_badge(trace)
            result = json.loads(json.dumps(rows, ensure_ascii=False, default=str))
            return json.dumps({"rows": result, "__source_badge__": badge, "__trace__": trace}, ensure_ascii=False)
        if name == "list_tables":
            return json.dumps(bq.list_tables(), ensure_ascii=False)
        if name == "describe_table":
            return json.dumps(bq.describe_table(args["table_id"]), ensure_ascii=False, default=str)
        if name == "render_chart":
            try:
                labels = json.loads(args["labels"])
                values = json.loads(args["values"])
            except Exception:
                return json.dumps({"error": "labels y values deben ser arrays JSON válidos"})
            chart_payload = {
                "chart_type": args.get("chart_type", "bar"),
                "title": args.get("title", ""),
                "dataset_label": args.get("dataset_label", ""),
                "labels": [str(l) for l in labels],
                "values": [float(v) if v is not None else 0 for v in values],
            }
            return json.dumps({"__chart__": chart_payload})
        return json.dumps({"error": f"tool desconocida: {name}"})
    except Exception as e:
        return json.dumps({"error": str(e)})


# ── History helpers ───────────────────────────────────────────────────────────

def _ensure_traces_column() -> None:
    """Add traces and charts JSONB columns to chat_history if they don't exist yet."""
    conn = pg._get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("ALTER TABLE public.chat_history ADD COLUMN IF NOT EXISTS traces JSONB")
            cur.execute("ALTER TABLE public.chat_history ADD COLUMN IF NOT EXISTS charts JSONB")
        conn.commit()
    except Exception as e:
        _log.warning("Could not add columns to chat_history: %s", e)
        conn.rollback()
    finally:
        conn.close()


def _save_messages(username: str, user_msg: str, model_reply: str, traces: list = None, charts: list = None) -> None:
    conn = pg._get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO public.chat_history (username, role, message) VALUES (%s, %s, %s)",
                (username, "user", user_msg),
            )
            cur.execute(
                "INSERT INTO public.chat_history (username, role, message, traces, charts) VALUES (%s, %s, %s, %s, %s)",
                (username, "model", model_reply,
                 json.dumps(traces, default=str) if traces else None,
                 json.dumps(charts, default=str) if charts else None),
            )
        conn.commit()
    except Exception as e:
        _log.error("_save_messages failed user=%s error=%s", username, e)
        conn.rollback()
    finally:
        conn.close()


def _load_history(username: str, limit: int = 40) -> list[dict]:
    conn = pg._get_conn()
    try:
        with conn.cursor(cursor_factory=__import__("psycopg2").extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, role, message, traces, charts, created_at
                FROM public.chat_history
                WHERE username = %s
                ORDER BY id DESC
                LIMIT %s
            """, (username, limit))
            rows = cur.fetchall()
            result = []
            for r in reversed(rows):
                entry = {"role": r["role"], "message": r["message"], "created_at": str(r["created_at"])}
                if r["traces"]:
                    entry["traces"] = r["traces"] if isinstance(r["traces"], list) else json.loads(r["traces"])
                if r["charts"]:
                    entry["charts"] = r["charts"] if isinstance(r["charts"], list) else json.loads(r["charts"])
                result.append(entry)
            return result
    finally:
        conn.close()


# ── Init ──────────────────────────────────────────────────────────────────────

def _ensure_prompt_table() -> None:
    conn = pg._get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS public.chat_prompts (
                    id         SERIAL PRIMARY KEY,
                    prompt     TEXT NOT NULL,
                    is_active  BOOLEAN NOT NULL DEFAULT FALSE,
                    saved_at   TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """)
        conn.commit()
    finally:
        conn.close()


def _load_system_prompt() -> str:
    """Load active prompt from DB; seed with hardcoded default if table is empty."""
    global _SYSTEM
    try:
        _ensure_prompt_table()
        conn = pg._get_conn()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT prompt FROM public.chat_prompts WHERE is_active = TRUE ORDER BY id DESC LIMIT 1"
                )
                row = cur.fetchone()
        finally:
            conn.close()
        if row and row[0].strip():
            _SYSTEM = row[0]
        else:
            # Seed DB with the hardcoded default
            _save_system_prompt(_SYSTEM)
    except Exception as e:
        _log.warning("Could not load prompt from DB, using hardcoded default: %s", e)
    return _SYSTEM


def _save_system_prompt(new_prompt: str) -> int:
    """Archive current active prompt as inactive, insert new one as active. Returns new row id."""
    global _SYSTEM
    conn = pg._get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE public.chat_prompts SET is_active = FALSE WHERE is_active = TRUE")
            cur.execute(
                "INSERT INTO public.chat_prompts (prompt, is_active) VALUES (%s, TRUE) RETURNING id",
                (new_prompt,),
            )
            new_id = cur.fetchone()[0]
        conn.commit()
    finally:
        conn.close()
    _SYSTEM = new_prompt
    return new_id


def init(templates: Jinja2Templates):
    global _templates, _client
    _templates = templates
    _load_system_prompt()
    _ensure_traces_column()
    cache._ensure_charts_column()

    # Support credentials from env var (base64 JSON) or local file path
    key_b64 = os.environ.get("BIGQUERY_KEY_B64")
    if key_b64:
        import base64, tempfile
        key_json = base64.b64decode(key_b64)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
        tmp.write(key_json)
        tmp.close()
        key_file = tmp.name
    else:
        key_file = KEY_PATH

    credentials = service_account.Credentials.from_service_account_file(
        key_file,
        scopes=["https://www.googleapis.com/auth/cloud-platform"],
    )
    _client = genai.Client(
        vertexai=True,
        project=GCP_PROJECT,
        location=GCP_LOCATION,
        credentials=credentials,
    )


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
async def chat_page(request: Request):
    return _templates.TemplateResponse(request, "chat.html")


@router.get("/history")
async def chat_history(request: Request):
    from auth import get_current_user
    username = get_current_user(request) or "anonymous"
    rows = _load_history(username, limit=60)
    return JSONResponse({"history": rows})


@router.delete("/history")
async def clear_history(request: Request):
    from auth import get_current_user
    username = get_current_user(request) or "anonymous"
    conn = pg._get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.chat_history WHERE username = %s", (username,))
        conn.commit()
    finally:
        conn.close()
    return JSONResponse({"ok": True})


@router.get("/prompt-editor", response_class=HTMLResponse)
async def prompt_editor_page(request: Request, _admin=Depends(require_admin)):
    return _templates.TemplateResponse(request, "prompt_editor.html")


@router.get("/prompt")
async def get_prompt(_admin=Depends(require_admin)):
    return JSONResponse({"prompt": _SYSTEM})


@router.post("/prompt")
async def update_prompt(request: Request, _admin=Depends(require_admin)):
    from fastapi import HTTPException
    body = await request.json()
    new_prompt = body.get("prompt", "").strip()
    if not new_prompt or len(new_prompt) < 100:
        raise HTTPException(status_code=400, detail="Prompt demasiado corto o vacío")
    version = _save_system_prompt(new_prompt)
    return JSONResponse({"ok": True, "version": version})


@router.get("/prompt/versions")
async def list_prompt_versions():
    conn = pg._get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, saved_at, LENGTH(prompt) AS chars
                FROM public.chat_prompts
                ORDER BY id DESC
                LIMIT 30
            """)
            rows = cur.fetchall()
    finally:
        conn.close()
    versions = [
        {
            "name": str(r[0]),
            "saved_at": r[1].strftime("%d/%m/%Y %H:%M") if r[1] else "—",
            "size": r[2] or 0,
        }
        for r in rows
    ]
    return JSONResponse({"versions": versions})


@router.get("/prompt/versions/{version_id}")
async def get_prompt_version(version_id: str):
    from fastapi import HTTPException
    if not version_id.isdigit():
        raise HTTPException(status_code=400, detail="ID de versión inválido")
    conn = pg._get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT prompt, id, is_active, saved_at FROM public.chat_prompts WHERE id = %s",
                (int(version_id),),
            )
            row = cur.fetchone()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Versión no encontrada")
    return JSONResponse({
        "prompt": row[0],
        "name": str(row[1]),
        "is_active": row[2],
        "saved_at": row[3].strftime("%d/%m/%Y %H:%M") if row[3] else "—",
    })


_ROLE_RESTRICTION_USER = """
---
RESTRICCIÓN DE ACCESO ACTIVA — ROL: USUARIO
El usuario autenticado tiene rol "usuario" (acceso básico).
NO debes responder consultas ni extraer datos de:
- Remuneraciones, sueldos, liquidaciones, nómina, RRHH
- Módulo de Ventas: facturación a clientes, márgenes, precios de venta, cuentas por cobrar
Si el usuario pregunta sobre estos temas, responde exactamente:
"No tengo permisos suficientes para acceder a esta información. Contacta a un administrador."
---
"""

@router.post("/ask")
async def chat_ask(request: Request):
    from auth import get_current_user, get_user_role
    from fastapi import HTTPException
    username = get_current_user(request) or "anonymous"
    user_role = get_user_role(username)

    body = await request.json()
    messages = body.get("messages", [])

    if not messages:
        raise HTTPException(status_code=400, detail="No messages provided")

    user_question = messages[-1].get("parts", "")
    if not isinstance(user_question, str) or len(user_question) > 4000:
        raise HTTPException(status_code=400, detail="Mensaje inválido o demasiado largo")

    _log.info("chat/ask user=%s question_len=%d", username, len(user_question))

    # ── Semantic cache lookup ─────────────────────────────────────────
    embedding = _get_embedding(user_question)
    cache_hit = False
    if embedding:
        cached = cache.get(user_question, embedding)
        if cached:
            cache_hit = True
            cached_reply, cached_traces, cached_charts = cached
            # Clean stale <br> from old cached replies
            cached_reply = re.sub(r"^(\s*<br\s*/?>)+\s*", "", cached_reply, flags=re.IGNORECASE).strip()
            _save_messages(username, user_question, cached_reply, cached_traces, cached_charts or None)
            _log.info("cache hit user=%s", username)
            return JSONResponse({
                "reply": cached_reply,
                "traces": cached_traces,
                "charts": cached_charts,
                "from_cache": True,
            })
    # Lazy cleanup — 1% of requests to avoid dedicated cron
    import random
    if random.random() < 0.01:
        cache.cleanup()

    history = []
    for m in messages[:-1]:
        role = "user" if m["role"] == "user" else "model"
        history.append(types.Content(role=role, parts=[types.Part(text=m["parts"])]))

    effective_system = _SYSTEM + (_ROLE_RESTRICTION_USER if user_role != "admin" else "")
    chat = _client.chats.create(
        model=MODEL,
        config=types.GenerateContentConfig(
            system_instruction=effective_system,
            tools=[types.Tool(function_declarations=_TOOLS)],
        ),
        history=history,
    )

    response = chat.send_message(messages[-1]["parts"])

    # Agentic loop — Gemini may call multiple tools before answering
    collected_badges: list[str] = []
    collected_traces: list[dict] = []
    collected_charts: list[dict] = []
    for _ in range(8):
        fn_calls = [p for p in response.candidates[0].content.parts if p.function_call]
        if not fn_calls:
            break

        tool_parts = []
        for part in fn_calls:
            fn = part.function_call
            result = _call_tool(fn.name, dict(fn.args))
            _log.info("tool_call name=%s result_len=%d", fn.name, len(result))
            # Collect source badges, traces and charts
            try:
                parsed = json.loads(result)
                badge = parsed.get("__source_badge__", "")
                trace = parsed.get("__trace__")
                chart = parsed.get("__chart__")
                _log.info("tool_result has_trace=%s has_badge=%s has_chart=%s row_count=%s",
                          trace is not None, bool(badge), chart is not None,
                          trace.get("row_count") if trace else "n/a")
                if badge:
                    collected_badges.append(badge.strip())
                if trace:
                    collected_traces.append(trace)
                if chart:
                    collected_charts.append(chart)
            except Exception as e:
                _log.error("tool_result parse error: %s", e)
            tool_parts.append(
                types.Part(
                    function_response=types.FunctionResponse(
                        name=fn.name,
                        response={"result": result},
                    )
                )
            )

        response = chat.send_message(tool_parts)

    _log.info("agentic_loop done badges=%d traces=%d charts=%d",
              len(collected_badges), len(collected_traces), len(collected_charts))

    try:
        # Join all text parts (Gemini may split text + function_response across parts)
        text_parts = [p.text for p in response.candidates[0].content.parts if hasattr(p, "text") and p.text]
        final_text = "\n".join(text_parts).strip()
    except (IndexError, AttributeError) as e:
        _log.error("chat/ask failed to extract response user=%s error=%s", username, e)
        raise HTTPException(status_code=502, detail="No se pudo obtener respuesta del modelo")

    # Strip stray HTML tags Gemini sometimes emits (e.g. <br>)
    final_text = re.sub(r"^(\s*<br\s*/?>)+\s*", "", final_text, flags=re.IGNORECASE).strip()

    # Extract __chart__ payloads Gemini sometimes embeds as raw JSON in the text
    # instead of calling the render_chart tool properly
    inline_chart_re = re.compile(r'\{["\s]*__chart__["\s]*:.*?\}\s*\}', re.DOTALL)
    for match in inline_chart_re.finditer(final_text):
        try:
            payload = json.loads(match.group())
            chart = payload.get("__chart__")
            if chart and isinstance(chart, dict):
                collected_charts.append(chart)
        except Exception:
            pass
    final_text = inline_chart_re.sub("", final_text).strip()

    # If there's a chart and the text is empty/trivial, use a default message
    if collected_charts and not final_text:
        final_text = "Aquí tienes el gráfico con los datos consultados."

    # Backend safety net: if Gemini omitted the source badge, append it
    if collected_badges:
        missing = [b for b in collected_badges if b not in final_text]
        if missing:
            final_text = final_text.rstrip() + "\n\n" + "\n".join(missing)

    # Anti-hallucination guard: block responses with numbers but no source badge
    # and no tool was called (i.e. Gemini answered from memory)
    _has_numbers = bool(re.search(r'\b\d[\d\.]{2,}\b', final_text))
    _has_source  = "__source_badge__" in final_text or "📊 Fuente:" in final_text
    _tool_was_called = len(collected_badges) > 0 or len(collected_charts) > 0 or len(collected_traces) > 0

    if _has_numbers and not _has_source and not _tool_was_called:
        _log.warning(
            "chat/ask blocked hallucination — numbers without source user=%s reply_preview=%.120s",
            username, final_text
        )
        final_text = (
            "⚠️ No puedo responder esta pregunta sin consultar la base de datos primero. "
            "Por favor reformula tu pregunta sobre datos internos de Empresas Donar "
            "(facturas, remuneraciones, tarjas, despacho, sensores, etc.)."
        )

    # Store in cache only if a tool was called (i.e. real data was queried)
    if embedding and _tool_was_called:
        cache.put(user_question, embedding, final_text, collected_traces, collected_charts or None)

    # Persist to DB (fire and forget — don't block the response)
    _log.info("saving user=%s traces=%d charts=%d tool_was_called=%s",
              username, len(collected_traces), len(collected_charts), _tool_was_called)
    _save_messages(username, user_question, final_text,
                   collected_traces if collected_traces else None,
                   collected_charts if collected_charts else None)

    return JSONResponse({"reply": final_text, "traces": collected_traces, "charts": collected_charts, "from_cache": False})


@router.post("/download")
async def chat_download(request: Request):
    from fastapi import HTTPException
    body = await request.json()
    fmt = body.get("format", "xlsx").lower()
    sql = body.get("sql", "").strip()
    system = body.get("system", "bigquery")

    if not sql or _WRITE_RE.match(sql):
        raise HTTPException(status_code=400, detail="SQL inválida")

    try:
        if system == "bigquery" or "ace-scarab" in sql:
            rows = bq.run_query(sql, 5000)
        else:
            rows = pg.run_pg_query(sql, 5000)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al consultar: {e}")

    if not rows:
        raise HTTPException(status_code=404, detail="La consulta no devolvió datos")

    if fmt == "pdf":
        return _build_pdf(rows, sql)
    return _build_xlsx(rows, sql)


def _build_xlsx(rows: list[dict], sql: str) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    headers = list(rows[0].keys())
    header_fill = PatternFill("solid", fgColor="2D6A4F")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row.get(header)
            ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="F0F4F0")

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(headers[col_idx - 1])),
            max((len(str(row.get(headers[col_idx - 1]) or "")) for row in rows[:50]), default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=datos_donar.xlsx"},
    )


def _build_pdf(rows: list[dict], sql: str) -> StreamingResponse:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    headers = list(rows[0].keys())
    page_size = landscape(A4) if len(headers) > 6 else A4

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=page_size,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Datos — Empresas Donar", styles["Title"]))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Total registros: {len(rows)}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    # Build table data
    col_widths = None
    page_width = page_size[0] - 3*cm
    col_w = page_width / len(headers)
    col_widths = [col_w] * len(headers)

    table_data = [headers]
    for row in rows:
        table_data.append([str(row.get(h) or "") for h in headers])

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D6A4F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F0")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t)

    doc.build(story)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=datos_donar.pdf"},
    )
