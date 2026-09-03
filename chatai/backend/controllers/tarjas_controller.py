"""
controllers/tarjas_controller.py
---------------------------------
HTTP layer for the Tarjas reports (replaces Looker Studio pages).

Routes:
  GET  /tarjas/general                  → General operational overview
  GET  /api/tarjas/general/filters      → Filter options (general)
  GET  /api/tarjas/general              → Aggregated data (labor summary, person ranking, chart)

  GET  /tarjas/detalle                  → Weekly detail page
  GET  /api/tarjas/detalle/filters      → Filter options (detalle)
  GET  /api/tarjas/detalle              → Detail data

  GET  /tarjas/detalle-tractorista       → Weekly tractorista detail (Looker)
  GET  /api/tarjas/detalle-tractorista/filters
  GET  /api/tarjas/detalle-tractorista   → Nested Looker pivot (fecha × trabajador × labor)
  PATCH /api/tarjas/detalle-tractorista/fila  → Update estado for contratista × fecha
  PATCH /api/tarjas/detalle-tractorista/celda → Update total_tractor for a pivot cell

  GET  /tarjas/contratista              → Contractor/worker pivot page
  GET  /api/tarjas/contratista/filters  → Filter options (contratista)
  GET  /api/tarjas/contratista          → Worker-level data for pivot table

  GET  /tarjas/contratista-tractorista       → Tractorista worker pivot (Looker)
  GET  /api/tarjas/contratista-tractorista/filters
  GET  /api/tarjas/contratista-tractorista   → Raw tarjas_pagos rows (tractorista only)

  GET  /tarjas/notas                    → Notas de crédito page (contractor payment report)
  GET  /api/tarjas/notas/filters        → Filter options (notas)
  GET  /api/tarjas/notas                → Report data

  GET  /tarjas/bono-mensual             → Bonos mensuales page
  GET  /api/tarjas/bono-mensual/filters → Filter options (bono mensual)
  GET  /api/tarjas/bono-mensual         → Report data (bonos mensuales del mes)

  GET  /tarjas/hora-ponderada-9h              → Hora ponderada 9h page (Labor x CC pivot)
  GET  /api/tarjas/hora-ponderada-9h/filters  → Filter options (hora ponderada 9h)
  GET  /api/tarjas/hora-ponderada-9h          → Report data, one row per Labor+CC+Fecha cell

  GET  /tarjas/registros-campo               → App field-record timeline (audit)
  GET  /api/tarjas/registros-campo/filters   → Distinct filter dropdowns
  GET  /api/tarjas/registros-campo           → Paginated raw tarjas_pagos rows
  PATCH /api/tarjas/registros-campo/{id}     → Edit mal-digitado fields (allowlisted)
  GET  /api/tarjas/registros-campo/download-excel

  GET  /tarjas/calendario                    → App monthly calendar of field records + plans
  GET  /api/tarjas/calendario                → Per-day registro counts + overlapping plans
  GET  /api/tarjas/calendario/planes         → Planificación overlapping a given day
"""

import base64
import datetime
import decimal
import io
from pathlib import Path
import logging
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from xhtml2pdf import pisa

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from psycopg2 import sql as psql

from auth import require_auth
from db import get_connection

logger = logging.getLogger("controllers.tarjas")

router = APIRouter(dependencies=[Depends(require_auth)])

_templates: Jinja2Templates = None
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_MONTH_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")

_TRACTORISTA_PAGOS_SQL = "(LOWER(TRIM(tipo_pago)) = 'tractorista')"
_DETALLE_TRACT_ESTADOS = ("Pendiente", "Aprobado")


def _fmt_clp(v) -> str:
    try:
        return f"${int(float(v or 0)):,}".replace(",", ".")
    except Exception:
        return "-"


def _fmt_date_display(iso: str) -> str:
    try:
        d = datetime.date.fromisoformat(iso)
        months = [
            "ene",
            "feb",
            "mar",
            "abr",
            "may",
            "jun",
            "jul",
            "ago",
            "sep",
            "oct",
            "nov",
            "dic",
        ]
        return f"{d.day} {months[d.month - 1]} {d.year}"
    except Exception:
        return iso


def _fmt_date_slash(iso: str) -> str:
    """DD/MM/YYYY — the project's date-display convention (CLAUDE.md: "sin
    excepciones"). Used in PDF headers (_pdf_header's Desde/Hasta chips and
    resumen-persona's custom header); kept separate from _fmt_date_display()
    above, which uses a different "1 jul 2026" style already relied on for
    table body date cells across several PDFs."""
    try:
        d = datetime.date.fromisoformat(iso)
        return f"{d.day:02d}/{d.month:02d}/{d.year}"
    except Exception:
        return iso


def _empresa_to_campo(empresa: str | None) -> str | None:
    return empresa or None


def _get_empresas(
    cur, table: str = "appsheet.tarjas_pagos", extra_where: str = ""
) -> list[str]:
    """Return distinct nombre_campo values as empresa options."""
    where = f"WHERE nombre_campo IS NOT NULL {('AND ' + extra_where) if extra_where else ''}"
    cur.execute(
        f"SELECT DISTINCT nombre_campo FROM {table} {where} ORDER BY nombre_campo"
    )
    return [r[0] for r in cur.fetchall()]


def _resolve_maquina_column(cur):
    """Return appsheet.tarjas_pagos column for machine/tractor, if any."""
    cur.execute(
        """
        SELECT column_name FROM information_schema.columns
        WHERE table_schema = 'appsheet' AND table_name = 'tarjas_pagos'
          AND (
            lower(column_name) IN ('maquina', 'máquina')
            OR lower(column_name) LIKE '%maquin%'
          )
        ORDER BY CASE WHEN lower(column_name) = 'maquina' THEN 0 ELSE 1 END,
                 length(column_name)
        LIMIT 1
        """
    )
    r = cur.fetchone()
    return r[0] if r else None


def init(templates: Jinja2Templates) -> None:
    global _templates
    _templates = templates


def _serialize(v):
    if isinstance(v, decimal.Decimal):
        return float(v)
    if isinstance(v, (datetime.date, datetime.datetime)):
        return str(v)
    return v


def _rows_to_dicts(cur):
    cols = [d[0] for d in cur.description]
    return [{k: _serialize(v) for k, v in zip(cols, r)} for r in cur.fetchall()]


# ===========================================================================
# PDF helpers
# ===========================================================================


def _logo_b64() -> str:
    """Return the Donar logo as a base64-encoded PNG, resized to max 200 px wide.

    The original PNG is 1288×539 px (~680 KB, ~907 KB base64). Embedding that
    full size into the HTML string fed to xhtml2pdf causes text-layout corruption
    in table cells (issue #12). Resizing to ≤200 px wide brings the base64 to
    ~29 KB, which xhtml2pdf handles correctly.
    """
    path = (
        Path(__file__).parent.parent.parent
        / "frontend"
        / "static"
        / "img"
        / "donar_logo.png"
    )
    try:
        from PIL import Image

        img = Image.open(path)
        max_w = 200
        w, h = img.size
        if w > max_w:
            img = img.resize((max_w, int(h * max_w / w)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        return base64.b64encode(buf.getvalue()).decode()
    except Exception:
        return ""


def _summary_table_html(resumen: list[dict], total: float, jornadas) -> str:
    """Build the Resumen table HTML for the Detalle PDF — same columns,
    currency format, and Total row as the on-screen td-summary-table, plus
    a "%" column with each row's share of the grand total."""

    def _pct(value):
        return f"{(float(value or 0) / total * 100):.1f} %" if total > 0 else "—"

    # Column widths are % of this table's own (now narrower-than-100%) box
    # — without them "Jornadas"/"%" stretched much wider than their short
    # values need, across the full page (issue #132).
    total_trab = sum(float(r.get("total_trabajado") or 0) for r in resumen)
    W = {
        "tipo": "width:22%",
        "total": "width:22%",
        "trab": "width:22%",
        "jornadas": "width:17%",
        "pct": "width:17%",
    }
    rows_html = "".join(
        f'<tr><td style="{W["tipo"]}"><span class="{_tipo_pago_badge_class(r["tipo_pago"])}">'
        f'{_escape_html(_tipo_pago_label(r["tipo_pago"]))}</span></td>'
        f'<td class="num" style="{W["total"]}">{_fmt_clp(r["total_pagar"])}</td>'
        f'<td class="num" style="{W["trab"]}">{_fmt_clp(r.get("total_trabajado"))}</td>'
        f'<td class="num" style="{W["jornadas"]}">{r["jornadas"]}</td>'
        f'<td class="num" style="{W["pct"]}">{_pct(r["total_pagar"])}</td></tr>'
        for r in resumen
    )
    return f"""
    <table class="summary-table" style="width:62%;table-layout:fixed">
      <thead><tr>
        <th style="{W["tipo"]}">Tipo de pago</th>
        <th class="num" style="{W["total"]}">Total a pagar</th>
        <th class="num" style="{W["trab"]}">Total trabajado</th>
        <th class="num" style="{W["jornadas"]}">Jornadas</th>
        <th class="num" style="{W["pct"]}">%</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
      <tfoot><tr>
        <td style="{W["tipo"]}">Total</td>
        <td class="num" style="{W["total"]}">{_fmt_clp(total)}</td>
        <td class="num" style="{W["trab"]}">{_fmt_clp(total_trab)}</td>
        <td class="num" style="{W["jornadas"]}">{jornadas}</td>
        <td class="num" style="{W["pct"]}">{"100.0 %" if total > 0 else "—"}</td>
      </tr></tfoot>
    </table>
    """


_PDF_CSS = """
@page { size: A4 landscape; margin: 12mm 10mm; }
body { font-family: Helvetica, Arial, sans-serif; font-size: 8pt; color: #111111; margin: 0; }
.ph { border-bottom: 1px solid #cccccc; padding: 6px 0 8px 0; margin-bottom: 6px; }
.ph h1 { font-size: 12pt; font-weight: bold; color: #111111; margin: 0 0 2px 0; }
.ph-sub { font-size: 7pt; color: #555555; margin: 0; }
.ph-chips { padding: 4px 0; margin-bottom: 8px; }
.chip { display: inline-block; border: 1px solid #aaaaaa; border-radius: 2px; padding: 1px 6px; font-size: 7pt; color: #333333; margin: 2px 3px 2px 0; }
.chip b { color: #111111; }
table { width: 100%; border-collapse: collapse; font-size: 7.5pt; }
th { background: #eeeeee; color: #111111; padding: 5px 6px; text-align: left; border: 1px solid #aaaaaa; font-weight: bold; }
th.num { text-align: right; }
td { padding: 3px 6px; border: 1px solid #cccccc; vertical-align: middle; }
tr:nth-child(even) td { background: #f7f7f7; }
tr.worker-first td { border-top: 1.5px solid #888888; }
.num { text-align: right; }
.total { text-align: right; font-weight: bold; border-left: 1.5px solid #888888; }
.section-title { font-size: 9pt; font-weight: bold; margin: 12px 0 5px 0; color: #111111; }
table.pivot-wide { table-layout: fixed; font-size: 6.5pt; }
table.pivot-wide th, table.pivot-wide td { padding: 3px 3px; overflow: hidden; }
.pdf-note { font-size: 7pt; font-style: italic; color: #666666; margin: 0 0 8px 0; }
.pdf-summary { width: 100%; border-collapse: collapse; margin: 0 0 8px 0; }
.pdf-summary td { width: 33.33%; border: 1px solid #cccccc; background: #f7f7f7; padding: 6px 10px; font-size: 8pt; text-align: center; }
.pdf-summary b { display: block; font-size: 11pt; color: #111111; }
.summary-table { width: 100%; border-collapse: collapse; font-size: 7.5pt; margin-bottom: 4px; }
.summary-table th { background: #ea8c1e; color: #ffffff; padding: 6px 8px; text-align: left; border: 1px solid #ea8c1e; font-weight: bold; }
.summary-table th.num { text-align: right; }
.summary-table td { padding: 5px 6px; border: 1px solid #cccccc; }
.summary-table tbody tr:nth-child(even) td { background: #fdf3d0; }
.summary-table tfoot td { border-top: 1.5px solid #888888; font-weight: bold; }
.badge-trato { display: inline-block; padding: 2px 10px; border-radius: 4px; background: #dbeafe; color: #1d4ed8; font-weight: bold; }
.badge-aldia { display: inline-block; padding: 2px 10px; border-radius: 4px; background: #ffedd5; color: #c2410c; font-weight: bold; }
table.detalle-table th { background: #1a1a1a; color: #f5d87a; padding: 6px 8px; text-align: left; border: 1px solid #1a1a1a; font-weight: bold; }
table.detalle-table th.num { text-align: right; }
table.detalle-table tbody tr:nth-child(even) td { background: #f0ebe1; }
table.detalle-table tfoot td { border-top: 1.5px solid #888888; font-weight: bold; background: #fdf3d0; }
table.detalle-tract-pivot { margin-top: 12px; }
table.detalle-tract-pivot th.th-group { background: #1F4E79; color: #ffffff; text-align: center; }
table.detalle-tract-pivot tfoot td { font-weight: bold; background: #e8eef4; }
"""


# Beyond this many date columns, each column becomes too narrow for
# reportlab to lay out text without overlapping neighboring cells (the
# wide-pivot PDFs are meant for weekly/biweekly ranges, matching how these
# reports are actually used — not multi-month exports).
MAX_PIVOT_DATES = 45


def _check_pivot_date_range(
    dates: list[str], report_label: str, max_dates: int = MAX_PIVOT_DATES
) -> None:
    if len(dates) > max_dates:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El rango de fechas es demasiado amplio para el PDF de "
                f"'{report_label}' ({len(dates)} días). Reduce el rango a "
                f"{max_dates} días o menos, o usa la descarga en Excel."
            ),
        )


def _pdf_title(report_name: str, contratista: str | None) -> str:
    """Build the PDF title as 'Tarjas-Reporte {report_name}-{Contratista}',
    matching the naming issue #54 asked for. The contratista suffix is
    omitted when the report isn't filtered to a single contratista."""
    title = f"Tarjas-Reporte {report_name}"
    if contratista:
        title += f"-{contratista.title()}"
    return title


def _escape_html(s: str) -> str:
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _pivot_col_widths(
    fixed_pct: dict[str, float], n_dates: int, date_pct: float = 3.5
) -> dict[str, str]:
    """Compute inline width styles for a wide date-pivot PDF table so xhtml2pdf's
    table-layout:fixed always fits the page — without explicit widths, reportlab
    raises 'negative availWidth' once there are enough date columns to overflow
    the page (e.g. a single week already crashes it: 7 dates + 5 fixed columns).

    Each date column gets a fixed percent-of-page width (enough for a DD/MM
    header or a short money/hours value) instead of always stretching to
    fill 100% — a range with only a handful of dates renders as a narrower
    table instead of one stretched edge-to-edge with mostly empty columns
    (issue #132). `widths["table"]` carries the resulting overall table
    width; callers must apply it to the <table> tag itself, since the
    other returned widths are percentages OF THAT (possibly narrower)
    table width, not of the page. Once enough dates would overflow 100%,
    falls back to the old behavior: fixed columns keep their given %, the
    remainder is split evenly across dates, table stays full width.
    """
    fixed_total = sum(fixed_pct.values())
    natural_total = fixed_total + date_pct * n_dates
    if n_dates and natural_total <= 100.0:
        table_pct = natural_total
        scale = 100.0 / table_pct
        widths = {k: f"width:{v * scale}%" for k, v in fixed_pct.items()}
        widths["date"] = f"width:{date_pct * scale}%"
    else:
        table_pct = 100.0 if n_dates else fixed_total
        remaining = max(0.0, 100.0 - fixed_total)
        dpct = (remaining / n_dates) if n_dates else 0.0
        widths = {k: f"width:{v}%" for k, v in fixed_pct.items()}
        widths["date"] = f"width:{dpct}%"
    widths["table"] = f"width:{table_pct}%"
    return widths


def _pdf_header(
    title: str, fecha_inicio: str, fecha_termino: str, filters: dict
) -> str:
    now = datetime.date.today().strftime("%-d de %B de %Y")
    chips = ""
    for k, v in filters.items():
        if v:
            chips += f'<span class="chip"><b>{k}:</b> {v}</span> '
    logo = _logo_b64()
    logo_cell = (
        (
            f'<td style="border:none;width:90px;padding:4px 8px;background:#1e293b;vertical-align:middle">'
            f'<img src="data:image/png;base64,{logo}" style="width:80px;height:auto" /></td>'
        )
        if logo
        else ""
    )
    return f"""
    <table style="width:100%;border:none;margin-bottom:6px">
      <tr>
        {logo_cell}
        <td style="border:none;padding:0 0 0 10px;vertical-align:middle">
          <div class="ph">
            <h1>{title}</h1>
            <p class="ph-sub">Generado el {now}</p>
          </div>
        </td>
      </tr>
    </table>
    <div class="ph-chips">
      <span class="chip"><b>Desde:</b> {_fmt_date_slash(fecha_inicio)}</span>
      <span class="chip"><b>Hasta:</b> {_fmt_date_slash(fecha_termino)}</span>
      {chips}
    </div>
    """


def _render_pdf(html: str, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


# ===========================================================================
# General operacional — Labor averages, person ranking, daily chart
# ===========================================================================


@router.get("/tarjas/general", response_class=HTMLResponse)
async def tarjas_general_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_general.html")


@router.get("/api/tarjas/general/filters")
async def get_tarjas_general_filters():
    """Distinct values for filter dropdowns (from tarjas_pagos)."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT cuartel_cc FROM appsheet.tarjas_pagos "
                "WHERE cuartel_cc IS NOT NULL ORDER BY cuartel_cc"
            )
            centros_costo = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT tipo_pago FROM appsheet.tarjas_pagos "
                "WHERE tipo_pago IS NOT NULL ORDER BY tipo_pago"
            )
            tipos_pago = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT labor FROM appsheet.tarjas_pagos "
                "WHERE labor IS NOT NULL ORDER BY labor"
            )
            labores = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(cur, "appsheet.tarjas_pagos")
    finally:
        conn.close()

    return {
        "centros_costo": centros_costo,
        "tipos_pago": tipos_pago,
        "labores": labores,
        "contratistas": contratistas,
        "empresas": empresas,
    }


def _build_pagos_where(
    fecha_inicio,
    fecha_termino,
    centro_costo,
    tipo_pago,
    labor,
    alias="",
    contratista=None,
    nombre_campo=None,
):
    """Build WHERE clause + params for tarjas_pagos queries."""
    pfx = f"{alias}." if alias else ""
    filters = [f"{pfx}fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if centro_costo:
        filters.append(f"{pfx}cuartel_cc = %s")
        params.append(centro_costo)
    if tipo_pago:
        filters.append(f"{pfx}tipo_pago = %s")
        params.append(tipo_pago)
    if labor:
        filters.append(f"{pfx}labor = %s")
        params.append(labor)
    if contratista:
        filters.append(f"{pfx}contratista = %s")
        params.append(contratista)
    if nombre_campo:
        filters.append(f"{pfx}nombre_campo = %s")
        params.append(nombre_campo)
    return "WHERE " + " AND ".join(filters), params


@router.get("/api/tarjas/general")
async def get_tarjas_general(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    where, params = _build_pagos_where(
        fecha_inicio,
        fecha_termino,
        centro_costo,
        tipo_pago,
        labor,
        contratista=contratista,
        nombre_campo=_empresa_to_campo(empresa),
    )

    try:
        with conn.cursor() as cur:
            _horas_expr = "NULLIF(SUM(horas_trabajadas), 0)"
            _horas_sum  = "COALESCE(SUM(horas_trabajadas), 0)"
            # 1) Average earnings per labor
            cur.execute(
                f"""
                SELECT
                    labor,
                    ROUND(AVG(total_trabajado)::numeric, 0)                        AS promedio_diario,
                    ROUND(SUM(total_trabajado)::numeric / {_horas_expr}, 0)        AS ganancia_hora,
                    ROUND(({_horas_sum})::numeric, 1)                              AS total_horas,
                    COALESCE(SUM(total_trabajado), 0)                              AS total
                FROM appsheet.tarjas_pagos
                {where}
                GROUP BY labor
                ORDER BY total DESC
            """,
                params,
            )
            labor_summary = _rows_to_dicts(cur)

            # 2) Person ranking (top 6 earners)
            cur.execute(
                f"""
                SELECT
                    trabajador,
                    contratista,
                    ROUND(AVG(total_trabajado)::numeric, 0)                        AS promedio_diario,
                    ROUND(SUM(total_trabajado)::numeric / {_horas_expr}, 0)        AS ganancia_hora,
                    ROUND(({_horas_sum})::numeric, 1)                              AS total_horas,
                    COALESCE(SUM(total_trabajado), 0)                              AS total
                FROM appsheet.tarjas_pagos
                {where}
                GROUP BY trabajador, contratista
                ORDER BY total DESC
                LIMIT 6
            """,
                params,
            )
            person_ranking = _rows_to_dicts(cur)

            # 3) Daily average by labor × cuadrilla (top 6 workers)
            p_where, p_params = _build_pagos_where(
                fecha_inicio,
                fecha_termino,
                centro_costo,
                tipo_pago,
                labor,
                alias="p",
                contratista=contratista,
                nombre_campo=_empresa_to_campo(empresa),
            )
            cur.execute(
                f"""
                WITH top_workers AS (
                    SELECT trabajador
                    FROM appsheet.tarjas_pagos
                    {where}
                    GROUP BY trabajador
                    ORDER BY SUM(total_trabajado) DESC
                    LIMIT 6
                )
                SELECT
                    p.labor,
                    p.trabajador,
                    ROUND(AVG(p.total_trabajado)::numeric, 2) AS avg_daily
                FROM appsheet.tarjas_pagos p
                JOIN top_workers tw ON tw.trabajador = p.trabajador
                {p_where}
                GROUP BY p.labor, p.trabajador
                ORDER BY p.labor, avg_daily DESC
            """,
                params + p_params,
            )
            chart_data = _rows_to_dicts(cur)
    finally:
        conn.close()

    return {
        "labor_summary": labor_summary,
        "person_ranking": person_ranking,
        "chart_data": chart_data,
    }


# ===========================================================================
# Detalle semanal
# ===========================================================================


@router.get("/tarjas/detalle", response_class=HTMLResponse)
async def tarjas_detail_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_detail.html")


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------


@router.get("/api/tarjas/detalle/filters")
async def get_tarjas_filters():
    """Distinct values for each filter dropdown."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_reporte ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]

            cur.execute(
                'SELECT DISTINCT "CC" FROM appsheet.tarjas_reporte WHERE "CC" IS NOT NULL ORDER BY "CC"'
            )
            centros_costo = [r[0] for r in cur.fetchall()]

            cur.execute(
                'SELECT DISTINCT "Nombre Labor" FROM appsheet.tarjas_reporte ORDER BY "Nombre Labor"'
            )
            labores = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT nombre_campo FROM appsheet.tarjas_reporte ORDER BY nombre_campo"
            )
            campos = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(cur, "appsheet.tarjas_reporte")
    finally:
        conn.close()

    return {
        "contratistas": contratistas,
        "empresas": empresas,
        "centros_costo": centros_costo,
        "labores": labores,
        "campos": campos,
    }


def _build_detalle_filters(
    fecha_inicio,
    fecha_termino,
    contratista=None,
    empresa=None,
    centro_costo=None,
    tipo_pago=None,
    labor=None,
    campo=None,
):
    """WHERE for Detalle operacional on tarjas_pagos (Aprobado only).

    Costo total / Total a pagar stay on total_pagar. Total trabajado is
    SUM(total_trabajado) so the cuadrilla figure still shows when AppSheet
    leaves total_pagar at 0.
    """
    filters = [
        "fecha::date BETWEEN %s AND %s",
        "estado = 'Aprobado'",
    ]
    params: list = [fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if centro_costo:
        filters.append("cuartel_cc = %s")
        params.append(centro_costo)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if labor:
        filters.append("labor = %s")
        params.append(labor)
    if campo:
        filters.append("nombre_campo = %s")
        params.append(campo)
    return "WHERE " + " AND ".join(filters), params


def _query_detalle_rows(cur, where, params):
    cur.execute(
        f"""
        SELECT
            p.tipo_pago,
            p.labor                                                   AS labor,
            p.cuartel_cc                                              AS centro_costo,
            cc.cultivo                                                AS centro_costo_nombre,
            COUNT(*)                                                  AS jornadas,
            SUM(p.horas_trabajadas)                                   AS horas_trabajadas,
            CASE WHEN COUNT(*) > 0
                 THEN ROUND((SUM(p.total_pagar) / COUNT(*))::numeric, 2)
                 ELSE NULL END                                        AS total_unitario,
            SUM(p.total_pagar)                                        AS costo_total,
            SUM(p.total_trabajado)                                    AS total_trabajado,
            CASE WHEN SUM(p.horas_trabajadas) > 0
                 THEN ROUND((SUM(p.total_pagar) / SUM(p.horas_trabajadas))::numeric, 0)
                 ELSE NULL END                                        AS costo_hora,
            ROUND(
                SUM(p.total_pagar)::numeric
                / NULLIF(
                    SUM(SUM(p.total_pagar)) FILTER (
                        WHERE p.tipo_pago IN ('trato', 'Al dia', 'Al día')
                    ) OVER (),
                    0
                  ) * 100,
                2
            )                                                         AS pct_pago,
            p.nombre_campo
        FROM appsheet.tarjas_pagos p
        LEFT JOIN appsheet.tarjas_cc cc ON cc.id_cc::text = p.cuartel_cc::text
        {where}
        GROUP BY p.tipo_pago, p.labor, p.cuartel_cc, cc.cultivo, p.nombre_campo
        ORDER BY p.tipo_pago DESC, p.labor, p.cuartel_cc
    """,
        params,
    )
    return _rows_to_dicts(cur)


def _query_detalle_resumen(cur, where, params):
    """Sum total_pagar and total_trabajado by tipo_pago — Resumen table
    and pie chart (pie stays on total_pagar)."""
    cur.execute(
        f"""
        SELECT tipo_pago,
               COALESCE(SUM(total_pagar), 0)     AS total_pagar,
               COALESCE(SUM(total_trabajado), 0) AS total_trabajado,
               COUNT(*)                          AS jornadas
        FROM appsheet.tarjas_pagos {where}
        GROUP BY tipo_pago ORDER BY tipo_pago
    """,
        params,
    )
    return _rows_to_dicts(cur)


# Mirrors the TIPO_LABELS / TIPO_CLASS maps in tarjas_detail.js so the PDF
# summary table and pie chart match the screen exactly, including the
# fallback for tipo_pago values outside "trato"/"Al dia" (e.g. "Tractorista",
# "Bono"): plain label, no badge class, orange slice.
_TIPO_PAGO_LABELS = {"trato": "Trato", "Al dia": "Al Día", "Al día": "Al Día"}
_TIPO_PAGO_BADGE_CLASS = {
    "trato": "badge-trato",
    "Al dia": "badge-aldia",
    "Al día": "badge-aldia",
}


def _tipo_pago_label(tipo_pago: str) -> str:
    return _TIPO_PAGO_LABELS.get(tipo_pago, tipo_pago)


def _tipo_pago_badge_class(tipo_pago: str) -> str:
    return _TIPO_PAGO_BADGE_CLASS.get(tipo_pago, "")


@router.get("/api/tarjas/detalle")
async def get_tarjas_detail(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
    campo: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    where, params = _build_detalle_filters(
        fecha_inicio,
        fecha_termino,
        contratista,
        empresa,
        centro_costo,
        tipo_pago,
        labor,
        campo,
    )

    try:
        with conn.cursor() as cur:
            resumen = _query_detalle_resumen(cur, where, params)
            rows = _query_detalle_rows(cur, where, params)
    finally:
        conn.close()

    total_general = sum(r["total_pagar"] for r in resumen)
    total_trabajado = sum(float(r["total_trabajado"] or 0) for r in resumen)
    jornadas_general = sum(r["jornadas"] for r in resumen)

    return {
        "resumen": resumen,
        "total": total_general,
        "total_trabajado": total_trabajado,
        "jornadas": jornadas_general,
        "rows": rows,
        "count": len(rows),
    }


# ===========================================================================
# Detalle de la semana tractorista (Looker: Detalle tractorista)
# ===========================================================================


def _detalle_tractorista_where(
    fecha_inicio: str,
    fecha_termino: str,
    contratista: str | None = None,
    empresa: str | None = None,
    centro_costo: str | None = None,
    labor: str | None = None,
    campo: str | None = None,
) -> tuple[str, list]:
    """WHERE for Detalle tractorista on tarjas_pagos (all estados).

    tarjas_reporte hides Pendiente and sums total_pagar; tractorista money
    is total_tractor (issue #152).
    """
    filters = ["fecha::date BETWEEN %s AND %s", _TRACTORISTA_PAGOS_SQL]
    params: list = [fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if centro_costo:
        filters.append("cuartel_cc = %s")
        params.append(centro_costo)
    if labor:
        filters.append("labor = %s")
        params.append(labor)
    if campo:
        filters.append("nombre_campo = %s")
        params.append(campo)
    return "WHERE " + " AND ".join(filters), params


def _fetch_detalle_tractorista_filter_options(cur, where: str, params: list) -> dict:
    """Distinct dropdown values for the current tractorista WHERE (one scan)."""
    cur.execute(
        f"""
        SELECT contratista, nombre_campo, cuartel_cc, labor
        FROM appsheet.tarjas_pagos
        {where}
        """,
        params,
    )
    contratistas, empresas, centros, labores = set(), set(), set(), set()
    for contratista, campo, cc, labor in cur.fetchall():
        if contratista:
            contratistas.add(contratista)
        if campo:
            empresas.add(campo)
        if cc:
            centros.add(str(cc))
        if labor:
            labores.add(labor)
    return {
        "contratistas": sorted(contratistas),
        "empresas": sorted(empresas),
        "centros_costo": sorted(centros),
        "labores": sorted(labores),
    }


def _fetch_detalle_tractorista_rows(cur, where: str, params: list) -> list[dict]:
    """Raw (fecha, contratista, trabajador, labor, monto) cells for the
    Looker nested pivot. Shared by JSON, Excel and PDF (issues #152/#153)."""
    cur.execute(
        f"""
        SELECT
            fecha::date::text AS fecha,
            contratista,
            trabajador,
            labor,
            COALESCE(SUM(total_tractor), 0) AS monto,
            MIN(estado) AS estado,
            COUNT(DISTINCT estado) AS n_estados
        FROM appsheet.tarjas_pagos
        {where}
        GROUP BY fecha::date, contratista, trabajador, labor
        ORDER BY contratista, fecha::date, trabajador, labor
        """,
        params,
    )
    return _rows_to_dicts(cur)


def _general_tractorista_where(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    centro_costo: str | None = None,
    labor: str | None = None,
    maquina: str | None = None,
    contratista: str | None = None,
    empresa: str | None = None,
) -> tuple[str, list]:
    """Same WHERE as the General tractorista screen — JSON, Excel and PDF."""
    filters = ["fecha::date BETWEEN %s AND %s", _TRACTORISTA_PAGOS_SQL]
    params: list = [fecha_inicio, fecha_termino]
    if centro_costo:
        filters.append("cuartel_cc = %s")
        params.append(centro_costo)
    if labor:
        filters.append("labor = %s")
        params.append(labor)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    maq_col = _resolve_maquina_column(cur)
    if maq_col and maquina:
        frag = psql.SQL("{} = %s").format(psql.Identifier(maq_col))
        filters.append(frag.as_string(cur.connection))
        params.append(maquina)
    return "WHERE " + " AND ".join(filters), params


def _resumen_persona_tractorista_where(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    trabajador: str | None = None,
    tipo_pago: str | None = None,
    maquina: str | None = None,
    contratista: str | None = None,
    empresa: str | None = None,
) -> tuple[str, list]:
    """Same WHERE as the Resumen tractorista screen — JSON, Excel and PDF."""
    filters = ["fecha::date BETWEEN %s AND %s", _TRACTORISTA_PAGOS_SQL]
    params: list = [fecha_inicio, fecha_termino]
    if trabajador:
        filters.append("trabajador = %s")
        params.append(trabajador)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    maq_col = _resolve_maquina_column(cur)
    if maq_col and maquina:
        frag = psql.SQL("{} = %s").format(psql.Identifier(maq_col))
        filters.append(frag.as_string(cur.connection))
        params.append(maquina)
    return "WHERE " + " AND ".join(filters), params


def _fetch_general_tractorista_tables(cur, where: str, params: list) -> tuple[list, list]:
    """Labor summary + person ranking. Shared by JSON, Excel and PDF."""
    cur.execute(
        f"""
        SELECT
            labor,
            ROUND(AVG(total_tractor)::numeric, 2) AS avg_rate,
            COALESCE(SUM(total_tractor), 0)       AS total
        FROM appsheet.tarjas_pagos
        {where}
        GROUP BY labor
        ORDER BY total DESC
        """,
        params,
    )
    labor_summary = _rows_to_dicts(cur)
    cur.execute(
        f"""
        SELECT
            trabajador,
            contratista,
            ROUND(AVG(total_tractor)::numeric, 2) AS avg_rate,
            COALESCE(SUM(total_tractor), 0)       AS total
        FROM appsheet.tarjas_pagos
        {where}
        GROUP BY trabajador, contratista
        ORDER BY total DESC
        """,
        params,
    )
    return labor_summary, _rows_to_dicts(cur)


def _fetch_resumen_persona_tractorista_rows(cur, where: str, params: list) -> list[dict]:
    """Worker × contratista × máquina × horas × fecha cells. Shared by
    JSON, Excel and PDF (screen / standalone / bulk /reportes)."""
    conn = cur.connection
    maq_col = _resolve_maquina_column(cur)
    maq_select = (
        psql.SQL(", {col} AS maquina")
        .format(col=psql.Identifier(maq_col))
        .as_string(conn)
        if maq_col
        else ", NULL AS maquina"
    )
    cur.execute(
        f"""
        SELECT trabajador, contratista{maq_select}, horas_extras,
               fecha::date::text AS fecha,
               COALESCE(SUM(total_tractor), 0) AS total_tractor
        FROM appsheet.tarjas_pagos {where}
        GROUP BY trabajador, contratista, maquina, horas_extras, fecha::date
        ORDER BY trabajador, contratista, maquina, fecha::date
        """,
        params,
    )
    return _rows_to_dicts(cur)


def _pivot_resumen_persona_tractorista(rows: list[dict]) -> tuple[list[str], list[dict]]:
    """Same grouping as the screen: trabajador+contratista, then máquina+horas."""
    dates = sorted({r["fecha"] for r in rows if r.get("fecha")})
    groups: dict[tuple, dict] = {}
    for r in rows:
        gkey = (r.get("trabajador") or "", r.get("contratista") or "")
        maq = r.get("maquina") if r.get("maquina") is not None else ""
        hextra = r.get("horas_extras") if r.get("horas_extras") is not None else ""
        rkey = (maq, hextra)
        if gkey not in groups:
            groups[gkey] = {
                "trabajador": gkey[0],
                "contratista": gkey[1],
                "rows": {},
                "grand": 0.0,
            }
        g = groups[gkey]
        if rkey not in g["rows"]:
            g["rows"][rkey] = {
                "maquina": maq,
                "horas_extras": hextra,
                "by_date": {},
                "total": 0.0,
            }
        amt = float(r.get("total_tractor") or 0)
        entry = g["rows"][rkey]
        fecha = r["fecha"]
        entry["by_date"][fecha] = entry["by_date"].get(fecha, 0.0) + amt
        entry["total"] += amt
        g["grand"] += amt
    sorted_groups = sorted(groups.values(), key=lambda g: -g["grand"])
    for g in sorted_groups:
        g["row_list"] = list(g["rows"].values())
    return dates, sorted_groups


def _detalle_tract_col_key(trabajador: str, labor: str) -> str:
    return f"{trabajador}||{labor}"


def _build_detalle_tractorista_pivots(rows: list[dict]) -> list[dict]:
    """One nested pivot block per contratista: fecha × (trabajador, labor).

    Only dates that appear in `rows` are included (no empty calendar days).
    Only labores that exist for that worker in the range become columns.
    Missing cells stay None so the UI/Excel/PDF can render them empty.
    """
    from collections import defaultdict

    by_contratista: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        name = r.get("contratista") or "(sin contratista)"
        by_contratista[name].append(r)

    pivots = []
    for contratista in sorted(by_contratista):
        crows = by_contratista[contratista]
        dates = sorted({r["fecha"] for r in crows if r.get("fecha")})
        worker_labores: dict[str, set[str]] = defaultdict(set)
        for r in crows:
            worker = r.get("trabajador") or "(sin nombre)"
            worker_labores[worker].add(r.get("labor") or "")

        workers = []
        columns = []
        for worker in sorted(worker_labores):
            labores = sorted(worker_labores[worker])
            workers.append({"name": worker, "labores": labores})
            for labor in labores:
                columns.append(
                    {
                        "trabajador": worker,
                        "labor": labor,
                        "key": _detalle_tract_col_key(worker, labor),
                    }
                )

        keys = [c["key"] for c in columns]
        matrix = {d: {k: None for k in keys} for d in dates}
        for r in crows:
            fecha = r.get("fecha")
            if not fecha:
                continue
            worker = r.get("trabajador") or "(sin nombre)"
            labor = r.get("labor") or ""
            key = _detalle_tract_col_key(worker, labor)
            amount = float(r.get("monto") or 0)
            current = matrix[fecha].get(key)
            matrix[fecha][key] = amount if current is None else current + amount

        col_totals = {
            k: sum(matrix[d][k] or 0.0 for d in dates) for k in keys
        }
        date_totals = {
            d: sum(v or 0.0 for v in matrix[d].values()) for d in dates
        }
        estados_by_date: dict[str, set[str]] = defaultdict(set)
        for r in crows:
            fecha = r.get("fecha")
            if not fecha:
                continue
            n_est = int(r.get("n_estados") or 1)
            est = (r.get("estado") or "").strip()
            if n_est > 1:
                estados_by_date[fecha].add("Mixto")
            elif est:
                estados_by_date[fecha].add(est)
        date_estados = {}
        for d in dates:
            found = estados_by_date.get(d) or set()
            if len(found) == 1:
                date_estados[d] = next(iter(found))
            elif len(found) > 1:
                date_estados[d] = "Mixto"
            else:
                date_estados[d] = "Pendiente"
        pivots.append(
            {
                "contratista": contratista,
                "dates": dates,
                "workers": workers,
                "columns": columns,
                "matrix": matrix,
                "col_totals": col_totals,
                "date_totals": date_totals,
                "date_estados": date_estados,
                "grand_total": sum(col_totals.values()),
            }
        )
    return pivots


@router.get("/tarjas/detalle-tractorista", response_class=HTMLResponse)
async def tarjas_detalle_tractorista_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_detalle_tractorista.html")


@router.get("/tractoristas")
async def redirect_tractoristas_legacy():
    """Old menu URL → weekly tractorista detail."""
    return RedirectResponse(url="/tarjas/detalle-tractorista", status_code=302)


@router.get("/api/tarjas/detalle-tractorista/filters")
async def get_tarjas_detalle_tractorista_filters():
    """Filter options limited to Tractorista rows in tarjas_pagos."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            options = _fetch_detalle_tractorista_filter_options(
                cur, f"WHERE {_TRACTORISTA_PAGOS_SQL}", []
            )
    finally:
        conn.close()

    return options


@router.get("/api/tarjas/detalle-tractorista")
async def get_tarjas_detalle_tractorista(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    labor: str = Query(None),
    campo: str = Query(None),
):
    """Looker-style nested pivot: fecha × trabajador × labor, one block per
    contratista. Amounts are SUM(total_tractor) from tarjas_pagos."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    where, params = _detalle_tractorista_where(
        fecha_inicio,
        fecha_termino,
        contratista=contratista,
        empresa=empresa,
        centro_costo=centro_costo,
        labor=labor,
        campo=campo,
    )

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                    COALESCE(SUM(total_tractor), 0) AS total,
                    COUNT(*) AS jornadas
                FROM appsheet.tarjas_pagos
                {where}
                """,
                params,
            )
            total, jornadas = cur.fetchone()
            rows = _fetch_detalle_tractorista_rows(cur, where, params)
            date_where, date_params = _detalle_tractorista_where(
                fecha_inicio, fecha_termino
            )
            filter_options = _fetch_detalle_tractorista_filter_options(
                cur, date_where, date_params
            )
    finally:
        conn.close()

    pivots = _build_detalle_tractorista_pivots(rows)
    return {
        "pivots": pivots,
        "total": float(total or 0),
        "jornadas": int(jornadas or 0),
        "count": len(rows),
        "filter_options": filter_options,
        "estados": list(_DETALLE_TRACT_ESTADOS),
    }


def _detalle_tract_name_clause(col: str, value: str | None, empty_label: str) -> tuple[str, list]:
    """Match a text column to the pivot's display name, including blanks."""
    label = (value or "").strip()
    if not label or label == empty_label:
        return f"({col} IS NULL OR TRIM({col}) = '')", []
    return f"{col} = %s", [label]


def _parse_clp_int(raw) -> int:
    """CLP integer: '$34.500', '34500' or 34500 → 34500."""
    if raw is None or str(raw).strip() == "":
        raise HTTPException(status_code=422, detail="Monto requerido")
    if isinstance(raw, bool):
        raise HTTPException(status_code=422, detail="Monto inválido")
    if isinstance(raw, (int, float, decimal.Decimal)):
        n = int(round(float(raw)))
        if n < 0:
            raise HTTPException(status_code=422, detail="El monto no puede ser negativo")
        return n
    digits = re.sub(r"[^\d]", "", str(raw))
    if not digits:
        raise HTTPException(status_code=422, detail="Monto inválido")
    return int(digits)


def _update_detalle_tractorista_estado(
    cur, contratista: str, fecha: str, estado: str
) -> int:
    """Set estado on every tractorista row for that contratista + date."""
    name_sql, name_params = _detalle_tract_name_clause(
        "contratista", contratista, "(sin contratista)"
    )
    cur.execute(
        f"""
        UPDATE appsheet.tarjas_pagos
        SET estado = %s
        WHERE {_TRACTORISTA_PAGOS_SQL}
          AND fecha::date = %s
          AND {name_sql}
        """,
        [estado, fecha, *name_params],
    )
    return cur.rowcount


def _update_detalle_tractorista_monto(
    cur,
    contratista: str,
    fecha: str,
    trabajador: str,
    labor: str,
    monto: int,
) -> int:
    """Set SUM(total_tractor) for one pivot cell. If several source rows
    exist, the amount goes on the first id_Resumen and the rest become 0."""
    cont_sql, cont_params = _detalle_tract_name_clause(
        "contratista", contratista, "(sin contratista)"
    )
    trab_sql, trab_params = _detalle_tract_name_clause(
        "trabajador", trabajador, "(sin nombre)"
    )
    labor_val = labor or ""
    cur.execute(
        f"""
        SELECT "id_Resumen"
        FROM appsheet.tarjas_pagos
        WHERE {_TRACTORISTA_PAGOS_SQL}
          AND fecha::date = %s
          AND {cont_sql}
          AND {trab_sql}
          AND COALESCE(labor, '') = %s
        ORDER BY "id_Resumen"
        """,
        [fecha, *cont_params, *trab_params, labor_val],
    )
    ids = [r[0] for r in cur.fetchall()]
    if not ids:
        raise HTTPException(status_code=404, detail="Celda sin registros")
    first, rest = ids[0], ids[1:]
    cur.execute(
        'UPDATE appsheet.tarjas_pagos SET total_tractor = %s WHERE "id_Resumen" = %s',
        [monto, first],
    )
    if rest:
        cur.execute(
            'UPDATE appsheet.tarjas_pagos SET total_tractor = 0 WHERE "id_Resumen" IN %s',
            [tuple(rest)],
        )
    return len(ids)


@router.patch("/api/tarjas/detalle-tractorista/fila")
async def patch_tarjas_detalle_tractorista_fila(payload: dict = Body(...)):
    """Update estado for every tractorista tarja on a contratista × fecha row."""
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="JSON requerido")
    fecha = payload.get("fecha") or ""
    if not _DATE_RE.match(str(fecha)):
        raise HTTPException(status_code=400, detail="fecha must be YYYY-MM-DD")
    estado = (payload.get("estado") or "").strip()
    if estado not in _DETALLE_TRACT_ESTADOS:
        raise HTTPException(
            status_code=422,
            detail=f"estado debe ser {' o '.join(_DETALLE_TRACT_ESTADOS)}",
        )
    contratista = payload.get("contratista")
    try:
        conn = get_connection()
    except Exception:
        logger.exception("detalle-tractorista fila: DB connect failed")
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn:
            with conn.cursor() as cur:
                n = _update_detalle_tractorista_estado(
                    cur, contratista, fecha, estado
                )
    finally:
        conn.close()
    if n == 0:
        raise HTTPException(status_code=404, detail="Fila sin registros")
    return {"ok": True, "updated": n, "estado": estado}


@router.patch("/api/tarjas/detalle-tractorista/celda")
async def patch_tarjas_detalle_tractorista_celda(payload: dict = Body(...)):
    """Update total_tractor for one trabajador × labor × fecha cell."""
    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="JSON requerido")
    fecha = payload.get("fecha") or ""
    if not _DATE_RE.match(str(fecha)):
        raise HTTPException(status_code=400, detail="fecha must be YYYY-MM-DD")
    monto = _parse_clp_int(payload.get("total_tractor"))
    try:
        conn = get_connection()
    except Exception:
        logger.exception("detalle-tractorista celda: DB connect failed")
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn:
            with conn.cursor() as cur:
                n = _update_detalle_tractorista_monto(
                    cur,
                    payload.get("contratista"),
                    fecha,
                    payload.get("trabajador"),
                    payload.get("labor") or "",
                    monto,
                )
    finally:
        conn.close()
    return {"ok": True, "updated": n, "total_tractor": monto}


# ===========================================================================
# Detalle Contratista — Pivot by worker × date
# ===========================================================================


@router.get("/tarjas/contratista", response_class=HTMLResponse)
async def tarjas_contractor_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_contractor.html")


@router.get("/api/tarjas/contratista/filters")
async def get_tarjas_contractor_filters():
    """Distinct values for each filter dropdown (from tarjas_pagos directly)."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT cuartel_cc FROM appsheet.tarjas_pagos "
                "WHERE cuartel_cc IS NOT NULL ORDER BY cuartel_cc"
            )
            centros_costo = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT labor FROM appsheet.tarjas_pagos "
                "WHERE labor IS NOT NULL ORDER BY labor"
            )
            labores = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT tipo_pago FROM appsheet.tarjas_pagos "
                "WHERE tipo_pago IS NOT NULL ORDER BY tipo_pago"
            )
            tipos_pago = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(cur, "appsheet.tarjas_pagos")
    finally:
        conn.close()

    return {
        "contratistas": contratistas,
        "empresas": empresas,
        "centros_costo": centros_costo,
        "labores": labores,
        "tipos_pago": tipos_pago,
    }


@router.get("/api/tarjas/contratista")
async def get_tarjas_contractor_data(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
):
    """Return raw tarjas_pagos rows for the pivot table."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]

    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if centro_costo:
        filters.append("cuartel_cc = %s")
        params.append(centro_costo)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if labor:
        filters.append("labor = %s")
        params.append(labor)

    where = "WHERE " + " AND ".join(filters)

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT *, fecha::date::text AS fecha "
                f"FROM appsheet.tarjas_pagos {where} "
                "ORDER BY contratista, labor, fecha::date",
                params,
            )
            cols = [d[0] for d in cur.description]
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    return {
        "columns": cols,
        "rows": rows,
        "count": len(rows),
    }


# ===========================================================================
# Por persona Tractorista — Worker pivot (tractorista only, Looker)
# ===========================================================================


@router.get("/tarjas/contratista-tractorista", response_class=HTMLResponse)
async def tarjas_contractor_tractorista_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_contractor_tractorista.html")


@router.get("/api/tarjas/contratista-tractorista/filters")
async def get_tarjas_contractor_tractorista_filters():
    """Filter options from tarjas_pagos limited to Tractorista rows."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            maq_col = _resolve_maquina_column(cur)
            base_where = f" FROM appsheet.tarjas_pagos WHERE {_TRACTORISTA_PAGOS_SQL} "

            cur.execute(
                "SELECT DISTINCT contratista "
                + base_where
                + "AND contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT cuartel_cc "
                + base_where
                + "AND cuartel_cc IS NOT NULL ORDER BY cuartel_cc"
            )
            centros_costo = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT labor "
                + base_where
                + "AND labor IS NOT NULL ORDER BY labor"
            )
            labores = [r[0] for r in cur.fetchall()]

            maquinas = []
            if maq_col:
                cur.execute(
                    psql.SQL(
                        "SELECT DISTINCT {col} FROM appsheet.tarjas_pagos WHERE "
                        + _TRACTORISTA_PAGOS_SQL
                        + " AND {col} IS NOT NULL ORDER BY {col}"
                    ).format(col=psql.Identifier(maq_col))
                )
                maquinas = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(
                cur,
                "appsheet.tarjas_pagos",
                extra_where="LOWER(TRIM(tipo_pago)) = 'tractorista'",
            )
    finally:
        conn.close()

    return {
        "contratistas": contratistas,
        "empresas": empresas,
        "centros_costo": centros_costo,
        "labores": labores,
        "maquinas": maquinas,
        "maquina_column": maq_col,
    }


@router.get("/api/tarjas/contratista-tractorista")
async def get_tarjas_contractor_tractorista_data(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    labor: str = Query(None),
    maquina: str = Query(None),
):
    """Raw tarjas_pagos rows for the tractorista contractor pivot."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = ["fecha::date BETWEEN %s AND %s", _TRACTORISTA_PAGOS_SQL]
    params: list = [fecha_inicio, fecha_termino]

    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if centro_costo:
        filters.append("cuartel_cc = %s")
        params.append(centro_costo)
    if labor:
        filters.append("labor = %s")
        params.append(labor)

    try:
        with conn.cursor() as cur:
            maq_col = _resolve_maquina_column(cur)
            if maq_col and maquina is not None and maquina != "":
                frag = psql.SQL("{} = %s").format(psql.Identifier(maq_col))
                filters.append(frag.as_string(conn))
                params.append(maquina)

            where = "WHERE " + " AND ".join(filters)
            cur.execute(
                f"SELECT *, fecha::date::text AS fecha "
                f"FROM appsheet.tarjas_pagos {where} "
                "ORDER BY contratista, labor, fecha::date",
                params,
            )
            cols = [d[0] for d in cur.description]
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    return {
        "columns": cols,
        "rows": rows,
        "count": len(rows),
    }


# ===========================================================================
# Resumen por persona operacional — Worker summary pivot (worker × date)
# ===========================================================================


@router.get("/tarjas/resumen-persona", response_class=HTMLResponse)
async def tarjas_resumen_persona_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_resumen_persona.html")


@router.get("/api/tarjas/resumen-persona/filters")
async def get_tarjas_resumen_persona_filters():
    """Distinct trabajador + tipo_pago + contratista for filter dropdowns."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT trabajador, COALESCE(SUM(total_trabajado), 0) AS total "
                "FROM appsheet.tarjas_pagos "
                "WHERE trabajador IS NOT NULL "
                "GROUP BY trabajador ORDER BY total DESC"
            )
            trabajadores = _rows_to_dicts(cur)

            cur.execute(
                "SELECT DISTINCT tipo_pago FROM appsheet.tarjas_pagos "
                "WHERE tipo_pago IS NOT NULL ORDER BY tipo_pago"
            )
            tipos_pago = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(cur, "appsheet.tarjas_pagos")
    finally:
        conn.close()

    return {
        "trabajadores": trabajadores,
        "tipos_pago": tipos_pago,
        "contratistas": contratistas,
        "empresas": empresas,
    }


@router.get("/api/tarjas/resumen-persona")
async def get_tarjas_resumen_persona(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    """Return worker-level rows for the resumen pivot table."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]

    if trabajador:
        filters.append("trabajador = %s")
        params.append(trabajador)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))

    where = "WHERE " + " AND ".join(filters)

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT trabajador, tipo_pago, fecha::date::text AS fecha, "
                f"COALESCE(SUM(total_trabajado), 0) AS total_trabajado "
                f"FROM appsheet.tarjas_pagos {where} "
                "GROUP BY trabajador, tipo_pago, fecha::date "
                "ORDER BY trabajador, tipo_pago, fecha::date",
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    return {
        "rows": rows,
        "count": len(rows),
    }


@router.get("/api/tarjas/resumen-persona/download-excel")
async def download_tarjas_resumen_persona_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    """Descarga la tabla pivot resumen-persona como Excel."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if trabajador:
        filters.append("trabajador = %s")
        params.append(trabajador)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))

    where = "WHERE " + " AND ".join(filters)

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT trabajador, contratista, tipo_pago, fecha::date::text AS fecha, "
                f"COALESCE(SUM(total_trabajado), 0) AS total_trabajado "
                f"FROM appsheet.tarjas_pagos {where} "
                "GROUP BY trabajador, contratista, tipo_pago, fecha::date "
                "ORDER BY trabajador, tipo_pago, fecha::date",
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    # Build pivot: worker+tipo → date → total
    dates = sorted({r["fecha"] for r in rows})

    workers: dict = {}
    for r in rows:
        key = (r["trabajador"] or "", r["contratista"] or "", r["tipo_pago"] or "")
        if key not in workers:
            workers[key] = {"by_date": {}, "total": 0}
        workers[key]["by_date"][r["fecha"]] = workers[key]["by_date"].get(
            r["fecha"], 0
        ) + float(r["total_trabajado"] or 0)
        workers[key]["total"] += float(r["total_trabajado"] or 0)

    sorted_workers = sorted(workers.items(), key=lambda x: -x[1]["total"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen por persona"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="D6E4F0")
    money_fmt = "#,##0"

    # Header row
    headers = (
        ["Trabajador", "Contratista", "Tipo de pago"]
        + [datetime.date.fromisoformat(d).strftime("%d/%m/%Y") for d in dates]
        + ["Total"]
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    row_num = 2
    for (trab, cont, tipo), entry in sorted_workers:
        ws.cell(row=row_num, column=1, value=trab)
        ws.cell(row=row_num, column=2, value=cont)
        ws.cell(row=row_num, column=3, value=tipo)
        for col, d in enumerate(dates, 4):
            val = entry["by_date"].get(d, 0)
            c = ws.cell(row=row_num, column=col, value=val if val else None)
            c.number_format = money_fmt
        total_cell = ws.cell(row=row_num, column=4 + len(dates), value=entry["total"])
        total_cell.number_format = money_fmt
        total_cell.fill = total_fill
        total_cell.font = Font(bold=True)
        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    for i in range(len(dates) + 1):
        col_letter = openpyxl.utils.get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"resumen_persona_{fecha_inicio}_{fecha_termino}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ===========================================================================
# Resumen hora extra por persona — Worker hours pivot (worker × date)
# ===========================================================================


@router.get("/tarjas/resumen-horas", response_class=HTMLResponse)
async def tarjas_resumen_horas_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_resumen_horas.html")


@router.get("/api/tarjas/resumen-horas/filters")
async def get_tarjas_resumen_horas_filters():
    """Distinct trabajador + tipo_pago + contratista for the hours report."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT trabajador, COALESCE(SUM(horas_extras), 0)::numeric AS total "
                "FROM appsheet.tarjas_pagos "
                "WHERE trabajador IS NOT NULL "
                "GROUP BY trabajador ORDER BY total DESC"
            )
            trabajadores = _rows_to_dicts(cur)

            cur.execute(
                "SELECT DISTINCT tipo_pago FROM appsheet.tarjas_pagos "
                "WHERE tipo_pago IS NOT NULL ORDER BY tipo_pago"
            )
            tipos_pago = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(cur, "appsheet.tarjas_pagos")
    finally:
        conn.close()

    return {
        "trabajadores": trabajadores,
        "tipos_pago": tipos_pago,
        "contratistas": contratistas,
        "empresas": empresas,
    }


@router.get("/api/tarjas/resumen-horas")
async def get_tarjas_resumen_horas(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    """Return worker hours grouped by date for the pivot table."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]

    if trabajador:
        filters.append("trabajador = %s")
        params.append(trabajador)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))

    where = "WHERE " + " AND ".join(filters)

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT trabajador, tipo_pago, fecha::date::text AS fecha, "
                f"COALESCE(SUM(horas_extras), 0)::numeric AS horas_trabajadas, "
                f"COALESCE(SUM(total_hora_extra), 0)::numeric AS monto_hora_extra "
                f"FROM appsheet.tarjas_pagos {where} "
                "GROUP BY trabajador, tipo_pago, fecha::date "
                "ORDER BY trabajador, tipo_pago, fecha::date",
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    # Only show workers with a non-zero horas_extras total across the whole
    # period — a row-by-row filter would also drop the zero days of a
    # worker who did have overtime on other dates, which must stay visible.
    totals_by_worker: dict = {}
    for r in rows:
        totals_by_worker[r["trabajador"]] = totals_by_worker.get(
            r["trabajador"], 0
        ) + (r["horas_trabajadas"] or 0)
    rows = [r for r in rows if totals_by_worker.get(r["trabajador"], 0) > 0]

    return {
        "rows": rows,
        "count": len(rows),
        "resumen": {
            "total_horas": sum(float(r["horas_trabajadas"] or 0) for r in rows),
            "total_trabajadores": len({r["trabajador"] for r in rows}),
            "total_monto": sum(float(r["monto_hora_extra"] or 0) for r in rows),
        },
    }


# ===========================================================================
# Resumen por persona tractorista — Worker × date pivot (total_tractor)
# ===========================================================================


@router.get("/tarjas/resumen-persona-tractorista", response_class=HTMLResponse)
async def tarjas_resumen_persona_tractorista_page(request: Request):
    return _templates.TemplateResponse(
        request, "tarjas_resumen_persona_tractorista.html"
    )


@router.get("/api/tarjas/resumen-persona-tractorista/filters")
async def get_tarjas_resumen_persona_tractorista_filters():
    """Filter options for the tractorista worker pivot."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            maq_col = _resolve_maquina_column(cur)

            cur.execute(
                "SELECT trabajador, COALESCE(SUM(total_tractor), 0) AS total "
                "FROM appsheet.tarjas_pagos "
                f"WHERE {_TRACTORISTA_PAGOS_SQL} AND trabajador IS NOT NULL "
                "GROUP BY trabajador ORDER BY total DESC"
            )
            trabajadores = _rows_to_dicts(cur)

            cur.execute(
                "SELECT DISTINCT tipo_pago FROM appsheet.tarjas_pagos "
                f"WHERE {_TRACTORISTA_PAGOS_SQL} AND tipo_pago IS NOT NULL ORDER BY tipo_pago"
            )
            tipos_pago = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                f"WHERE {_TRACTORISTA_PAGOS_SQL} AND contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]

            maquinas = []
            if maq_col:
                cur.execute(
                    psql.SQL(
                        "SELECT DISTINCT {col} FROM appsheet.tarjas_pagos WHERE "
                        + _TRACTORISTA_PAGOS_SQL
                        + " AND {col} IS NOT NULL ORDER BY {col}"
                    ).format(col=psql.Identifier(maq_col))
                )
                maquinas = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(
                cur,
                "appsheet.tarjas_pagos",
                extra_where="LOWER(TRIM(tipo_pago)) = 'tractorista'",
            )
    finally:
        conn.close()

    return {
        "trabajadores": trabajadores,
        "tipos_pago": tipos_pago,
        "contratistas": contratistas,
        "empresas": empresas,
        "maquinas": maquinas,
        "maquina_column": maq_col,
    }


@router.get("/api/tarjas/resumen-persona-tractorista")
async def get_tarjas_resumen_persona_tractorista(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    maquina: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    """Worker × date pivot rows using total_tractor (tractorista only)."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    try:
        with conn.cursor() as cur:
            where, params = _resumen_persona_tractorista_where(
                cur,
                fecha_inicio,
                fecha_termino,
                trabajador=trabajador,
                tipo_pago=tipo_pago,
                maquina=maquina,
                contratista=contratista,
                empresa=empresa,
            )
            rows = _fetch_resumen_persona_tractorista_rows(cur, where, params)
    finally:
        conn.close()

    return {
        "rows": rows,
        "count": len(rows),
    }


# ===========================================================================
# General Tractorista — Labor averages, person ranking, daily chart (tractorista only)
# ===========================================================================


@router.get("/tarjas/general-tractorista", response_class=HTMLResponse)
async def tarjas_general_tractorista_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_general_tractorista.html")


@router.get("/api/tarjas/general-tractorista/filters")
async def get_tarjas_general_tractorista_filters():
    """Distinct filter values for General Tractorista (tractorista rows only)."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            maq_col = _resolve_maquina_column(cur)
            base = f" FROM appsheet.tarjas_pagos WHERE {_TRACTORISTA_PAGOS_SQL} "

            cur.execute(
                "SELECT DISTINCT cuartel_cc "
                + base
                + "AND cuartel_cc IS NOT NULL ORDER BY cuartel_cc"
            )
            centros_costo = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT labor " + base + "AND labor IS NOT NULL ORDER BY labor"
            )
            labores = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT contratista "
                + base
                + "AND contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]

            maquinas = []
            if maq_col:
                cur.execute(
                    psql.SQL(
                        "SELECT DISTINCT {col} FROM appsheet.tarjas_pagos WHERE "
                        + _TRACTORISTA_PAGOS_SQL
                        + " AND {col} IS NOT NULL ORDER BY {col}"
                    ).format(col=psql.Identifier(maq_col))
                )
                maquinas = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(
                cur,
                "appsheet.tarjas_pagos",
                extra_where="LOWER(TRIM(tipo_pago)) = 'tractorista'",
            )
    finally:
        conn.close()

    return {
        "centros_costo": centros_costo,
        "labores": labores,
        "contratistas": contratistas,
        "empresas": empresas,
        "maquinas": maquinas,
        "maquina_column": maq_col,
    }


@router.get("/api/tarjas/general-tractorista")
async def get_tarjas_general_tractorista(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    centro_costo: str = Query(None),
    labor: str = Query(None),
    maquina: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    try:
        with conn.cursor() as cur:
            where, params = _general_tractorista_where(
                cur,
                fecha_inicio,
                fecha_termino,
                centro_costo=centro_costo,
                labor=labor,
                maquina=maquina,
                contratista=contratista,
                empresa=empresa,
            )
            labor_summary, person_ranking = _fetch_general_tractorista_tables(
                cur, where, params
            )
            maq_col = _resolve_maquina_column(cur)

            # 3) Daily average by labor × worker (top 6) — same filters as ranking
            p_filters = [
                "p.fecha::date BETWEEN %s AND %s",
                "(LOWER(TRIM(p.tipo_pago)) = 'tractorista')",
            ]
            p_params: list = [fecha_inicio, fecha_termino]
            if centro_costo:
                p_filters.append("p.cuartel_cc = %s")
                p_params.append(centro_costo)
            if labor:
                p_filters.append("p.labor = %s")
                p_params.append(labor)
            if contratista:
                p_filters.append("p.contratista = %s")
                p_params.append(contratista)
            if empresa:
                p_filters.append("p.nombre_campo = %s")
                p_params.append(_empresa_to_campo(empresa))
            if maq_col and maquina:
                frag = psql.SQL("p.{} = %s").format(psql.Identifier(maq_col))
                p_filters.append(frag.as_string(conn))
                p_params.append(maquina)
            p_where = "WHERE " + " AND ".join(p_filters)

            cur.execute(
                f"""
                WITH top_workers AS (
                    SELECT trabajador
                    FROM appsheet.tarjas_pagos
                    {where}
                    GROUP BY trabajador
                    ORDER BY SUM(total_tractor) DESC
                    LIMIT 6
                )
                SELECT
                    p.labor,
                    p.trabajador,
                    ROUND(AVG(p.total_tractor)::numeric, 2) AS avg_daily
                FROM appsheet.tarjas_pagos p
                JOIN top_workers tw ON tw.trabajador = p.trabajador
                {p_where}
                GROUP BY p.labor, p.trabajador
                ORDER BY p.labor, avg_daily DESC
            """,
                params + p_params,
            )
            chart_data = _rows_to_dicts(cur)
    finally:
        conn.close()

    return {
        "labor_summary": labor_summary,
        "person_ranking": person_ranking,
        "chart_data": chart_data,
    }


# ===========================================================================
# Excel download helpers
# ===========================================================================


def _excel_response(wb: "openpyxl.Workbook", filename: str) -> StreamingResponse:
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _excel_header_style():
    from openpyxl.styles import Font, PatternFill, Alignment

    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center")
    return fill, font, align


def _apply_header(ws, headers):
    fill, font, align = _excel_header_style()
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font
        c.alignment = align


@router.get("/api/tarjas/general/download-excel")
async def download_tarjas_general_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    where, params = _build_pagos_where(
        fecha_inicio,
        fecha_termino,
        centro_costo,
        tipo_pago,
        labor,
        contratista=contratista,
        nombre_campo=_empresa_to_campo(empresa),
    )
    try:
        _horas_expr = "NULLIF(SUM(horas_trabajadas), 0)"
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT trabajador, contratista, labor, tipo_pago,
                       ROUND(AVG(total_trabajado)::numeric, 0)                  AS promedio_diario,
                       ROUND(SUM(total_trabajado)::numeric / {_horas_expr}, 0)  AS ganancia_hora,
                       ROUND(SUM(total_pagar)::numeric    / {_horas_expr}, 0)   AS costo_hora,
                       COALESCE(SUM(total_trabajado),0) AS total
                FROM appsheet.tarjas_pagos {where}
                GROUP BY trabajador, contratista, labor, tipo_pago
                ORDER BY total DESC
            """,
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General"
    _apply_header(
        ws,
        [
            "Trabajador",
            "Contratista",
            "Labor",
            "Tipo de pago",
            "Promedio diario",
            "Ganancia por hora",
            "Costo por hora",
            "Total",
        ],
    )
    money = "#,##0"
    for i, r in enumerate(rows, 2):
        ws.cell(i, 1, r["trabajador"])
        ws.cell(i, 2, r["contratista"])
        ws.cell(i, 3, r["labor"])
        ws.cell(i, 4, r["tipo_pago"])
        c5 = ws.cell(i, 5, float(r["promedio_diario"] or 0))
        c5.number_format = money
        c6 = ws.cell(
            i,
            6,
            float(r["ganancia_hora"] or 0) if r["ganancia_hora"] is not None else None,
        )
        c6.number_format = money
        c7 = ws.cell(
            i, 7, float(r["costo_hora"] or 0) if r["costo_hora"] is not None else None
        )
        c7.number_format = money
        c8 = ws.cell(i, 8, float(r["total"] or 0))
        c8.number_format = money
    for col, w in zip("ABCDEFGH", [28, 24, 28, 16, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, f"tarjas_general_{fecha_inicio}_{fecha_termino}.xlsx")


@router.get("/api/tarjas/detalle/download-excel")
async def download_tarjas_detalle_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
    campo: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    where, params = _build_detalle_filters(
        fecha_inicio,
        fecha_termino,
        contratista,
        empresa,
        centro_costo,
        tipo_pago,
        labor,
        campo,
    )
    try:
        with conn.cursor() as cur:
            rows = _query_detalle_rows(cur, where, params)
    finally:
        conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle"
    _apply_header(
        ws,
        [
            "Tipo de pago",
            "Labor",
            "CC",
            "Nombre CC",
            "Costo por hora",
            "Jornadas",
            "Total unitario",
            "Costo total",
            "Total trabajado",
            "% Tipo pago",
            "Campo",
        ],
    )
    money = "#,##0"
    for i, r in enumerate(rows, 2):
        ws.cell(i, 1, r["tipo_pago"])
        ws.cell(i, 2, r["labor"])
        ws.cell(i, 3, r["centro_costo"])
        ws.cell(i, 4, r["centro_costo_nombre"] or "—")
        c5 = ws.cell(
            i, 5, float(r["costo_hora"]) if r["costo_hora"] is not None else None
        )
        c5.number_format = money
        ws.cell(i, 6, r["jornadas"])
        c7 = ws.cell(i, 7, float(r["total_unitario"] or 0))
        c7.number_format = money
        c8 = ws.cell(i, 8, float(r["costo_total"] or 0))
        c8.number_format = money
        c9 = ws.cell(i, 9, float(r["total_trabajado"] or 0))
        c9.number_format = money
        ws.cell(i, 10, float(r["pct_pago"] or 0))
        ws.cell(i, 11, r["nombre_campo"])
    last = len(rows) + 2
    ws.cell(last, 1, "Total")
    ws.cell(last, 6, sum(float(r["jornadas"] or 0) for r in rows))
    c8t = ws.cell(last, 8, sum(float(r["costo_total"] or 0) for r in rows))
    c8t.number_format = money
    c9t = ws.cell(last, 9, sum(float(r.get("total_trabajado") or 0) for r in rows))
    c9t.number_format = money
    ws.cell(last, 10, 100 if any(float(r["costo_total"] or 0) for r in rows) else None)
    for col in range(1, 12):
        ws.cell(last, col).font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJK", [14, 28, 10, 28, 14, 10, 14, 14, 16, 12, 20]):
        ws.column_dimensions[col].width = w
    return _excel_response(wb, f"tarjas_detalle_{fecha_inicio}_{fecha_termino}.xlsx")


@router.get("/api/tarjas/contratista/download-excel")
async def download_tarjas_contratista_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if centro_costo:
        filters.append("cuartel_cc = %s")
        params.append(centro_costo)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if labor:
        filters.append("labor = %s")
        params.append(labor)
    where = "WHERE " + " AND ".join(filters)
    _horas_sum = "COALESCE(SUM(horas_trabajadas), 0)"
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT trabajador, contratista, labor, tipo_pago,
                       fecha::date::text AS fecha,
                       COALESCE(SUM(total_trabajado), 0) AS total,
                       {_horas_sum} AS horas
                FROM appsheet.tarjas_pagos {where}
                GROUP BY trabajador, contratista, labor, tipo_pago, fecha::date
                ORDER BY contratista, trabajador, labor, fecha::date
                """,
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    # Build pivot in Python
    from collections import OrderedDict
    dates = sorted({r["fecha"] for r in rows})
    groups: dict = OrderedDict()
    for r in rows:
        key = (r["trabajador"], r["contratista"], r["labor"], r["tipo_pago"])
        if key not in groups:
            groups[key] = {"total_ganado": 0.0, "total_horas": 0.0, "by_date": {}}
        g = groups[key]
        g["total_ganado"] += float(r["total"] or 0)
        g["total_horas"] += float(r["horas"] or 0)
        g["by_date"][r["fecha"]] = float(r["total"] or 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Por contratista"
    fixed_headers = ["Trabajador", "Contratista", "Labor", "Tipo de pago", "Costo/hr"]
    date_headers = [d[5:] for d in dates]  # MM-DD
    _apply_header(ws, fixed_headers + date_headers + ["Total"])

    fill_hdr, font_hdr, align_hdr = _excel_header_style()
    total_col_idx = len(fixed_headers) + len(dates) + 1

    # Style total column header
    ws.cell(1, total_col_idx).fill = fill_hdr
    ws.cell(1, total_col_idx).font = font_hdr
    ws.cell(1, total_col_idx).alignment = align_hdr

    money = "#,##0"
    for i, ((trabajador, cont, lab, tipo), g) in enumerate(groups.items(), 2):
        costo_hora = round(g["total_ganado"] / g["total_horas"]) if g["total_horas"] > 0 else None
        ws.cell(i, 1, trabajador)
        ws.cell(i, 2, cont)
        ws.cell(i, 3, lab)
        ws.cell(i, 4, tipo)
        c = ws.cell(i, 5, costo_hora)
        if costo_hora is not None:
            c.number_format = money
        for j, d in enumerate(dates, 6):
            val = g["by_date"].get(d)
            if val:
                dc = ws.cell(i, j, val)
                dc.number_format = money
        tc = ws.cell(i, total_col_idx, g["total_ganado"])
        tc.number_format = money
        from openpyxl.styles import Font
        tc.font = Font(bold=True)

    fixed_widths = [28, 22, 28, 14, 12]
    date_widths = [10] * len(dates)
    for col_letter, w in zip(
        [chr(65 + i) for i in range(len(fixed_widths) + len(date_widths) + 1)],
        fixed_widths + date_widths + [12],
    ):
        ws.column_dimensions[col_letter].width = w

    return _excel_response(
        wb, f"tarjas_contratista_{fecha_inicio}_{fecha_termino}.xlsx"
    )


@router.get("/api/tarjas/detalle-tractorista/download-excel")
async def download_tarjas_detalle_tractorista_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    labor: str = Query(None),
    campo: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    where, params = _detalle_tractorista_where(
        fecha_inicio,
        fecha_termino,
        contratista=contratista,
        empresa=empresa,
        centro_costo=centro_costo,
        labor=labor,
        campo=campo,
    )
    try:
        with conn.cursor() as cur:
            rows = _fetch_detalle_tractorista_rows(cur, where, params)
    finally:
        conn.close()
    pivots = _build_detalle_tractorista_pivots(rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle tractorista"
    row = 1
    for pivot in pivots:
        row = _write_detalle_tractorista_excel_pivot(ws, pivot, row)
    if not pivots:
        ws.cell(1, 1, "Sin resultados")
    return _excel_response(
        wb, f"tarjas_detalle_tractorista_{fecha_inicio}_{fecha_termino}.xlsx"
    )


def _write_detalle_tractorista_excel_pivot(ws, pivot: dict, start_row: int) -> int:
    """Write one Looker nested pivot block. Returns the next free row."""
    from openpyxl.utils import get_column_letter

    fill, font, _align = _excel_header_style()
    wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    n = len(pivot["columns"])
    if n == 0:
        return start_row
    total_col = 2 + n
    money = "$#,##0"

    def _paint(cell, value=None):
        if value is not None:
            cell.value = value
        cell.fill = fill
        cell.font = font
        cell.alignment = wrap

    _paint(ws.cell(start_row, 1), "Fecha")
    ws.merge_cells(
        start_row=start_row, start_column=2, end_row=start_row, end_column=1 + n
    )
    _paint(ws.cell(start_row, 2), pivot["contratista"])
    for col in range(3, 2 + n):
        _paint(ws.cell(start_row, col))
    _paint(ws.cell(start_row, total_col), f"Total {pivot['contratista']}")

    ws.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row + 2, end_column=1
    )
    ws.merge_cells(
        start_row=start_row,
        start_column=total_col,
        end_row=start_row + 2,
        end_column=total_col,
    )

    col = 2
    for wk in pivot["workers"]:
        span = len(wk["labores"])
        end_col = col + span - 1
        if span > 1:
            ws.merge_cells(
                start_row=start_row + 1,
                start_column=col,
                end_row=start_row + 1,
                end_column=end_col,
            )
        _paint(ws.cell(start_row + 1, col), wk["name"])
        for i in range(span):
            _paint(ws.cell(start_row + 1, col + i))
        col += span

    for i, c in enumerate(pivot["columns"]):
        _paint(ws.cell(start_row + 2, 2 + i), c["labor"])

    for r in range(start_row, start_row + 3):
        _paint(ws.cell(r, 1))
        _paint(ws.cell(r, total_col))

    right = Alignment(horizontal="right")
    bold = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="E8EEF4")
    row = start_row + 3
    for fecha in pivot["dates"]:
        ws.cell(row, 1, _fmt_date_slash(fecha))
        for i, c in enumerate(pivot["columns"]):
            val = pivot["matrix"][fecha].get(c["key"])
            if val is not None:
                cell = ws.cell(row, 2 + i, int(round(val)))
                cell.number_format = money
                cell.alignment = right
        tot = ws.cell(row, total_col, int(round(pivot["date_totals"][fecha])))
        tot.number_format = money
        tot.font = bold
        tot.alignment = right
        row += 1

    ws.cell(row, 1, "Suma total").font = bold
    ws.cell(row, 1).fill = total_fill
    for i, c in enumerate(pivot["columns"]):
        cell = ws.cell(row, 2 + i, int(round(pivot["col_totals"][c["key"]])))
        cell.number_format = money
        cell.font = bold
        cell.fill = total_fill
        cell.alignment = right
    grand = ws.cell(row, total_col, int(round(pivot["grand_total"])))
    grand.number_format = money
    grand.font = bold
    grand.fill = total_fill
    grand.alignment = right

    ws.column_dimensions["A"].width = 14
    for i in range(n):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16
    ws.column_dimensions[get_column_letter(total_col)].width = max(
        18, min(36, len(f"Total {pivot['contratista']}") + 2)
    )
    return row + 2


def _detalle_tractorista_pivot_table_html(pivot: dict) -> str:
    """One nested pivot HTML table for the PDF (no rowspan — xhtml2pdf)."""
    n = len(pivot["columns"])
    if n == 0:
        return ""
    _check_pivot_date_range(pivot["columns"], "Detalle tractorista")
    w = _pivot_col_widths({"fecha": 10, "total": 12}, n, date_pct=6.0)
    contratista = _escape_html(pivot["contratista"])
    group = 'class="th-group" style="text-align:center;background:#1F4E79;color:#ffffff"'
    worker_ths = "".join(
        f'<th {group} colspan="{len(wk["labores"])}">{_escape_html(wk["name"])}</th>'
        for wk in pivot["workers"]
    )
    labor_ths = "".join(
        f'<th class="num" style="{w["date"]};background:#1F4E79;color:#ffffff">'
        f'{_escape_html(c["labor"])}</th>'
        for c in pivot["columns"]
    )

    def money(v) -> str:
        return _fmt_clp(v) if v is not None else ""

    body = ""
    for d in pivot["dates"]:
        cells = "".join(
            f'<td class="num" style="{w["date"]}">{money(pivot["matrix"][d].get(c["key"]))}</td>'
            for c in pivot["columns"]
        )
        body += (
            f'<tr><td style="{w["fecha"]}">{_fmt_date_slash(d)}</td>{cells}'
            f'<td class="num" style="{w["total"]}">{_fmt_clp(pivot["date_totals"][d])}</td></tr>'
        )
    foot = "".join(
        f'<td class="num" style="{w["date"]}">{_fmt_clp(pivot["col_totals"][c["key"]])}</td>'
        for c in pivot["columns"]
    )
    return f"""
    <table class="detalle-tract-pivot pivot-wide" style="{w['table']}">
      <thead>
        <tr>
          <th style="{w["fecha"]}"></th>
          <th {group} colspan="{n}">{contratista}</th>
          <th style="{w["total"]}"></th>
        </tr>
        <tr>
          <th style="{w["fecha"]}"></th>
          {worker_ths}
          <th style="{w["total"]}"></th>
        </tr>
        <tr>
          <th style="{w["fecha"]};background:#1F4E79;color:#ffffff">Fecha</th>
          {labor_ths}
          <th class="num" style="{w["total"]};background:#1F4E79;color:#ffffff">Total {contratista}</th>
        </tr>
      </thead>
      <tbody>{body}</tbody>
      <tfoot>
        <tr>
          <td style="{w["fecha"]}">Suma total</td>
          {foot}
          <td class="num" style="{w["total"]}">{_fmt_clp(pivot["grand_total"])}</td>
        </tr>
      </tfoot>
    </table>
    """


def _build_detalle_tractorista_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    contratista: str | None = None,
    empresa: str | None = None,
    centro_costo: str | None = None,
    labor: str | None = None,
    campo: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical."""
    where, params = _detalle_tractorista_where(
        fecha_inicio,
        fecha_termino,
        contratista=contratista,
        empresa=empresa,
        centro_costo=centro_costo,
        labor=labor,
        campo=campo,
    )
    rows = _fetch_detalle_tractorista_rows(cur, where, params)
    pivots = _build_detalle_tractorista_pivots(rows)
    tables = (
        "".join(_detalle_tractorista_pivot_table_html(p) for p in pivots)
        or "<p>Sin resultados</p>"
    )
    header = _pdf_header(
        _pdf_title("Detalle Tractoristas", contratista),
        fecha_inicio,
        fecha_termino,
        {
            "Empresa": empresa,
            "Contratista": contratista,
            "CC": centro_costo,
            "Labor": labor,
        },
    )
    return f"""
    {header}
    {tables}
    """


@router.get("/api/tarjas/detalle-tractorista/download-pdf")
async def download_tarjas_detalle_tractorista_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    labor: str = Query(None),
    campo: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            body = _build_detalle_tractorista_html(
                cur,
                fecha_inicio,
                fecha_termino,
                contratista,
                empresa,
                centro_costo,
                labor,
                campo,
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}</style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(
        html, f"detalle_tractorista_{fecha_inicio}_{fecha_termino}.pdf"
    )


@router.get("/api/tarjas/general-tractorista/download-excel")
async def download_tarjas_general_tractorista_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    centro_costo: str = Query(None),
    labor: str = Query(None),
    maquina: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            where, params = _general_tractorista_where(
                cur,
                fecha_inicio,
                fecha_termino,
                centro_costo=centro_costo,
                labor=labor,
                maquina=maquina,
                contratista=contratista,
                empresa=empresa,
            )
            labor_summary, person_ranking = _fetch_general_tractorista_tables(
                cur, where, params
            )
    finally:
        conn.close()
    wb = openpyxl.Workbook()
    money = "#,##0"
    money2 = "#,##0.00"
    ws = wb.active
    ws.title = "Por labor"
    _apply_header(ws, ["Labor", "Promedio", "Total"])
    for i, r in enumerate(labor_summary, 2):
        ws.cell(i, 1, r["labor"])
        c2 = ws.cell(i, 2, float(r["avg_rate"] or 0))
        c2.number_format = money2
        c3 = ws.cell(i, 3, float(r["total"] or 0))
        c3.number_format = money
    foot = 2 + len(labor_summary)
    ws.cell(foot, 1, "Suma total")
    tot = ws.cell(
        foot, 3, sum(float(r["total"] or 0) for r in labor_summary)
    )
    tot.number_format = money
    tot.font = Font(bold=True)
    for col, w in zip("ABC", [32, 14, 14]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Ranking")
    _apply_header(ws2, ["Trabajador", "Contratista", "Promedio", "Total"])
    for i, r in enumerate(person_ranking, 2):
        ws2.cell(i, 1, r["trabajador"])
        ws2.cell(i, 2, r["contratista"])
        c3 = ws2.cell(i, 3, float(r["avg_rate"] or 0))
        c3.number_format = money2
        c4 = ws2.cell(i, 4, float(r["total"] or 0))
        c4.number_format = money
    foot = 2 + len(person_ranking)
    ws2.cell(foot, 1, "Suma total")
    tot = ws2.cell(
        foot, 4, sum(float(r["total"] or 0) for r in person_ranking)
    )
    tot.number_format = money
    tot.font = Font(bold=True)
    for col, w in zip("ABCD", [28, 28, 14, 14]):
        ws2.column_dimensions[col].width = w
    return _excel_response(
        wb, f"tarjas_general_tractorista_{fecha_inicio}_{fecha_termino}.xlsx"
    )


def _build_general_tractorista_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    centro_costo: str | None = None,
    labor: str | None = None,
    contratista: str | None = None,
    empresa: str | None = None,
    maquina: str | None = None,
) -> str:
    """Same two tables as the screen (labor + ranking). Shared by standalone
    PDF and bulk /reportes (issue #116)."""
    where, params = _general_tractorista_where(
        cur,
        fecha_inicio,
        fecha_termino,
        centro_costo=centro_costo,
        labor=labor,
        maquina=maquina,
        contratista=contratista,
        empresa=empresa,
    )
    labor_summary, person_ranking = _fetch_general_tractorista_tables(
        cur, where, params
    )

    def money(v, decimals=False):
        if v is None:
            return "—"
        if decimals:
            n = f"{float(v):,.2f}"
            return "$" + n.replace(",", "X").replace(".", ",").replace("X", ".")
        return _fmt_clp(v)

    labor_rows = "".join(
        f'<tr><td>{_escape_html(r["labor"] or "")}</td>'
        f'<td class="num">{money(r["avg_rate"], True)}</td>'
        f'<td class="total">{money(r["total"])}</td></tr>'
        for r in labor_summary
    )
    labor_sum = sum(float(r["total"] or 0) for r in labor_summary)
    rank_rows = "".join(
        f'<tr><td>{_escape_html(r["trabajador"] or "")}</td>'
        f'<td>{_escape_html(r["contratista"] or "")}</td>'
        f'<td class="num">{money(r["avg_rate"], True)}</td>'
        f'<td class="total">{money(r["total"])}</td></tr>'
        for r in person_ranking
    )
    rank_sum = sum(float(r["total"] or 0) for r in person_ranking)
    header = _pdf_header(
        _pdf_title("General Tractoristas", contratista),
        fecha_inicio,
        fecha_termino,
        {
            "Empresa": empresa,
            "Contratista": contratista,
            "CC": centro_costo,
            "Labor": labor,
            "Máquina": maquina,
        },
    )
    return f"""
    {header}
    <h2 style="font-size:12pt;margin:10px 0 6px">Ganancia promedio por labor</h2>
    <table style="width:70%;table-layout:fixed"><thead>
      <tr><th style="width:50%">Labor</th>
          <th class="num" style="width:25%">Promedio</th>
          <th class="num" style="width:25%">Total</th></tr>
    </thead><tbody>{labor_rows}</tbody>
    <tfoot><tr><td>Suma total</td><td></td>
      <td class="total">{money(labor_sum)}</td></tr></tfoot></table>
    <h2 style="font-size:12pt;margin:16px 0 6px">Ranking por persona</h2>
    <table style="width:85%;table-layout:fixed"><thead>
      <tr><th style="width:32%">Trabajador</th><th style="width:32%">Contratista</th>
          <th class="num" style="width:18%">Promedio</th>
          <th class="num" style="width:18%">Total</th></tr>
    </thead><tbody>{rank_rows}</tbody>
    <tfoot><tr><td colspan="3">Suma total</td>
      <td class="total">{money(rank_sum)}</td></tr></tfoot></table>
    """


@router.get("/api/tarjas/general-tractorista/download-pdf")
async def download_tarjas_general_tractorista_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    centro_costo: str = Query(None),
    labor: str = Query(None),
    maquina: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            body = _build_general_tractorista_html(
                cur,
                fecha_inicio,
                fecha_termino,
                centro_costo,
                labor,
                contratista,
                empresa,
                maquina,
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}</style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(
        html, f"general_tractorista_{fecha_inicio}_{fecha_termino}.pdf"
    )


@router.get("/api/tarjas/resumen-horas/download-excel")
async def download_tarjas_resumen_horas_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if trabajador:
        filters.append("trabajador = %s")
        params.append(trabajador)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    where = "WHERE " + " AND ".join(filters)
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT trabajador, contratista, tipo_pago, fecha::date::text AS fecha, "
                f"COALESCE(SUM(horas_extras), 0)::numeric AS horas "
                f"FROM appsheet.tarjas_pagos {where} "
                "GROUP BY trabajador, contratista, tipo_pago, fecha::date "
                "ORDER BY trabajador, tipo_pago, fecha::date",
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()
    # Build pivot: worker → tipo_pago → date → hours
    dates = sorted({r["fecha"] for r in rows})
    workers: dict = {}
    for r in rows:
        key = (r["trabajador"] or "", r["contratista"] or "", r["tipo_pago"] or "")
        if key not in workers:
            workers[key] = {"by_date": {}, "total": 0}
        workers[key]["by_date"][r["fecha"]] = workers[key]["by_date"].get(
            r["fecha"], 0
        ) + (r["horas"] or 0)
        workers[key]["total"] += r["horas"] or 0

    # Only keep workers whose horas_extras total for the whole period (across
    # every contratista/tipo_pago combination) is greater than zero.
    totals_by_worker: dict = {}
    for (trab, _cont, _tipo), entry in workers.items():
        totals_by_worker[trab] = totals_by_worker.get(trab, 0) + entry["total"]
    workers = {
        k: v for k, v in workers.items() if totals_by_worker.get(k[0], 0) > 0
    }

    sorted_workers = sorted(workers.items(), key=lambda x: -x[1]["total"])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horas extra"
    headers = (
        ["Trabajador", "Contratista", "Tipo de pago"]
        + [datetime.date.fromisoformat(d).strftime("%d/%m/%Y") for d in dates]
        + ["Total"]
    )
    _apply_header(ws, headers)
    from openpyxl.styles import PatternFill, Font

    total_fill = PatternFill("solid", fgColor="D6E4F0")
    for row_num, ((trab, cont, tipo), entry) in enumerate(sorted_workers, 2):
        ws.cell(row_num, 1, trab)
        ws.cell(row_num, 2, cont)
        ws.cell(row_num, 3, tipo)
        for col, d in enumerate(dates, 4):
            v = entry["by_date"].get(d, 0)
            ws.cell(row_num, col, v if v else None)
        tc = ws.cell(row_num, 4 + len(dates), entry["total"])
        tc.fill = total_fill
        tc.font = Font(bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    for i in range(len(dates) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(4 + i)].width = 12
    return _excel_response(wb, f"tarjas_horas_{fecha_inicio}_{fecha_termino}.xlsx")


@router.get("/api/tarjas/resumen-persona-tractorista/download-excel")
async def download_tarjas_resumen_persona_tractorista_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    maquina: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            where, params = _resumen_persona_tractorista_where(
                cur,
                fecha_inicio,
                fecha_termino,
                trabajador=trabajador,
                tipo_pago=tipo_pago,
                maquina=maquina,
                contratista=contratista,
                empresa=empresa,
            )
            rows = _fetch_resumen_persona_tractorista_rows(cur, where, params)
    finally:
        conn.close()
    dates, groups = _pivot_resumen_persona_tractorista(rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tractorista"
    headers = (
        ["Trabajador", "Contratista", "Máquina", "Horas extra"]
        + [datetime.date.fromisoformat(d).strftime("%d/%m/%Y") for d in dates]
        + ["Total"]
    )
    _apply_header(ws, headers)
    total_fill = PatternFill("solid", fgColor="D6E4F0")
    money = "#,##0"
    date_totals = {d: 0.0 for d in dates}
    overall = 0.0
    row_num = 2
    for g in groups:
        for entry in g["row_list"]:
            ws.cell(row_num, 1, g["trabajador"])
            ws.cell(row_num, 2, g["contratista"] or None)
            ws.cell(row_num, 3, entry["maquina"] or None)
            ws.cell(row_num, 4, entry["horas_extras"] if entry["horas_extras"] != "" else None)
            for col, d in enumerate(dates, 5):
                v = entry["by_date"].get(d, 0)
                c = ws.cell(row_num, col, v if v else None)
                if v:
                    c.number_format = money
                    date_totals[d] += v
            tc = ws.cell(row_num, 5 + len(dates), entry["total"])
            tc.number_format = money
            tc.fill = total_fill
            tc.font = Font(bold=True)
            overall += entry["total"]
            row_num += 1
    ws.cell(row_num, 1, "Suma total")
    ws.cell(row_num, 1).font = Font(bold=True)
    for col, d in enumerate(dates, 5):
        c = ws.cell(row_num, col, date_totals[d] if date_totals[d] else None)
        if date_totals[d]:
            c.number_format = money
        c.font = Font(bold=True)
        c.fill = total_fill
    tot = ws.cell(row_num, 5 + len(dates), overall)
    tot.number_format = money
    tot.font = Font(bold=True)
    tot.fill = total_fill
    for col, w in zip("ABCD", [28, 24, 18, 14]):
        ws.column_dimensions[col].width = w
    for i in range(len(dates) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(5 + i)].width = 14
    return _excel_response(
        wb, f"tarjas_tractorista_{fecha_inicio}_{fecha_termino}.xlsx"
    )


def _build_resumen_persona_tractorista_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    trabajador: str | None = None,
    tipo_pago: str | None = None,
    contratista: str | None = None,
    empresa: str | None = None,
    maquina: str | None = None,
) -> str:
    """Same pivot as the screen (Trabajador, Contratista, Máquina, Horas extra,
    dates DD/MM/YYYY, Total). Shared by standalone PDF and bulk /reportes."""
    where, params = _resumen_persona_tractorista_where(
        cur,
        fecha_inicio,
        fecha_termino,
        trabajador=trabajador,
        tipo_pago=tipo_pago,
        maquina=maquina,
        contratista=contratista,
        empresa=empresa,
    )
    rows = _fetch_resumen_persona_tractorista_rows(cur, where, params)
    dates, groups = _pivot_resumen_persona_tractorista(rows)
    _check_pivot_date_range(dates, "Resumen tractorista")
    w = _pivot_col_widths(
        {"worker": 16, "cont": 16, "maq": 10, "hextra": 8, "total": 10},
        len(dates),
    )
    date_headers = "".join(
        f'<th class="num" style="{w["date"]}">{_fmt_date_slash(d)}</th>'
        for d in dates
    )
    date_totals = {d: 0.0 for d in dates}
    overall = 0.0
    rows_html = ""
    for g in groups:
        first = True
        for entry in g["row_list"]:
            cls = "worker-first" if first else ""
            first = False
            maq = _escape_html(str(entry["maquina"])) if entry["maquina"] not in ("", None) else "—"
            hextra = (
                _escape_html(str(entry["horas_extras"]))
                if entry["horas_extras"] not in ("", None)
                else "—"
            )
            rows_html += (
                f'<tr class="{cls}">'
                f'<td style="{w["worker"]}">{_escape_html(g["trabajador"])}</td>'
                f'<td style="{w["cont"]}">{_escape_html(g["contratista"]) or "—"}</td>'
                f'<td style="{w["maq"]}">{maq}</td>'
                f'<td style="{w["hextra"]}">{hextra}</td>'
            )
            for d in dates:
                v = entry["by_date"].get(d, 0)
                if v:
                    date_totals[d] += v
                    rows_html += (
                        f'<td class="num" style="{w["date"]}">{_fmt_clp(v)}</td>'
                    )
                else:
                    rows_html += f'<td class="num" style="{w["date"]}">0</td>'
            rows_html += (
                f'<td class="total" style="{w["total"]}">{_fmt_clp(entry["total"])}</td></tr>'
            )
            overall += entry["total"]
    foot_dates = "".join(
        f'<td class="total" style="{w["date"]}">'
        f'{_fmt_clp(date_totals[d]) if date_totals[d] else ""}</td>'
        for d in dates
    )
    header = _pdf_header(
        _pdf_title("Resumen Tractorista", contratista),
        fecha_inicio,
        fecha_termino,
        {
            "Empresa": empresa,
            "Contratista": contratista,
            "Trabajador": trabajador,
            "Máquina": maquina,
        },
    )
    return f"""
    {header}
    <table class="pivot-wide" style="{w['table']}"><thead>
      <tr>
        <th style="{w['worker']}">Trabajador</th>
        <th style="{w['cont']}">Contratista</th>
        <th style="{w['maq']}">Máquina</th>
        <th style="{w['hextra']}">Horas extra</th>
        {date_headers}
        <th class="num" style="{w['total']}">Total</th>
      </tr>
    </thead><tbody>{rows_html}</tbody>
    <tfoot><tr>
      <td colspan="4">Suma total</td>
      {foot_dates}
      <td class="total" style="{w['total']}">{_fmt_clp(overall)}</td>
    </tr></tfoot></table>
    """


@router.get("/api/tarjas/resumen-persona-tractorista/download-pdf")
async def download_tarjas_resumen_persona_tractorista_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    maquina: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            body = _build_resumen_persona_tractorista_html(
                cur,
                fecha_inicio,
                fecha_termino,
                trabajador=trabajador,
                tipo_pago=tipo_pago,
                contratista=contratista,
                empresa=empresa,
                maquina=maquina,
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}</style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(
        html, f"resumen_tractorista_{fecha_inicio}_{fecha_termino}.pdf"
    )


# ===========================================================================
# PDF download endpoints
# ===========================================================================



def _build_resumen_persona_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    trabajador: str | None = None,
    tipo_pago: str | None = None,
    contratista: str | None = None,
    empresa: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical."""
    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if trabajador:
        filters.append("trabajador = %s")
        params.append(trabajador)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    where = "WHERE " + " AND ".join(filters)
    cur.execute(
        f"SELECT trabajador, tipo_pago, fecha::date::text AS fecha, "
        f"COALESCE(SUM(total_trabajado),0) AS total "
        f"FROM appsheet.tarjas_pagos {where} "
        "GROUP BY trabajador, tipo_pago, fecha::date ORDER BY trabajador, tipo_pago, fecha::date",
        params,
    )
    rows = _rows_to_dicts(cur)

    dates = sorted({r["fecha"] for r in rows})
    workers: dict = {}
    for r in rows:
        k = (r["trabajador"] or "", r["tipo_pago"] or "")
        if k not in workers:
            workers[k] = {"by_date": {}, "total": 0}
        workers[k]["by_date"][r["fecha"]] = workers[k]["by_date"].get(
            r["fecha"], 0
        ) + float(r["total"] or 0)
        workers[k]["total"] += float(r["total"] or 0)
    # Alphabetical by worker name, keeping each worker's "Al día"/"trato"
    # rows contiguous — required for the is_first "print name once" logic
    # below (issue #137).
    sorted_workers = sorted(
        workers.items(), key=lambda x: ((x[0][0] or "").lower(), x[0][1] or "")
    )

    # Per-date pivot (one column per date), matching the on-screen table
    # exactly (issue #134) — same _pivot_col_widths pattern already used by
    # Detalle Contratistas / Horas Extra / Hora Ponderada 9h.
    _check_pivot_date_range(dates, "Resumen por persona")
    w = _pivot_col_widths({"worker": 18, "tipo": 9, "total": 11}, len(dates))
    date_headers = "".join(
        f'<th class="num" style="{w["date"]}">'
        f'{datetime.date.fromisoformat(d).strftime("%d/%m")}</th>'
        for d in dates
    )
    rows_html = ""
    prev = None
    for (trab, tipo), entry in sorted_workers:
        is_first = prev != trab
        prev = trab
        cls = "worker-first" if is_first else ""
        rows_html += (
            f'<tr class="{cls}">'
            f'<td style="{w["worker"]}">{"<b>" + trab + "</b>" if is_first else ""}</td>'
            f'<td style="{w["tipo"]}">{tipo}</td>'
        )
        for d in dates:
            v = entry["by_date"].get(d, 0)
            rows_html += (
                f'<td class="num" style="{w["date"]}">{_fmt_clp(v) if v else "0"}</td>'
            )
        rows_html += f'<td class="total" style="{w["total"]}">{_fmt_clp(entry["total"])}</td></tr>'

    logo = _logo_b64()
    logo_html = f'<img src="data:image/png;base64,{logo}" style="width:80px;height:auto" />' if logo else ""
    title = _pdf_title("Resumen Por Persona", contratista)
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#1e293b;margin-bottom:12px">
      <tr>
        <td style="padding:10px 16px">{logo_html}</td>
        <td style="padding:10px 16px;color:#ffffff">
          <b style="font-size:14pt">{title}</b><br/>
          <span style="font-size:9pt">Desde: {_fmt_date_slash(fecha_inicio)} &nbsp; Hasta: {_fmt_date_slash(fecha_termino)}</span>
        </td>
      </tr>
    </table>
    <table class="pivot-wide" style="{w['table']}"><thead>
      <tr><th style="{w['worker']}">Trabajador</th><th style="{w['tipo']}">Tipo de pago</th>{date_headers}<th class="num" style="{w['total']}">Total</th></tr>
    </thead><tbody>{rows_html}</tbody></table>
    """


@router.get("/api/tarjas/resumen-persona/download-pdf")
async def download_tarjas_resumen_persona_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    try:
        with conn.cursor() as cur:
            body = _build_resumen_persona_html(
                cur, fecha_inicio, fecha_termino, trabajador, tipo_pago, contratista, empresa
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}</style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(html, f"resumen_persona_{fecha_inicio}_{fecha_termino}.pdf")


def _build_general_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    centro_costo: str | None = None,
    tipo_pago: str | None = None,
    labor: str | None = None,
    contratista: str | None = None,
    empresa: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical."""
    where, params = _build_pagos_where(
        fecha_inicio,
        fecha_termino,
        centro_costo,
        tipo_pago,
        labor,
        contratista=contratista,
        nombre_campo=_empresa_to_campo(empresa),
    )
    _horas_expr = "NULLIF(SUM(horas_trabajadas), 0)"
    _horas_sum = "COALESCE(SUM(horas_trabajadas), 0)"
    cur.execute(
        f"""
        SELECT labor,
               ROUND(AVG(total_trabajado)::numeric, 0)           AS promedio_diario,
               ROUND(SUM(total_trabajado)::numeric / {_horas_expr}, 0) AS ganancia_hora,
               ROUND(({_horas_sum})::numeric, 1)                 AS total_horas,
               COALESCE(SUM(total_trabajado), 0)                 AS total
        FROM appsheet.tarjas_pagos {where}
        GROUP BY labor ORDER BY total DESC
    """,
        params,
    )
    labor_rows = _rows_to_dicts(cur)
    cur.execute(
        f"""
        SELECT trabajador, contratista,
               ROUND(AVG(total_trabajado)::numeric, 0)           AS promedio_diario,
               ROUND(SUM(total_trabajado)::numeric / {_horas_expr}, 0) AS ganancia_hora,
               ROUND(({_horas_sum})::numeric, 1)                 AS total_horas,
               COALESCE(SUM(total_trabajado), 0)                 AS total
        FROM appsheet.tarjas_pagos {where}
        GROUP BY trabajador, contratista ORDER BY total DESC
    """,
        params,
    )
    ranking_rows = _rows_to_dicts(cur)

    fmtCLP = lambda v: f"${float(v):,.0f}".replace(",", ".") if v is not None else "—"
    fmtHrs = lambda v: f"{float(v):,.1f} h".replace(",", ".") if v else "—"

    # Explicit widths on narrower-than-100% tables — Labor/Trabajador/
    # Contratista keep enough room for long text, the money/hour columns
    # don't need nearly as much as auto-layout gave them (issue #132).
    LW = {"labor": "width:39%", "prom": "width:16%", "gan": "width:16%", "horas": "width:12%", "total": "width:17%"}
    RW = {
        "trab": "width:23%",
        "cont": "width:23%",
        "prom": "width:14%",
        "gan": "width:14%",
        "horas": "width:10%",
        "total": "width:16%",
    }
    labor_html = "".join(
        f'<tr><td style="{LW["labor"]}">{r["labor"]}</td>'
        f'<td class="num" style="{LW["prom"]}">{fmtCLP(r["promedio_diario"])}</td>'
        f'<td class="num" style="{LW["gan"]}">{fmtCLP(r["ganancia_hora"])}</td>'
        f'<td class="num" style="{LW["horas"]}">{fmtHrs(r["total_horas"])}</td>'
        f'<td class="total" style="{LW["total"]}">{fmtCLP(r["total"])}</td></tr>'
        for r in labor_rows
    )
    ranking_html = "".join(
        f'<tr><td style="{RW["trab"]}">{r["trabajador"]}</td>'
        f'<td style="{RW["cont"]}">{r["contratista"]}</td>'
        f'<td class="num" style="{RW["prom"]}">{fmtCLP(r["promedio_diario"])}</td>'
        f'<td class="num" style="{RW["gan"]}">{fmtCLP(r["ganancia_hora"])}</td>'
        f'<td class="num" style="{RW["horas"]}">{fmtHrs(r["total_horas"])}</td>'
        f'<td class="total" style="{RW["total"]}">{fmtCLP(r["total"])}</td></tr>'
        for r in ranking_rows
    )
    header = _pdf_header(
        _pdf_title("General Operacional", contratista),
        fecha_inicio,
        fecha_termino,
        {
            "Empresa": empresa,
            "Contratista": contratista,
            "CC": centro_costo,
            "Tipo de pago": tipo_pago,
            "Labor": labor,
        },
    )
    return f"""
    {header}
    <p class="section-title">Ganancia promedio por labor</p>
    <table style="width:80%;table-layout:fixed"><thead>
      <tr><th style="{LW["labor"]}">Labor</th><th class="num" style="{LW["prom"]}">Promedio diario</th>
      <th class="num" style="{LW["gan"]}">Ganancia por hora</th><th class="num" style="{LW["horas"]}">Horas</th>
      <th class="num" style="{LW["total"]}">Total</th></tr>
    </thead><tbody>{labor_html}</tbody></table>
    <p class="section-title">Ranking por persona</p>
    <table style="width:80%;table-layout:fixed"><thead>
      <tr><th style="{RW["trab"]}">Trabajador</th><th style="{RW["cont"]}">Contratista</th>
      <th class="num" style="{RW["prom"]}">Promedio diario</th>
      <th class="num" style="{RW["gan"]}">Ganancia por hora</th>
      <th class="num" style="{RW["horas"]}">Horas</th><th class="num" style="{RW["total"]}">Total</th></tr>
    </thead><tbody>{ranking_html}</tbody></table>
    """


@router.get("/api/tarjas/general/download-pdf")
async def download_tarjas_general_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    try:
        with conn.cursor() as cur:
            body = _build_general_html(
                cur,
                fecha_inicio,
                fecha_termino,
                centro_costo,
                tipo_pago,
                labor,
                contratista,
                empresa,
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}</style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(html, f"general_{fecha_inicio}_{fecha_termino}.pdf")


def _build_detalle_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    contratista: str | None = None,
    centro_costo: str | None = None,
    tipo_pago: str | None = None,
    labor: str | None = None,
    campo: str | None = None,
    empresa: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical."""
    where, params = _build_detalle_filters(
        fecha_inicio,
        fecha_termino,
        contratista,
        empresa,
        centro_costo,
        tipo_pago,
        labor,
        campo,
    )
    resumen = _query_detalle_resumen(cur, where, params)
    rows = _query_detalle_rows(cur, where, params)

    fmtCLP = lambda v: f"${float(v):,.0f}".replace(",", ".") if v else "—"
    fmtPct = lambda v: f"{float(v):.2f} %" if v is not None else "—"
    # Labor/Nombre CC hold free-text descriptions that can run long, so this
    # table keeps close to full width — but the numeric columns (Costo/hora,
    # Jornadas, Total, % pago) had much more room than their short values
    # ever need, before these explicit widths (issue #132).
    DW = {
        "tipo": "width:8%",
        "labor": "width:16%",
        "cc": "width:5%",
        "nombre_cc": "width:15%",
        "costo_hora": "width:9%",
        "jornadas": "width:7%",
        "total": "width:12%",
        "trab": "width:13%",
        "pct": "width:15%",
    }
    rows_html = "".join(
        f'<tr><td style="{DW["tipo"]}"><span class="{_tipo_pago_badge_class(r["tipo_pago"])}">'
        f'{_escape_html(_tipo_pago_label(r["tipo_pago"]))}</span></td>'
        f'<td style="{DW["labor"]}">{r["labor"]}</td><td style="{DW["cc"]}">{r["centro_costo"]}</td>'
        f'<td style="{DW["nombre_cc"]}">{_escape_html(r["centro_costo_nombre"] or "—")}</td>'
        f'<td class="num" style="{DW["costo_hora"]}">{fmtCLP(r["costo_hora"])}</td>'
        f'<td class="num" style="{DW["jornadas"]}">{r["jornadas"]}</td>'
        f'<td class="total" style="{DW["total"]}">{fmtCLP(r["costo_total"])}</td>'
        f'<td class="num" style="{DW["trab"]}">{fmtCLP(r.get("total_trabajado"))}</td>'
        f'<td class="num" style="{DW["pct"]}">{fmtPct(r["pct_pago"])}</td></tr>'
        for r in rows
    )
    sum_jornadas = sum(float(r["jornadas"] or 0) for r in rows)
    sum_costo = sum(float(r["costo_total"] or 0) for r in rows)
    sum_trab = sum(float(r.get("total_trabajado") or 0) for r in rows)
    foot_pct = "100.00 %" if sum_costo > 0 else "—"
    foot_html = (
        f'<tr><td colspan="4"><strong>Total</strong></td>'
        f'<td class="num" style="{DW["costo_hora"]}"></td>'
        f'<td class="num" style="{DW["jornadas"]}"><strong>{int(sum_jornadas)}</strong></td>'
        f'<td class="total" style="{DW["total"]}"><strong>{fmtCLP(sum_costo)}</strong></td>'
        f'<td class="num" style="{DW["trab"]}"><strong>{fmtCLP(sum_trab)}</strong></td>'
        f'<td class="num" style="{DW["pct"]}"><strong>{foot_pct}</strong></td></tr>'
    )

    # Resumen — mismos datos y colores que la pantalla (issue #96): la
    # pantalla ya calculaba esto vía /api/tarjas/detalle, el PDF no lo
    # incluía. Sin gráfico (issue #122): solo la tabla, ahora con % por fila.
    summary_section = ""
    if resumen:
        total_general = sum(float(r["total_pagar"] or 0) for r in resumen)
        jornadas_general = sum(r["jornadas"] or 0 for r in resumen)
        summary_html = _summary_table_html(resumen, total_general, jornadas_general)
        summary_section = f"""
        <p class="section-title">Resumen</p>
        {summary_html}
        """

    header = _pdf_header(
        _pdf_title("Detalle Operacional", contratista),
        fecha_inicio,
        fecha_termino,
        {
            "Empresa": empresa,
            "Contratista": contratista,
            "CC": centro_costo,
            "Tipo de pago": tipo_pago,
            "Labor": labor,
        },
    )
    return f"""
    {header}
    {summary_section}
    <p class="section-title">Detalle</p>
    <table class="detalle-table" style="table-layout:fixed"><thead>
      <tr><th style="{DW["tipo"]}">Tipo pago</th><th style="{DW["labor"]}">Labor</th>
      <th style="{DW["cc"]}">CC</th><th style="{DW["nombre_cc"]}">Nombre CC</th>
      <th class="num" style="{DW["costo_hora"]}">Costo/hora</th>
      <th class="num" style="{DW["jornadas"]}">Jornadas</th>
      <th class="num" style="{DW["total"]}">Total</th>
      <th class="num" style="{DW["trab"]}">Total trabajado</th>
      <th class="num" style="{DW["pct"]}">% pago</th></tr>
    </thead><tbody>{rows_html}</tbody><tfoot>{foot_html}</tfoot></table>
    """


@router.get("/api/tarjas/detalle/download-pdf")
async def download_tarjas_detalle_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
    campo: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    try:
        with conn.cursor() as cur:
            body = _build_detalle_html(
                cur,
                fecha_inicio,
                fecha_termino,
                contratista,
                centro_costo,
                tipo_pago,
                labor,
                campo,
                empresa,
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}</style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(html, f"detalle_{fecha_inicio}_{fecha_termino}.pdf")


def _build_contratista_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    contratista: str | None = None,
    centro_costo: str | None = None,
    tipo_pago: str | None = None,
    labor: str | None = None,
    empresa: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical."""
    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if centro_costo:
        filters.append("cuartel_cc = %s")
        params.append(centro_costo)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if labor:
        filters.append("labor = %s")
        params.append(labor)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    where = "WHERE " + " AND ".join(filters)
    _horas_sum = "COALESCE(SUM(horas_trabajadas), 0)"
    cur.execute(
        f"""
        SELECT trabajador, contratista, labor, tipo_pago,
               fecha::date::text AS fecha,
               COALESCE(SUM(total_trabajado), 0) AS total,
               {_horas_sum} AS horas
        FROM appsheet.tarjas_pagos {where}
        GROUP BY trabajador, contratista, labor, tipo_pago, fecha::date
        ORDER BY contratista, trabajador, labor, fecha::date
        """,
        params,
    )
    rows = _rows_to_dicts(cur)

    def clp(v):
        return f"${float(v):,.0f}".replace(",", ".")

    # Same per-date pivot as download_tarjas_contratista_excel — one column
    # per date in the range, matching what the on-screen pivot table shows.
    from collections import OrderedDict

    dates = sorted({r["fecha"] for r in rows})
    _check_pivot_date_range(dates, "Detalle contratista")
    groups: "OrderedDict" = OrderedDict()
    for r in rows:
        key = (r["trabajador"], r["contratista"], r["labor"], r["tipo_pago"])
        if key not in groups:
            groups[key] = {"total_ganado": 0.0, "total_horas": 0.0, "by_date": {}}
        g = groups[key]
        g["total_ganado"] += float(r["total"] or 0)
        g["total_horas"] += float(r["horas"] or 0)
        g["by_date"][r["fecha"]] = g["by_date"].get(r["fecha"], 0) + float(r["total"] or 0)

    w = _pivot_col_widths(
        {"worker": 15, "labor": 17, "tipo": 7, "rate": 8, "total": 9}, len(dates)
    )
    date_headers = "".join(
        f'<th class="num" style="{w["date"]}">'
        f'{datetime.date.fromisoformat(d).strftime("%d/%m")}</th>'
        for d in dates
    )

    rows_html = ""
    prev_worker = None
    for (trabajador, cont, labor_, tipo), g in groups.items():
        trabajador = trabajador or ""
        is_new_worker = trabajador != prev_worker
        row_cls = "worker-first" if is_new_worker else ""
        worker_cell = trabajador if is_new_worker else ""
        prev_worker = trabajador
        costo_hora = (
            clp(round(g["total_ganado"] / g["total_horas"]))
            if g["total_horas"] > 0
            else "-"
        )
        date_cells = "".join(
            f'<td class="num" style="{w["date"]}">'
            f'{clp(g["by_date"][d]) if g["by_date"].get(d) else "-"}</td>'
            for d in dates
        )
        rows_html += (
            f"<tr class='{row_cls}'>"
            f"<td style='{w['worker']}'>{worker_cell}</td>"
            f"<td style='{w['labor']}'>{labor_ or ''}</td>"
            f"<td style='{w['tipo']}'>{tipo or ''}</td>"
            f"<td class='num' style='{w['rate']}'>{costo_hora}</td>"
            f"{date_cells}"
            f"<td class='total' style='{w['total']}'>{clp(g['total_ganado'])}</td></tr>"
        )

    header = _pdf_header(
        _pdf_title("Detalle Contratistas", contratista),
        fecha_inicio,
        fecha_termino,
        {
            "Empresa": empresa,
            "Contratista": contratista,
            "CC": centro_costo,
            "Tipo de pago": tipo_pago,
            "Labor": labor,
        },
    )
    return f"""
    {header}
    <table class="pivot-wide" style="{w['table']}"><thead>
      <tr>
        <th style="{w['worker']}">Trabajador</th>
        <th style="{w['labor']}">Labor</th>
        <th style="{w['tipo']}">Tipo</th>
        <th class="num" style="{w['rate']}">Costo/hr</th>
        {date_headers}
        <th class="num" style="{w['total']}">Total</th>
      </tr>
    </thead><tbody>{rows_html}</tbody></table>
    """


@router.get("/api/tarjas/contratista/download-pdf")
async def download_tarjas_contratista_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    centro_costo: str = Query(None),
    tipo_pago: str = Query(None),
    labor: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    try:
        with conn.cursor() as cur:
            body = _build_contratista_html(
                cur,
                fecha_inicio,
                fecha_termino,
                contratista,
                centro_costo,
                tipo_pago,
                labor,
                empresa,
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}</style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(html, f"contratista_{fecha_inicio}_{fecha_termino}.pdf")


def _build_resumen_horas_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    trabajador: str | None = None,
    tipo_pago: str | None = None,
    contratista: str | None = None,
    empresa: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical."""
    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if trabajador:
        filters.append("trabajador = %s")
        params.append(trabajador)
    if tipo_pago:
        filters.append("tipo_pago = %s")
        params.append(tipo_pago)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    where = "WHERE " + " AND ".join(filters)
    cur.execute(
        f"SELECT trabajador, tipo_pago, fecha::date::text AS fecha, "
        f"COALESCE(SUM(horas_extras), 0)::numeric AS horas, "
        f"COALESCE(SUM(total_hora_extra), 0)::numeric AS monto "
        f"FROM appsheet.tarjas_pagos {where} "
        "GROUP BY trabajador, tipo_pago, fecha::date ORDER BY trabajador, tipo_pago, fecha::date",
        params,
    )
    rows = _rows_to_dicts(cur)

    dates = sorted({r["fecha"] for r in rows})
    workers: dict = {}
    for r in rows:
        k = (r["trabajador"] or "", r["tipo_pago"] or "")
        if k not in workers:
            workers[k] = {"by_date": {}, "total": 0, "monto": 0}
        workers[k]["by_date"][r["fecha"]] = workers[k]["by_date"].get(r["fecha"], 0) + (
            r["horas"] or 0
        )
        workers[k]["total"] += r["horas"] or 0
        workers[k]["monto"] += r["monto"] or 0

    # Only keep workers whose horas_extras total for the whole period (across
    # every tipo_pago) is greater than zero.
    totals_by_worker: dict = {}
    for (trab, _tipo), entry in workers.items():
        totals_by_worker[trab] = totals_by_worker.get(trab, 0) + entry["total"]
    workers = {
        k: v for k, v in workers.items() if totals_by_worker.get(k[0], 0) > 0
    }

    sorted_workers = sorted(workers.items(), key=lambda x: -x[1]["total"])

    resumen_horas = sum(e["total"] for _, e in sorted_workers)
    resumen_trabajadores = len({k[0] for k in workers})
    resumen_monto = sum(e["monto"] for _, e in sorted_workers)

    _check_pivot_date_range(dates, "Horas extra por persona")
    w = _pivot_col_widths({"worker": 16, "tipo": 7, "total": 7, "monto": 10}, len(dates))
    date_headers = "".join(
        f'<th class="num" style="{w["date"]}">'
        f'{datetime.date.fromisoformat(d).strftime("%d/%m")}</th>'
        for d in dates
    )
    rows_html = ""
    prev = None
    for (trab, tipo), entry in sorted_workers:
        is_first = prev != trab
        prev = trab
        cls = "worker-first" if is_first else ""
        rows_html += (
            f'<tr class="{cls}">'
            f'<td style="{w["worker"]}">{"<b>" + trab + "</b>" if is_first else ""}</td>'
            f'<td style="{w["tipo"]}">{tipo}</td>'
        )
        for d in dates:
            v = entry["by_date"].get(d, 0)
            rows_html += f'<td class="num" style="{w["date"]}">{v if v else ""}</td>'
        rows_html += (
            f'<td class="total" style="{w["total"]}">{entry["total"]}</td>'
            f'<td class="total" style="{w["monto"]}">{_fmt_clp(entry["monto"])}</td></tr>'
        )

    header = _pdf_header(
        _pdf_title("Horas Extra", contratista),
        fecha_inicio,
        fecha_termino,
        {
            "Empresa": empresa,
            "Contratista": contratista,
            "Trabajador": trabajador,
            "Tipo de pago": tipo_pago,
        },
    )
    summary_html = (
        '<table class="pdf-summary"><tr>'
        f'<td>Total horas extra<b>{resumen_horas}</b></td>'
        f'<td>Trabajadores con horas extra<b>{resumen_trabajadores}</b></td>'
        f'<td>Total a pagar<b>{_fmt_clp(resumen_monto)}</b></td>'
        "</tr></table>"
    )
    return f"""
    {header}
    <p class="pdf-note">*Sólo se muestran aquellos trabajadores que cuentan con horas extras en el periodo especificado.</p>
    {summary_html}
    <table class="pivot-wide" style="{w['table']}"><thead>
      <tr><th style="{w['worker']}">Trabajador</th><th style="{w['tipo']}">Tipo de pago</th>{date_headers}<th class="num" style="{w['total']}">Total hrs</th><th class="num" style="{w['monto']}">Monto</th></tr>
    </thead><tbody>{rows_html}</tbody></table>
    """


@router.get("/api/tarjas/resumen-horas/download-pdf")
async def download_tarjas_resumen_horas_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    trabajador: str = Query(None),
    tipo_pago: str = Query(None),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    try:
        with conn.cursor() as cur:
            body = _build_resumen_horas_html(
                cur, fecha_inicio, fecha_termino, trabajador, tipo_pago, contratista, empresa
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}</style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(html, f"horas_{fecha_inicio}_{fecha_termino}.pdf")


# ===========================================================================
# Jornadas por trabajador — count of distinct work dates per worker
# ===========================================================================


@router.get("/tarjas/jornadas-trabajador", response_class=HTMLResponse)
async def tarjas_jornadas_trabajador_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_jornadas_trabajador.html")


@router.get("/api/tarjas/jornadas-trabajador/filters")
async def get_tarjas_jornadas_trabajador_filters():
    """Distinct contratistas and empresas for filter dropdowns."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(cur, "appsheet.tarjas_pagos")
    finally:
        conn.close()

    return {"contratistas": contratistas, "empresas": empresas}


@router.get("/api/tarjas/jornadas-trabajador")
async def get_tarjas_jornadas_trabajador(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    """Return jornada count (distinct work dates) per worker for the given period."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]

    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))

    where = "WHERE " + " AND ".join(filters)

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT trabajador, contratista, COUNT(DISTINCT fecha::date) AS jornadas "
                f"FROM appsheet.tarjas_pagos {where} "
                "GROUP BY trabajador, contratista "
                "ORDER BY contratista, trabajador",
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    return {"rows": rows, "count": len(rows)}


@router.get("/api/tarjas/jornadas-trabajador/download-excel")
async def download_tarjas_jornadas_trabajador_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    """Descarga el reporte de jornadas por trabajador como Excel."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))

    where = "WHERE " + " AND ".join(filters)

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT trabajador, contratista, COUNT(DISTINCT fecha::date) AS jornadas "
                f"FROM appsheet.tarjas_pagos {where} "
                "GROUP BY trabajador, contratista ORDER BY contratista, trabajador",
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jornadas por trabajador"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="D6E4F0")

    headers = ["Trabajador", "Contratista", "Jornadas"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_jornadas = 0
    for row_num, r in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=r["trabajador"] or "")
        ws.cell(row=row_num, column=2, value=r["contratista"] or "")
        jornadas = int(r["jornadas"] or 0)
        ws.cell(row=row_num, column=3, value=jornadas)
        total_jornadas += jornadas

    # Total row
    total_row = len(rows) + 2
    total_cell = ws.cell(row=total_row, column=1, value="Suma total")
    total_cell.font = Font(bold=True)
    total_cell.fill = total_fill
    ws.cell(row=total_row, column=2, value="").fill = total_fill
    t = ws.cell(row=total_row, column=3, value=total_jornadas)
    t.font = Font(bold=True)
    t.fill = total_fill

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"jornadas_trabajador_{fecha_inicio}_{fecha_termino}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_jornadas_trabajador_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    contratista: str | None = None,
    empresa: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical."""
    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    where = "WHERE " + " AND ".join(filters)
    cur.execute(
        f"SELECT trabajador, contratista, COUNT(DISTINCT fecha::date) AS jornadas "
        f"FROM appsheet.tarjas_pagos {where} "
        "GROUP BY trabajador, contratista ORDER BY contratista, trabajador",
        params,
    )
    rows = _rows_to_dicts(cur)

    total = sum(int(r["jornadas"] or 0) for r in rows)
    # Explicit widths (table-layout:fixed) on a narrower-than-100% table —
    # without them this 3-column table's auto layout collapsed "Contratista"
    # to near-zero width and overlapped its text (issue #132).
    W = {"trabajador": "width:47%", "contratista": "width:41%", "jornadas": "width:12%"}
    rows_html = "".join(
        f'<tr><td style="{W["trabajador"]}">{r["trabajador"] or ""}</td>'
        f'<td style="{W["contratista"]}">{r["contratista"] or ""}</td>'
        f'<td class="num" style="{W["jornadas"]}">{r["jornadas"] or 0}</td></tr>'
        for r in rows
    )
    rows_html += (
        f'<tr class="total-row"><td style="{W["trabajador"]}"><b>Suma total</b></td>'
        f'<td style="{W["contratista"]}"></td>'
        f'<td class="num" style="{W["jornadas"]}"><b>{total}</b></td></tr>'
    )

    header = _pdf_header(
        _pdf_title("Jornadas Por Trabajador", contratista),
        fecha_inicio,
        fecha_termino,
        {"Empresa": empresa, "Contratista": contratista},
    )
    return f"""
    {header}
    <table style="width:55%;table-layout:fixed"><thead>
      <tr><th style="{W["trabajador"]}">Trabajador</th><th style="{W["contratista"]}">Contratista</th><th class="num" style="{W["jornadas"]}">Jornadas</th></tr>
    </thead><tbody>{rows_html}</tbody></table>
    """


@router.get("/api/tarjas/jornadas-trabajador/download-pdf")
async def download_tarjas_jornadas_trabajador_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    try:
        with conn.cursor() as cur:
            body = _build_jornadas_trabajador_html(
                cur, fecha_inicio, fecha_termino, contratista, empresa
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}
    .total-row td {{ background:#D6E4F0; font-weight:bold; }}
    </style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(html, f"jornadas_trabajador_{fecha_inicio}_{fecha_termino}.pdf")


# ===========================================================================
# Bonos mensuales — monthly bonus payments recorded in tarjas_pagos
# ===========================================================================

_MES_RE = re.compile(r"^\d{4}-\d{2}$")
_BONO_MENSUAL_LABOR = "Bono mensual"


def _mes_range(mes: str) -> tuple[str, str]:
    """First and last ISO day of a YYYY-MM month string."""
    year, month = (int(p) for p in mes.split("-"))
    first = datetime.date(year, month, 1)
    next_month = datetime.date(
        year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1
    )
    return first.isoformat(), (next_month - datetime.timedelta(days=1)).isoformat()


def _build_bono_mensual_filters(mes, contratista=None, empresa=None, campo=None):
    fecha_inicio, fecha_termino = _mes_range(mes)
    filters = ["labor = %s", "fecha::date BETWEEN %s AND %s"]
    params: list = [_BONO_MENSUAL_LABOR, fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if campo:
        filters.append("nombre_campo = %s")
        params.append(campo)
    return "WHERE " + " AND ".join(filters), params


def _query_bono_mensual_rows(cur, where, params):
    cur.execute(
        f"""
        SELECT
            trabajador,
            rut_trabajador,
            contratista,
            nombre_campo,
            cuartel_cc AS cc,
            fecha::date AS fecha,
            total_pagar AS monto,
            estado
        FROM appsheet.tarjas_pagos
        {where}
        ORDER BY contratista, trabajador, fecha
        """,
        params,
    )
    return _rows_to_dicts(cur)


@router.get("/tarjas/bono-mensual", response_class=HTMLResponse)
async def tarjas_bono_mensual_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_bono_mensual.html")


@router.get("/api/tarjas/bono-mensual/filters")
async def get_tarjas_bono_mensual_filters():
    """Distinct contratistas, empresas and campos among bonos mensuales."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE labor = %s AND contratista IS NOT NULL ORDER BY contratista",
                (_BONO_MENSUAL_LABOR,),
            )
            contratistas = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT nombre_campo FROM appsheet.tarjas_pagos "
                "WHERE labor = %s AND nombre_campo IS NOT NULL ORDER BY nombre_campo",
                (_BONO_MENSUAL_LABOR,),
            )
            campos = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(
                cur, "appsheet.tarjas_pagos", extra_where="labor = 'Bono mensual'"
            )
    finally:
        conn.close()

    return {"contratistas": contratistas, "empresas": empresas, "campos": campos}


@router.get("/api/tarjas/bono-mensual")
async def get_tarjas_bono_mensual(
    mes: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    campo: str = Query(None),
):
    """Bonos mensuales registrados en tarjas_pagos para el mes seleccionado."""
    if not _MES_RE.match(mes):
        raise HTTPException(status_code=400, detail="mes must be YYYY-MM")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    where, params = _build_bono_mensual_filters(mes, contratista, empresa, campo)
    try:
        with conn.cursor() as cur:
            rows = _query_bono_mensual_rows(cur, where, params)
    finally:
        conn.close()

    total = sum(float(r["monto"] or 0) for r in rows)
    return {"rows": rows, "count": len(rows), "total": total}


@router.get("/api/tarjas/bono-mensual/download-excel")
async def download_tarjas_bono_mensual_excel(
    mes: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    campo: str = Query(None),
):
    if not _MES_RE.match(mes):
        raise HTTPException(status_code=400, detail="mes must be YYYY-MM")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    where, params = _build_bono_mensual_filters(mes, contratista, empresa, campo)
    try:
        with conn.cursor() as cur:
            rows = _query_bono_mensual_rows(cur, where, params)
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bonos mensuales"
    _apply_header(
        ws,
        [
            "Trabajador",
            "RUT",
            "Contratista",
            "Empresa/Campo",
            "CC",
            "Fecha",
            "Monto",
            "Estado",
        ],
    )
    money = "#,##0"
    total = 0.0
    for i, r in enumerate(rows, 2):
        ws.cell(i, 1, r["trabajador"])
        ws.cell(i, 2, r["rut_trabajador"])
        ws.cell(i, 3, r["contratista"])
        ws.cell(i, 4, r["nombre_campo"])
        ws.cell(i, 5, r["cc"])
        ws.cell(i, 6, r["fecha"])
        monto = float(r["monto"] or 0)
        total += monto
        c = ws.cell(i, 7, monto)
        c.number_format = money
        ws.cell(i, 8, r["estado"])

    total_row = len(rows) + 2
    fill = PatternFill("solid", fgColor="D6E4F0")
    tcell = ws.cell(total_row, 1, "Suma total")
    tcell.font = Font(bold=True)
    for col in range(1, 9):
        ws.cell(total_row, col).fill = fill
    t = ws.cell(total_row, 7, total)
    t.font = Font(bold=True)
    t.number_format = money

    for col, w in zip("ABCDEFGH", [26, 14, 28, 20, 10, 12, 14, 14]):
        ws.column_dimensions[col].width = w

    return _excel_response(wb, f"bonos_mensuales_{mes}.xlsx")


def _build_bono_mensual_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    empresa: str | None = None,
    contratista: str | None = None,
    campo: str | None = None,
    mes: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical. The
    standalone page filters by a whole calendar month (mes=YYYY-MM,
    resolved to fecha_inicio/fecha_termino by the caller); the bulk PDF
    passes an arbitrary date range instead. `mes` is only used for the
    header's "Mes" chip, shown when the caller has one."""
    filters = ["labor = %s", "fecha::date BETWEEN %s AND %s"]
    params: list = [_BONO_MENSUAL_LABOR, fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if campo:
        filters.append("nombre_campo = %s")
        params.append(campo)
    where = "WHERE " + " AND ".join(filters)
    rows = _query_bono_mensual_rows(cur, where, params)

    total = sum(float(r["monto"] or 0) for r in rows)

    # Explicit per-column widths (sum to 100%). Without these, reportlab's
    # auto column-width algorithm collapses several columns to near-zero
    # width and overlaps their text once the table has this many columns —
    # reproduced with real data and confirmed fixed by declaring widths on
    # every <th>/<td>, including the footer row (see specs/117-*).
    W = {
        "trabajador": "width:16%",
        "rut": "width:11%",
        "contratista": "width:20%",
        "campo": "width:14%",
        "cc": "width:7%",
        "fecha": "width:9%",
        "monto": "width:11%",
        "estado": "width:12%",
    }
    rows_html = "".join(
        f'<tr><td style="{W["trabajador"]}">{_escape_html(r["trabajador"] or "")}</td>'
        f'<td style="{W["rut"]}">{_escape_html(r["rut_trabajador"] or "")}</td>'
        f'<td style="{W["contratista"]}">{_escape_html(r["contratista"] or "")}</td>'
        f'<td style="{W["campo"]}">{_escape_html(r["nombre_campo"] or "")}</td>'
        f'<td style="{W["cc"]}">{_escape_html(str(r["cc"] or ""))}</td>'
        f'<td style="{W["fecha"]}">{_fmt_date_display(str(r["fecha"]))}</td>'
        f'<td class="num" style="{W["monto"]}">{_fmt_clp(r["monto"])}</td>'
        f'<td style="{W["estado"]}">{_escape_html(r["estado"] or "")}</td></tr>'
        for r in rows
    )
    rows_html += (
        f'<tr class="total-row"><td style="{W["trabajador"]}"><b>Suma total</b></td>'
        f'<td style="{W["rut"]}"></td><td style="{W["contratista"]}"></td>'
        f'<td style="{W["campo"]}"></td><td style="{W["cc"]}"></td><td style="{W["fecha"]}"></td>'
        f'<td class="num" style="{W["monto"]}"><b>{_fmt_clp(total)}</b></td>'
        f'<td style="{W["estado"]}"></td></tr>'
    )

    header = _pdf_header(
        _pdf_title("Bonos Mensuales", contratista),
        fecha_inicio,
        fecha_termino,
        {"Mes": mes, "Empresa": empresa, "Campo": campo, "Contratista": contratista},
    )
    return f"""
    {header}
    <table style="width:88%;table-layout:fixed"><thead>
      <tr><th style="{W["trabajador"]}">Trabajador</th><th style="{W["rut"]}">RUT</th>
      <th style="{W["contratista"]}">Contratista</th><th style="{W["campo"]}">Empresa/Campo</th>
      <th style="{W["cc"]}">CC</th><th style="{W["fecha"]}">Fecha</th>
      <th class="num" style="{W["monto"]}">Monto</th><th style="{W["estado"]}">Estado</th></tr>
    </thead><tbody>{rows_html}</tbody></table>
    """


@router.get("/api/tarjas/bono-mensual/download-pdf")
async def download_tarjas_bono_mensual_pdf(
    mes: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    campo: str = Query(None),
):
    if not _MES_RE.match(mes):
        raise HTTPException(status_code=400, detail="mes must be YYYY-MM")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    fecha_inicio, fecha_termino = _mes_range(mes)
    try:
        with conn.cursor() as cur:
            body = _build_bono_mensual_html(
                cur, fecha_inicio, fecha_termino, empresa, contratista, campo, mes
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}
    .total-row td {{ background:#D6E4F0; font-weight:bold; }}
    </style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(html, f"bonos_mensuales_{mes}.pdf")


# ===========================================================================
# Hora ponderada 9h — Labor x CC pivot, hourly rate projected to a 9h day
# ===========================================================================

# Any daily cell above this amount is highlighted on screen, Excel and PDF.
HORA_PONDERADA_HIGHLIGHT_THRESHOLD = 30000

# This pivot has 3 fixed columns (Labor/CC/Total) plus large currency values
# (can run into the millions), so it needs wider date columns than the
# 2-fixed-column pivots elsewhere in this file. Capped at roughly a full
# calendar month of working days (verified legible up to 23 date columns at
# the widths below); wider ranges are redirected to Excel instead of a PDF
# that would be uncomfortably cramped.
MAX_PIVOT_DATES_HORA_PONDERADA = 23


def _build_hora_ponderada_filters(
    fecha_inicio,
    fecha_termino,
    contratista=None,
    empresa=None,
    centro_costo=None,
    labor=None,
):
    filters = ["fecha::date BETWEEN %s AND %s", f"NOT {_TRACTORISTA_PAGOS_SQL}"]
    params: list = [fecha_inicio, fecha_termino]
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if centro_costo:
        filters.append("cuartel_cc = %s")
        params.append(centro_costo)
    if labor:
        filters.append("labor = %s")
        params.append(labor)
    return "WHERE " + " AND ".join(filters), params


def _query_hora_ponderada_rows(cur, where, params):
    """One row per Labor+CC+Fecha cell — same costo_hora formula as Detalle
    Operacional (_query_detalle_rows: SUM(total_labor)/SUM(horas_trabajadas)),
    aggregated at the pivot's own cell granularity (Labor+CC+Fecha) instead of
    trabajador level, since this report's rows are Labor -> CC, not workers."""
    cur.execute(
        f"""
        SELECT
            labor,
            cuartel_cc                               AS centro_costo,
            fecha::date::text                        AS fecha,
            COALESCE(SUM(total_trabajado), 0)        AS total_trabajado,
            COALESCE(SUM(horas_trabajadas), 0)       AS horas_trabajadas
        FROM appsheet.tarjas_pagos
        {where}
        GROUP BY labor, cuartel_cc, fecha::date
        ORDER BY labor, cuartel_cc, fecha::date
    """,
        params,
    )
    return _rows_to_dicts(cur)


def _hora_ponderada_9h(total_trabajado, horas_trabajadas) -> int | None:
    """hora_ponderada_9h = ROUND(costo_hora * 9, 0), costo_hora = total/horas.
    Returns None (rendered as '-') when horas_trabajadas is 0/NULL for the
    aggregate being projected."""
    if not horas_trabajadas:
        return None
    return round(total_trabajado / horas_trabajadas * 9)


@router.get("/tarjas/hora-ponderada-9h", response_class=HTMLResponse)
async def tarjas_hora_ponderada_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_hora_ponderada.html")


@router.get("/api/tarjas/hora-ponderada-9h/filters")
async def get_tarjas_hora_ponderada_filters():
    """Distinct values for each filter dropdown (from tarjas_pagos directly)."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                f"WHERE contratista IS NOT NULL AND NOT {_TRACTORISTA_PAGOS_SQL} "
                f"ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]

            cur.execute(
                f"SELECT DISTINCT cuartel_cc FROM appsheet.tarjas_pagos "
                f"WHERE cuartel_cc IS NOT NULL AND NOT {_TRACTORISTA_PAGOS_SQL} "
                f"ORDER BY cuartel_cc"
            )
            centros_costo = [r[0] for r in cur.fetchall()]

            cur.execute(
                f"SELECT DISTINCT labor FROM appsheet.tarjas_pagos "
                f"WHERE labor IS NOT NULL AND NOT {_TRACTORISTA_PAGOS_SQL} "
                f"ORDER BY labor"
            )
            labores = [r[0] for r in cur.fetchall()]
            empresas = _get_empresas(
                cur, "appsheet.tarjas_pagos", extra_where=f"NOT {_TRACTORISTA_PAGOS_SQL}"
            )
    finally:
        conn.close()

    return {
        "contratistas": contratistas,
        "empresas": empresas,
        "centros_costo": centros_costo,
        "labores": labores,
    }


@router.get("/api/tarjas/hora-ponderada-9h")
async def get_tarjas_hora_ponderada_data(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    labor: str = Query(None),
):
    """One row per Labor+CC+Fecha cell; the client pivots them into a
    Labor -> CC row grouping with one column per date."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    where, params = _build_hora_ponderada_filters(
        fecha_inicio, fecha_termino, contratista, empresa, centro_costo, labor
    )
    try:
        with conn.cursor() as cur:
            rows = _query_hora_ponderada_rows(cur, where, params)
    finally:
        conn.close()

    return {"rows": rows, "count": len(rows)}


@router.get("/api/tarjas/hora-ponderada-9h/download-excel")
async def download_tarjas_hora_ponderada_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    labor: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    where, params = _build_hora_ponderada_filters(
        fecha_inicio, fecha_termino, contratista, empresa, centro_costo, labor
    )
    try:
        with conn.cursor() as cur:
            rows = _query_hora_ponderada_rows(cur, where, params)
    finally:
        conn.close()

    # Pivot in Python: rows = Labor -> CC, columns = one per date, cell =
    # hora_ponderada_9h for that Labor+CC+Fecha combo (mirrors the per-date
    # pivot built in download_tarjas_contratista_excel).
    from collections import OrderedDict
    from openpyxl.utils import get_column_letter

    dates = sorted({r["fecha"] for r in rows})
    groups: "OrderedDict" = OrderedDict()
    grand_total = 0.0
    grand_horas = 0.0
    for r in rows:
        key = (r["labor"], r["centro_costo"])
        if key not in groups:
            groups[key] = {"total_trabajado": 0.0, "total_horas": 0.0, "by_date": {}}
        g = groups[key]
        total = float(r["total_trabajado"] or 0)
        horas = float(r["horas_trabajadas"] or 0)
        g["total_trabajado"] += total
        g["total_horas"] += horas
        g["by_date"][r["fecha"]] = (total, horas)
        grand_total += total
        grand_horas += horas

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hora ponderada 9h"
    fixed_headers = ["Labor", "CC"]
    date_headers = [d[5:] for d in dates]  # MM-DD
    _apply_header(ws, fixed_headers + date_headers + ["Hora ponderada 9h"])

    fill_hdr, font_hdr, align_hdr = _excel_header_style()
    total_col_idx = len(fixed_headers) + len(dates) + 1
    ws.cell(1, total_col_idx).fill = fill_hdr
    ws.cell(1, total_col_idx).font = font_hdr
    ws.cell(1, total_col_idx).alignment = align_hdr

    money = "#,##0"
    highlight_fill = PatternFill("solid", fgColor="FFEDD5")
    highlight_font = Font(bold=True, color="C2410C")
    for i, ((labor_, cc), g) in enumerate(groups.items(), 2):
        ws.cell(i, 1, labor_)
        ws.cell(i, 2, cc)
        for j, d in enumerate(dates, 3):
            cell = g["by_date"].get(d)
            if cell:
                val = _hora_ponderada_9h(*cell)
                if val is not None:
                    dc = ws.cell(i, j, val)
                    dc.number_format = money
                    if val > HORA_PONDERADA_HIGHLIGHT_THRESHOLD:
                        dc.fill = highlight_fill
                        dc.font = highlight_font
        row_total = _hora_ponderada_9h(g["total_trabajado"], g["total_horas"])
        tc = ws.cell(i, total_col_idx, row_total)
        if row_total is not None:
            tc.number_format = money
        tc.font = Font(bold=True)

    # Footer: blended hora_ponderada_9h across ALL Labor+CC rows — NOT a sum
    # of the per-row projected values (see spec Decisions: summing an
    # hourly-rate projection across independent Labor/CC rows is not
    # economically meaningful; instead the same formula is recomputed over
    # the grand totals, same approach as the "Costo/hr" column elsewhere).
    footer_row = len(groups) + 2
    fill = PatternFill("solid", fgColor="D6E4F0")
    tcell = ws.cell(footer_row, 1, "Hora ponderada 9h global")
    tcell.font = Font(bold=True)
    for col in range(1, total_col_idx + 1):
        ws.cell(footer_row, col).fill = fill
    footer_val = _hora_ponderada_9h(grand_total, grand_horas)
    fcell = ws.cell(footer_row, total_col_idx, footer_val)
    fcell.font = Font(bold=True)
    if footer_val is not None:
        fcell.number_format = money

    fixed_widths = [28, 14]
    date_widths = [10] * len(dates)
    for idx, w in enumerate(fixed_widths + date_widths + [18], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    return _excel_response(
        wb, f"tarjas_hora_ponderada_9h_{fecha_inicio}_{fecha_termino}.xlsx"
    )


def _build_hora_ponderada_html(
    cur,
    fecha_inicio: str,
    fecha_termino: str,
    contratista: str | None = None,
    empresa: str | None = None,
    centro_costo: str | None = None,
    labor: str | None = None,
) -> str:
    """Shared by the standalone PDF endpoint and the bulk /reportes PDF
    (issue #116) — a single source of truth so both stay identical."""
    where, params = _build_hora_ponderada_filters(
        fecha_inicio, fecha_termino, contratista, empresa, centro_costo, labor
    )
    rows = _query_hora_ponderada_rows(cur, where, params)

    # Same per-date pivot as download_tarjas_hora_ponderada_excel — one
    # column per date, matching what the on-screen pivot table shows.
    from collections import OrderedDict

    dates = sorted({r["fecha"] for r in rows})
    _check_pivot_date_range(
        dates, "Hora ponderada 9h", MAX_PIVOT_DATES_HORA_PONDERADA
    )
    groups: "OrderedDict" = OrderedDict()
    grand_total = 0.0
    grand_horas = 0.0
    for r in rows:
        key = (r["labor"], r["centro_costo"])
        if key not in groups:
            groups[key] = {"total_trabajado": 0.0, "total_horas": 0.0, "by_date": {}}
        g = groups[key]
        total = float(r["total_trabajado"] or 0)
        horas = float(r["horas_trabajadas"] or 0)
        g["total_trabajado"] += total
        g["total_horas"] += horas
        g["by_date"][r["fecha"]] = (total, horas)
        grand_total += total
        grand_horas += horas

    # date_pct wider than the default (5 vs 3.5): this pivot's cells hold
    # money values up to 7 digits ($X.XXX.XXX), unlike the smaller hour/count
    # values in other pivots — narrower columns risk reproducing #108's
    # column-overlap bug.
    w = _pivot_col_widths({"labor": 13, "cc": 8, "total": 13}, len(dates), date_pct=5.0)
    date_headers = "".join(
        f'<th class="num" style="{w["date"]}">'
        f'{datetime.date.fromisoformat(d).strftime("%d/%m")}</th>'
        for d in dates
    )

    # Daily cells above HORA_PONDERADA_HIGHLIGHT_THRESHOLD get a highlighted
    # style, matching the on-screen pivot and the Excel export.
    _HIGHLIGHT_STYLE = "background:#ffedd5;color:#c2410c;font-weight:bold;"

    def cell_html(cell, style):
        v = _hora_ponderada_9h(*cell) if cell else None
        text = _fmt_clp(v) if v is not None else "-"
        cell_style = style + ";"
        if v is not None and v > HORA_PONDERADA_HIGHLIGHT_THRESHOLD:
            cell_style += _HIGHLIGHT_STYLE
        return f'<td class="num" style="{cell_style}">{text}</td>'

    rows_html = ""
    prev_labor = None
    for (labor_, cc), g in groups.items():
        labor_ = labor_ or ""
        is_new_labor = labor_ != prev_labor
        row_cls = "worker-first" if is_new_labor else ""
        labor_cell = labor_ if is_new_labor else ""
        prev_labor = labor_
        date_cells = "".join(
            cell_html(g["by_date"].get(d), w["date"]) for d in dates
        )
        row_total = _hora_ponderada_9h(g["total_trabajado"], g["total_horas"])
        rows_html += (
            f"<tr class='{row_cls}'>"
            f"<td style='{w['labor']}'>{_escape_html(labor_cell)}</td>"
            f"<td style='{w['cc']}'>{_escape_html(cc or '')}</td>"
            f"{date_cells}"
            f"<td class='total' style='{w['total']}'>"
            f"{_fmt_clp(row_total) if row_total is not None else '-'}</td></tr>"
        )

    # Footer: same blended-rate approach as the Excel export (see Decisions).
    # Every cell here must carry the same inline width style as the rest of
    # the table's cells (w[...]) — xhtml2pdf's table-layout:fixed column-width
    # computation breaks for the WHOLE table (columns overlap/bleed into each
    # other) if even one row's cells omit it, which was the root cause of the
    # "everything overlaps" PDF bug.
    footer_val = _hora_ponderada_9h(grand_total, grand_horas)
    rows_html += (
        f"<tr class='total-row'><td style='{w['labor']}'><b>Hora ponderada 9h global</b></td>"
        f"<td style='{w['cc']}'></td>"
        + f'<td class="num" style="{w["date"]}"></td>' * len(dates)
        + f"<td class='total' style='{w['total']}'><b>"
        f"{_fmt_clp(footer_val) if footer_val is not None else '-'}</b></td></tr>"
    )

    header = _pdf_header(
        _pdf_title("Hora Ponderada 9h", contratista),
        fecha_inicio,
        fecha_termino,
        {
            "Empresa": empresa,
            "Contratista": contratista,
            "CC": centro_costo,
            "Labor": labor,
        },
    )
    return f"""
    {header}
    <table class="pivot-wide" style="{w['table']}"><thead>
      <tr>
        <th style="{w['labor']}">Labor</th>
        <th style="{w['cc']}">CC</th>
        {date_headers}
        <th class="num" style="{w['total']}">Hora ponderada 9h</th>
      </tr>
    </thead><tbody>{rows_html}</tbody></table>
    """


@router.get("/api/tarjas/hora-ponderada-9h/download-pdf")
async def download_tarjas_hora_ponderada_pdf(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    contratista: str = Query(None),
    empresa: str = Query(None),
    centro_costo: str = Query(None),
    labor: str = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    try:
        with conn.cursor() as cur:
            body = _build_hora_ponderada_html(
                cur, fecha_inicio, fecha_termino, contratista, empresa, centro_costo, labor
            )
    finally:
        conn.close()
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>{_PDF_CSS}
    .total-row td {{ background:#D6E4F0; font-weight:bold; }}
    </style></head><body>
    {body}
    </body></html>"""
    return _render_pdf(
        html, f"hora_ponderada_9h_{fecha_inicio}_{fecha_termino}.pdf"
    )


# ===========================================================================
# Notas de crédito — contractor payment report (moved from despacho)
# ===========================================================================


@router.get("/tarjas/notas", response_class=HTMLResponse)
async def tarjas_notas_page(request: Request):
    return _templates.TemplateResponse(request, "despacho_notas.html")


@router.get("/api/tarjas/notas/filters")
async def get_tarjas_notas_filters():
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT nombre_campo FROM appsheet.tarjas_pagos "
                "WHERE nombre_campo IS NOT NULL AND estado = 'Aprobado' ORDER BY nombre_campo"
            )
            campos = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE contratista IS NOT NULL AND estado = 'Aprobado' ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]
    finally:
        conn.close()

    return {"campos": campos, "contratistas": contratistas}


@router.get("/api/tarjas/notas")
async def get_tarjas_notas(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    campo: str = Query(None),
    contratista: str = Query(...),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = [
        "fecha::date BETWEEN %s AND %s",
        "contratista = %s",
        "estado = 'Aprobado'",
    ]
    params: list = [fecha_inicio, fecha_termino, contratista]

    if campo:
        filters.append("nombre_campo = %s")
        params.append(campo)

    where = "WHERE " + " AND ".join(filters)

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                    tipo_pago,
                    cuartel_cc              AS cc,
                    labor,
                    COUNT(*)                AS jornadas,
                    ROUND(
                      CASE
                        WHEN LOWER(TRIM(tipo_pago)) IN ('a trato', 'trato')
                          THEN AVG(NULLIF(total_trato, 0))
                        ELSE AVG(NULLIF(total_jornada, 0))
                      END::numeric, 0
                    )                       AS total_unitario,
                    COALESCE(SUM(total_pagar), 0) AS total_pagar
                FROM appsheet.tarjas_pagos
                {where}
                GROUP BY tipo_pago, cuartel_cc, labor
                ORDER BY
                    CASE WHEN LOWER(TRIM(tipo_pago)) IN ('a trato','trato') THEN 0 ELSE 1 END,
                    cuartel_cc, labor
            """,
                params,
            )
            rows = _rows_to_dicts(cur)

            cur.execute(
                f"""
                SELECT
                    tipo_pago,
                    COALESCE(SUM(total_pagar), 0) AS total
                FROM appsheet.tarjas_pagos
                {where}
                GROUP BY tipo_pago
            """,
                params,
            )
            totals_by_tipo = {r[0]: float(r[1]) for r in cur.fetchall()}

            cur.execute(
                f"""
                SELECT DISTINCT nombre_campo
                FROM appsheet.tarjas_pagos
                {where}
                LIMIT 1
            """,
                params,
            )
            row = cur.fetchone()
            nombre_campo = row[0] if row else ""

    finally:
        conn.close()

    total_trato = sum(
        r["total_pagar"]
        for r in rows
        if r["tipo_pago"] and r["tipo_pago"].lower().strip() in ("a trato", "trato")
    )
    total_aldia = sum(
        r["total_pagar"]
        for r in rows
        if r["tipo_pago"] and r["tipo_pago"].lower().strip() not in ("a trato", "trato")
    )
    total_general = total_trato + total_aldia

    return {
        "nombre_campo": nombre_campo,
        "contratista": contratista,
        "fecha_inicio": fecha_inicio,
        "fecha_termino": fecha_termino,
        "total_trato": total_trato,
        "total_aldia": total_aldia,
        "total_general": total_general,
        "rows": rows,
        "count": len(rows),
    }


@router.get("/api/tarjas/notas/odoo-export")
async def export_tarjas_notas_odoo(
    contratista: str = Query(...),
    campo: str = Query(...),
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    nc_total: int = Query(
        None, description="Monto total NC para redistribución proporcional"
    ),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    # Auto-sync unmapped labores (same as purchase orders export)
    from .purchase_orders_controller import _sync_labores

    try:
        _sync_labores(conn, fecha_inicio, fecha_termino, contratista, campo)
    except Exception as exc:
        logger.warning(f"Labor auto-sync failed (non-fatal): {exc}")

    excluded_amount = 0.0
    try:
        with conn.cursor() as cur:
            # Sum excluded rows (no product_id or unmapped CC)
            cur.execute(
                """
                SELECT COALESCE(SUM("order_line/product_qty" * "order_line/price_unit"), 0)
                FROM appsheet.tarjas_reporte_odoo
                WHERE "Vendedor"     = %s
                  AND nombre_campo   = %s
                  AND fecha BETWEEN %s AND %s
                  AND (
                      "order_line/product_id" IS NULL
                      OR "order_line/analytic_distribution" LIKE '%%"": %%'
                  )
            """,
                (contratista, campo, fecha_inicio, fecha_termino),
            )
            excluded_amount = float(cur.fetchone()[0] or 0)

            # Group and aggregate valid rows — same structure as purchase orders
            cur.execute(
                """
                SELECT
                    "partner_id",
                    "order_line/product_id",
                    SUM("order_line/product_qty")               AS qty,
                    "order_line/analytic_distribution",
                    CASE WHEN SUM("order_line/product_qty") > 0
                         THEN ROUND(
                             SUM("order_line/product_qty" * "order_line/price_unit")
                             / SUM("order_line/product_qty"), 2)
                         ELSE NULL END                          AS price_unit
                FROM appsheet.tarjas_reporte_odoo
                WHERE "Vendedor"      = %s
                  AND "nombre_campo"  = %s
                  AND "fecha" BETWEEN %s AND %s
                  AND "order_line/product_id" IS NOT NULL
                  AND "order_line/analytic_distribution" NOT LIKE '%%"": %%'
                GROUP BY
                    "partner_id",
                    "order_line/product_id",
                    "order_line/analytic_distribution"
                ORDER BY "order_line/product_id"
            """,
                (contratista, campo, fecha_inicio, fecha_termino),
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    # If nc_total provided, scale price_unit proportionally
    if nc_total and nc_total > 0 and rows:
        original_total = sum(
            float(qty or 0) * float(price or 0) for _, _, qty, _, price in rows
        )
        scale = nc_total / original_total if original_total else 1.0
        rows = [
            (
                partner,
                product,
                qty,
                analytic,
                round(float(price) * scale, 2) if price is not None else None,
            )
            for partner, product, qty, analytic, price in rows
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    headers = [
        "partner_id",
        "order_line/product_id",
        "order_line/product_qty",
        "order_line/analytic_distribution",
        "order_line/price_unit",
    ]
    ws.append(headers)

    for i, (partner_id, product_id, qty, analytic, price) in enumerate(rows):
        ws.append(
            [
                partner_id if i == 0 else None,
                product_id,
                float(qty) if qty is not None else None,
                analytic,
                float(price) if price is not None else None,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (
        f"odoo_nota_{contratista.replace(' ', '_')}_{campo.replace(' ', '_')}"
        f"_{fecha_inicio}_{fecha_termino}.xlsx"
    )
    response_headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Access-Control-Expose-Headers": "X-Excluded-Amount",
    }
    if excluded_amount > 0:
        response_headers["X-Excluded-Amount"] = str(int(excluded_amount))

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=response_headers,
    )


# ---------------------------------------------------------------------------
# Tractorista Odoo purchase order page + export
# GET /odoo/tarjas-tractorista  → HTML page (Orden de compra tractorista)
# GET /api/tarjas/tractorista/filters  → filter dropdowns
# GET /api/tarjas/tractorista/odoo-export  → Excel download
# ---------------------------------------------------------------------------


@router.get("/odoo/tarjas-tractorista", response_class=HTMLResponse)
async def odoo_tarjas_tractorista_page(request: Request):
    return _templates.TemplateResponse(request, "purchase_orders_tractorista.html")


@router.get("/api/tarjas/tractorista/filters")
async def get_tarjas_tractorista_filters():
    """Filter options for the Odoo tractorista purchase order page."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE LOWER(TRIM(tipo_pago)) = 'tractorista' "
                "AND contratista IS NOT NULL ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT nombre_campo FROM appsheet.tarjas_pagos "
                "WHERE LOWER(TRIM(tipo_pago)) = 'tractorista' "
                "AND nombre_campo IS NOT NULL ORDER BY nombre_campo"
            )
            campos = [r[0] for r in cur.fetchall()]
    finally:
        conn.close()

    return {"contratistas": contratistas, "campos": campos}


@router.get("/api/tarjas/tractorista/export-preview")
async def get_tarjas_tractorista_export_preview(
    contratista: str = Query(...),
    campo: str = Query(...),
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
):
    """Preview Odoo export rows for tractoristas: ok (mapped) vs excluded (sin mapeo / sin CC)."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")

    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    "order_line/product_id"             AS product_id,
                    "order_line/product_qty"            AS qty,
                    "order_line/analytic_distribution"  AS analytic,
                    "order_line/price_unit"             AS price_unit
                FROM appsheet.tarjas_reporte_odoo_tractorista
                WHERE "Vendedor"   = %s
                  AND nombre_campo = %s
                  AND fecha BETWEEN %s AND %s
                ORDER BY "order_line/product_id", fecha
                """,
                (contratista, campo, fecha_inicio, fecha_termino),
            )
            all_rows = cur.fetchall()
    finally:
        conn.close()

    # product_id viene de tarjas_labores (labor='TRACTORISTA') — configurable sin tocar código
    try:
        conn2 = get_connection()
        with conn2.cursor() as cur2:
            cur2.execute(
                "SELECT codigo_labor FROM appsheet.tarjas_labores WHERE labor = 'TRACTORISTA' LIMIT 1"
            )
            r = cur2.fetchone()
            product_id_tractorista = r[0] if r else "ARRIET-002"
        conn2.close()
    except Exception:
        product_id_tractorista = "ARRIET-002"

    ok_rows, excl_rows = [], []
    for _product_id, qty, analytic, price_unit in all_rows:
        qty_f = float(qty or 0)
        price_f = float(price_unit or 0)
        row = {
            "product_id": product_id_tractorista,
            "qty": qty_f,
            "analytic_distribution": analytic or "–",
            "price_unit": price_f,
            "total": round(qty_f * price_f, 0),
        }
        missing_cc = not analytic or '"": ' in (analytic or "")
        if missing_cc:
            row["reason"] = "Sin CC mapeado"
            excl_rows.append(row)
        else:
            ok_rows.append(row)

    return {
        "rows": ok_rows,
        "excluded": excl_rows,
        "contratista": contratista,
        "campo": campo,
        "fecha_inicio": fecha_inicio,
        "fecha_termino": fecha_termino,
    }


@router.get("/api/tarjas/tractorista/preview")
async def get_tarjas_tractorista_preview(
    contratista: str = Query(...),
    campo: str = Query(...),
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
):
    """Preview rows from tarjas_reporte_odoo_tractorista for the given filters."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    fecha::text                                                  AS fecha,
                    "Lineas del pedido/Código de Distribución Analítica/Código" AS cc,
                    "Lineas del pedido/Producto/Nombre"                         AS labor,
                    SUM("order_line/product_qty")                               AS horas,
                    CASE WHEN SUM("order_line/product_qty") > 0
                         THEN ROUND(
                             SUM("order_line/product_qty" * "order_line/price_unit")
                             / SUM("order_line/product_qty"), 2)
                         ELSE NULL END                                          AS precio_hora,
                    COALESCE(MAX("order_line/product_id"), '(sin mapeo)')       AS product_id
                FROM appsheet.tarjas_reporte_odoo_tractorista
                WHERE "Vendedor"   = %s
                  AND nombre_campo = %s
                  AND fecha BETWEEN %s AND %s
                GROUP BY
                    fecha,
                    "Lineas del pedido/Código de Distribución Analítica/Código",
                    "Lineas del pedido/Producto/Nombre"
                ORDER BY cc, fecha, labor
                """,
                (contratista, campo, fecha_inicio, fecha_termino),
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    total_horas = sum(float(r[3] or 0) for r in rows)
    total_monto = sum(float(r[3] or 0) * float(r[4] or 0) for r in rows)

    return {
        "rows": [
            {
                "fecha": r[0],
                "cc": r[1],
                "labor": r[2],
                "horas": float(r[3]) if r[3] is not None else None,
                "precio_hora": float(r[4]) if r[4] is not None else None,
                "total": (float(r[3] or 0) * float(r[4] or 0)) if r[4] is not None else None,
                "product_id": r[5],
            }
            for r in rows
        ],
        "total_horas": total_horas,
        "total_monto": total_monto,
        "contratista": contratista,
        "campo": campo,
        "fecha_inicio": fecha_inicio,
        "fecha_termino": fecha_termino,
    }


@router.get("/api/tarjas/tractorista/odoo-export")
async def export_tarjas_tractorista_odoo(
    contratista: str = Query(...),
    campo: str = Query(...),
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT codigo_labor FROM appsheet.tarjas_labores WHERE labor = 'TRACTORISTA' LIMIT 1"
            )
            row = cur.fetchone()
            product_id_tractorista = row[0] if row else "ARRIET-002"

            cur.execute(
                """
                SELECT
                    "partner_id",
                    SUM("order_line/product_qty")               AS qty,
                    "order_line/analytic_distribution",
                    CASE WHEN SUM("order_line/product_qty") > 0
                         THEN ROUND(
                             SUM("order_line/product_qty" * "order_line/price_unit")
                             / SUM("order_line/product_qty"), 2)
                         ELSE NULL END                          AS price_unit
                FROM appsheet.tarjas_reporte_odoo_tractorista
                WHERE "Vendedor"   = %s
                  AND nombre_campo = %s
                  AND fecha BETWEEN %s AND %s
                  AND "order_line/analytic_distribution" IS NOT NULL
                  AND "order_line/analytic_distribution" NOT LIKE '%%"": %%'
                GROUP BY
                    "partner_id",
                    "order_line/analytic_distribution"
                ORDER BY "order_line/analytic_distribution"
                """,
                (contratista, campo, fecha_inicio, fecha_termino),
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    headers = [
        "partner_id",
        "order_line/product_id",
        "order_line/product_qty",
        "order_line/analytic_distribution",
        "order_line/price_unit",
    ]
    ws.append(headers)

    for i, (partner_id, qty, analytic, price) in enumerate(rows):
        ws.append(
            [
                partner_id if i == 0 else None,
                product_id_tractorista,
                float(qty) if qty is not None else None,
                analytic,
                float(price) if price is not None else None,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (
        f"odoo_tractorista_{contratista.replace(' ', '_')}_{campo.replace(' ', '_')}"
        f"_{fecha_inicio}_{fecha_termino}.xlsx"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _fetch_tractorista_pivot_rows(
    conn,
    contratista: str,
    campo: str,
    fecha_inicio: str,
    fecha_termino: str,
    cc: str | None = None,
) -> list[dict]:
    """Raw (fecha, trabajador, labor, monto) rows behind every 'por operador'
    view of the tractorista purchase order — shared by the Excel download,
    the on-screen JSON preview, and the PDF pivot section, so the three never
    drift apart. `cc` optionally scopes it to a single centro de costo, to
    match a CC-filtered PDF."""
    cc_filter = "AND cuartel_cc = %s" if cc else ""
    params = [contratista, campo, fecha_inicio, fecha_termino]
    if cc:
        params.append(cc)
    with conn.cursor() as cur:
        cur.execute(
            f"""
            SELECT
                fecha::date::text AS fecha,
                trabajador,
                labor,
                SUM(total_tractor) AS monto
            FROM appsheet.tarjas_pagos
            WHERE LOWER(TRIM(tipo_pago)) = 'tractorista'
              AND contratista  = %s
              AND nombre_campo = %s
              AND fecha::date BETWEEN %s AND %s
              {cc_filter}
            GROUP BY fecha::date, trabajador, labor
            ORDER BY fecha::date, trabajador, labor
            """,
            params,
        )
        return _rows_to_dicts(cur)


def _build_tractorista_pivot(rows: list[dict]) -> dict:
    """Pivot (fecha, trabajador, labor, monto) rows into fecha x
    "trabajador — labor", with row/column/grand totals. Dates stay as
    ISO strings (YYYY-MM-DD) — callers format for display."""

    def _col(r: dict) -> str:
        return f'{r["trabajador"] or "(sin nombre)"} — {r["labor"] or ""}'

    dates = sorted({r["fecha"] for r in rows})
    columns = sorted({_col(r) for r in rows})
    matrix = {d: {c: 0.0 for c in columns} for d in dates}
    for r in rows:
        matrix[r["fecha"]][_col(r)] += float(r["monto"] or 0)
    col_totals = {c: sum(matrix[d][c] for d in dates) for c in columns}
    date_totals = {d: sum(matrix[d].values()) for d in dates}
    grand_total = sum(col_totals.values())
    return {
        "dates": dates,
        "columns": columns,
        "matrix": matrix,
        "col_totals": col_totals,
        "date_totals": date_totals,
        "grand_total": grand_total,
    }


@router.get("/api/tarjas/tractorista/pivot-preview")
async def preview_tarjas_tractorista_pivot(
    contratista: str = Query(...),
    campo: str = Query(...),
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
):
    """JSON pivot (fecha x operador — labor) for the on-screen table rendered
    below the CC sections in the tractorista purchase order screen."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")
    try:
        rows = _fetch_tractorista_pivot_rows(conn, contratista, campo, fecha_inicio, fecha_termino)
    finally:
        conn.close()
    return _build_tractorista_pivot(rows)


@router.get("/api/tarjas/tractorista/pivot-excel")
async def pivot_tarjas_tractorista_excel(
    contratista: str = Query(...),
    campo: str = Query(...),
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
):
    """Excel pivot: filas=fecha, columnas=trabajador — labor, valores=total_tractor."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")

    try:
        rows = _fetch_tractorista_pivot_rows(conn, contratista, campo, fecha_inicio, fecha_termino)
    finally:
        conn.close()

    if not rows:
        raise HTTPException(status_code=404, detail="Sin datos para los filtros seleccionados")

    pivot = _build_tractorista_pivot(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Por Operador"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    total_fill  = PatternFill("solid", fgColor="E0E7FF")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    bold_font   = Font(bold=True)

    all_columns = pivot["columns"] + ["Suma total"]

    # Header row
    ws.cell(1, 1, "Fecha").font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    for col_idx, col_name in enumerate(all_columns, start=2):
        cell = ws.cell(1, col_idx, col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 22

    ws.column_dimensions["A"].width = 14
    ws.row_dimensions[1].height = 40

    # Data rows
    row_idx = 2
    for fecha in pivot["dates"]:
        fecha_fmt = datetime.date.fromisoformat(fecha).strftime("%d/%m/%Y")
        ws.cell(row_idx, 1, fecha_fmt)
        for col_idx, col_name in enumerate(pivot["columns"], start=2):
            cell = ws.cell(row_idx, col_idx, int(pivot["matrix"][fecha][col_name]))
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal="right")
        total_cell = ws.cell(row_idx, len(all_columns) + 1, int(pivot["date_totals"][fecha]))
        total_cell.number_format = '$#,##0'
        total_cell.alignment = Alignment(horizontal="right")
        total_cell.fill = total_fill
        total_cell.font = bold_font
        row_idx += 1

    # Totals row
    ws.cell(row_idx, 1, "Suma total").font = bold_font
    ws.cell(row_idx, 1).fill = total_fill
    for col_idx, col_name in enumerate(pivot["columns"], start=2):
        cell = ws.cell(row_idx, col_idx, int(pivot["col_totals"][col_name]))
        cell.number_format = '$#,##0'
        cell.alignment = Alignment(horizontal="right")
        cell.fill = total_fill
        cell.font = bold_font
    grand_cell = ws.cell(row_idx, len(all_columns) + 1, int(pivot["grand_total"]))
    grand_cell.number_format = '$#,##0'
    grand_cell.alignment = Alignment(horizontal="right")
    grand_cell.fill = total_fill
    grand_cell.font = bold_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (
        f"tractorista_operadores_{contratista.replace(' ', '_')}_{campo.replace(' ', '_')}"
        f"_{fecha_inicio}_{fecha_termino}.xlsx"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/tarjas/tractorista/download-pdf")
async def download_tarjas_tractorista_pdf(
    contratista: str = Query(...),
    campo: str = Query(...),
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    cc: str = Query(None),
):
    """PDF de detalle tractorista: agrupado por fecha → trabajador → labor. Opcionalmente filtrado por CC."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Error de conexión")

    try:
        with conn.cursor() as cur:
            cc_filter = "AND cuartel_cc = %s" if cc else ""
            params = [contratista, campo, fecha_inicio, fecha_termino]
            if cc:
                params.append(cc)
            cur.execute(
                f"""
                SELECT
                    fecha::date::text   AS fecha,
                    trabajador,
                    labor,
                    maquina,
                    SUM(total_tractor)  AS monto
                FROM appsheet.tarjas_pagos
                WHERE LOWER(TRIM(tipo_pago)) = 'tractorista'
                  AND contratista   = %s
                  AND nombre_campo  = %s
                  AND fecha::date BETWEEN %s AND %s
                  {cc_filter}
                GROUP BY fecha::date, trabajador, labor, maquina
                ORDER BY fecha::date, trabajador, labor
                """,
                params,
            )
            rows = _rows_to_dicts(cur)
        pivot_rows = _fetch_tractorista_pivot_rows(
            conn, contratista, campo, fecha_inicio, fecha_termino, cc=cc
        )
    finally:
        conn.close()

    fmt = lambda v: f"${int(v):,}".replace(",", ".")

    # Group by date for subtotals
    from collections import defaultdict
    by_date: dict = defaultdict(list)
    for r in rows:
        by_date[r["fecha"]].append(r)

    rows_html = ""
    grand_total = 0.0
    for fecha, day_rows in sorted(by_date.items()):
        fecha_fmt = datetime.date.fromisoformat(fecha).strftime("%d/%m/%Y")
        day_total = sum(float(r["monto"] or 0) for r in day_rows)
        grand_total += day_total
        for i, r in enumerate(day_rows):
            monto = float(r["monto"] or 0)
            rows_html += (
                f'<tr>'
                f'<td style="width:10%">{"<b>" + fecha_fmt + "</b>" if i == 0 else ""}</td>'
                f'<td style="width:24%">{r["trabajador"] or ""}</td>'
                f'<td style="width:26%">{r["labor"] or ""}</td>'
                f'<td style="width:20%">{r["maquina"] or "–"}</td>'
                f'<td class="num" style="width:20%">{fmt(monto)}</td>'
                f'</tr>'
            )

    rows_html += (
        f'<tr style="background:#1e293b;color:#fff">'
        f'<td colspan="4" style="text-align:right;font-weight:bold">TOTAL</td>'
        f'<td class="num" style="width:20%;font-weight:bold">{fmt(grand_total)}</td>'
        f'</tr>'
    )

    # Segunda tabla: pivote por operador (fecha x trabajador — labor), misma
    # fuente que el Excel "Tabla por operador" y el preview en pantalla.
    pivot = _build_tractorista_pivot(pivot_rows)
    pivot_html = ""
    if pivot["dates"]:
        _check_pivot_date_range(pivot["columns"], "Por operador")
        widths = _pivot_col_widths({"fecha": 12, "total": 10}, len(pivot["columns"]))
        pivot_ths = "".join(
            f'<th class="num" style="{widths["date"]}">{_escape_html(c)}</th>'
            for c in pivot["columns"]
        )
        pivot_body = ""
        for d in pivot["dates"]:
            d_fmt = datetime.date.fromisoformat(d).strftime("%d/%m/%Y")
            cells = "".join(
                f'<td class="num" style="{widths["date"]}">{fmt(pivot["matrix"][d][c])}</td>'
                for c in pivot["columns"]
            )
            pivot_body += (
                f'<tr><td style="{widths["fecha"]}">{d_fmt}</td>{cells}'
                f'<td class="num" style="{widths["total"]}">{fmt(pivot["date_totals"][d])}</td></tr>'
            )
        pivot_totals = "".join(
            f'<td class="num" style="{widths["date"]};font-weight:bold">{fmt(pivot["col_totals"][c])}</td>'
            for c in pivot["columns"]
        )
        pivot_html = f"""
        <table class="data" style="{widths['table']};border-collapse:collapse;font-size:8pt;table-layout:fixed;margin-top:16px">
          <thead><tr>
            <th style="{widths["fecha"]}">Fecha</th>{pivot_ths}
            <th class="num" style="{widths["total"]}">Total</th>
          </tr></thead>
          <tbody>{pivot_body}
            <tr style="background:#1e293b;color:#fff">
              <td style="{widths["fecha"]};font-weight:bold">TOTAL</td>{pivot_totals}
              <td class="num" style="{widths["total"]};font-weight:bold">{fmt(pivot["grand_total"])}</td>
            </tr>
          </tbody>
        </table>
        """

    logo = _logo_b64()
    logo_html = f'<img src="data:image/png;base64,{logo}" style="width:80px;height:auto" />' if logo else ""
    fi_fmt = datetime.date.fromisoformat(fecha_inicio).strftime("%d/%m/%Y")
    ft_fmt = datetime.date.fromisoformat(fecha_termino).strftime("%d/%m/%Y")
    cc_line = f'<div class="sub">CC {cc}</div>' if cc else ""

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
    <style>
      @page {{ size: A4; margin: 12mm 14mm; }}
      body {{ font-family: Helvetica, Arial, sans-serif; font-size: 8.5pt; color: #111; margin: 0; }}
      .hdr {{ width: 100%; border-collapse: collapse; background: #1e293b; margin-bottom: 12px; }}
      .hdr td {{ padding: 10px 14px; color: #fff; vertical-align: middle; }}
      .hdr .title {{ font-size: 13pt; font-weight: 800; }}
      .hdr .sub {{ font-size: 8.5pt; margin-top: 2px; opacity: .85; }}
      table.data {{ width: 100%; border-collapse: collapse; font-size: 8pt; table-layout: fixed; }}
      table.data th {{ background: #1e293b; color: #fff; padding: 5px 7px; text-align: left; word-wrap: break-word; }}
      table.data td {{ padding: 4px 7px; border-bottom: 1px solid #e2e8f0; vertical-align: top; word-wrap: break-word; }}
      table.data tr:nth-child(even) {{ background: #f8fafc; }}
      .num {{ text-align: right; }}
      .section-title {{ font-size: 10pt; font-weight: 800; margin: 4px 0 6px; }}
    </style>
    </head><body>
    <table class="hdr">
      <tr>
        <td style="width:100px">{logo_html}</td>
        <td>
          <div class="title">Detalle Tarjas Tractoristas</div>
          <div class="sub">{contratista} &mdash; {campo}</div>
          {cc_line}
          <div class="sub">{fi_fmt} al {ft_fmt}</div>
        </td>
      </tr>
    </table>
    <table class="data">
      <thead><tr>
        <th style="width:10%">Fecha</th>
        <th style="width:24%">Trabajador</th>
        <th style="width:26%">Labor</th>
        <th style="width:20%">Máquina</th>
        <th class="num" style="width:20%">Total a pagar</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    {'<div class="section-title">Tabla por operador</div>' + pivot_html if pivot_html else ''}
    </body></html>"""

    cc_slug = f"_cc{cc.replace(' ', '')}" if cc else ""
    filename = (
        f"tractorista_{contratista.replace(' ', '_')}_{campo.replace(' ', '_')}"
        f"{cc_slug}_{fecha_inicio}_{fecha_termino}.pdf"
    )
    return _render_pdf(html, filename)


# ---------------------------------------------------------------------------
# Nota de crédito print-PDF  (opens in new tab, same layout as the web page)
# ---------------------------------------------------------------------------

_NOTA_PDF_CSS = """
@page { size: A4 landscape; margin: 12mm 14mm; }
body { font-family: Helvetica, Arial, sans-serif; font-size: 8pt; color: #111; margin: 0; }
.hdr-table { width: 100%; border-collapse: collapse; border: 1px solid #cbd5e1; margin-bottom: 10px; }
.hdr-logo  { width: 120px; padding: 8px 12px; vertical-align: middle; text-align: center;
             border-right: 1px solid #e2e8f0; background: #1e293b; }
.hdr-logo-img { width: 100px; height: auto; }
.hdr-mid   { padding: 10px 14px; vertical-align: top; }
.hdr-right { width: 190px; padding: 10px 14px; vertical-align: top; text-align: right;
             border-left: 1px solid #e2e8f0; }
.co-name   { font-size: 14pt; font-weight: 800; margin: 0 0 3px; }
.co-sub    { font-size: 9pt; font-weight: 600; margin: 0 0 2px; }
.co-week   { font-size: 7.5pt; color: #64748b; margin: 0; }
.dt-row    { margin-bottom: 5px; font-size: 8pt; }
.dt-label  { font-weight: 600; }
.dt-val    { background: #fef9c3; padding: 1px 7px; font-weight: 700; }
.grand-tot { font-size: 15pt; font-weight: 800; margin-top: 8px; }
.glosa-table { width: 100%; border-collapse: collapse; border: 1px solid #cbd5e1; margin-bottom: 10px; }
.glosa-title { background: #1e293b; color: white; text-align: center;
               padding: 6px; font-weight: 700; font-size: 8pt; letter-spacing: .5px; }
.glosa-body  { background: #fef08a; text-align: center; padding: 8px 12px;
               font-size: 8.5pt; font-weight: 700; color: #1e293b; }
.totals-row  { border-top: 2px solid #1e293b; }
.tot-cell    { text-align: center; padding: 8px 6px; width: 33%;
               border-right: 1px solid #e2e8f0; }
.tot-cell-hl { background: #fef08a; }
.tot-label   { font-size: 7pt; font-weight: 700; color: #64748b; text-transform: uppercase; }
.tot-value   { font-size: 13pt; font-weight: 800; color: #1e293b; margin-top: 2px; }
.detail-table { width: 100%; border-collapse: collapse; font-size: 7.5pt; margin-top: 8px; }
.detail-table thead tr { background: #1d4ed8; color: white; }
.detail-table thead th { padding: 6px 8px; text-align: left; font-weight: bold; }
.detail-table thead th.num { text-align: right; }
.detail-table tbody tr.even { background: #f8fafc; }
.detail-table tbody td { padding: 4px 8px; border-bottom: 1px solid #f1f5f9; }
.detail-table td.num { text-align: right; }
.badge-trato { color: #1d4ed8; font-weight: 700; }
.badge-aldia { color: #15803d; font-weight: 700; }
"""


@router.get("/api/tarjas/notas/print-pdf")
async def notas_print_pdf(
    contratista: str = Query(...),
    campo: str = Query(None),
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    nc_total: int = Query(None),
):
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    filters = [
        "fecha::date BETWEEN %s AND %s",
        "contratista = %s",
        "estado = 'Aprobado'",
    ]
    params: list = [fecha_inicio, fecha_termino, contratista]
    if campo:
        filters.append("nombre_campo = %s")
        params.append(campo)
    where = "WHERE " + " AND ".join(filters)

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT tipo_pago, cuartel_cc AS cc, labor,
                       COUNT(*) AS jornadas,
                       ROUND(CASE
                         WHEN LOWER(TRIM(tipo_pago)) IN ('a trato','trato')
                           THEN AVG(NULLIF(total_trato, 0))
                         ELSE AVG(NULLIF(total_jornada, 0))
                       END::numeric, 0) AS total_unitario,
                       COALESCE(SUM(total_pagar), 0) AS total_pagar
                FROM appsheet.tarjas_pagos
                {where}
                GROUP BY tipo_pago, cuartel_cc, labor
                ORDER BY
                    CASE WHEN LOWER(TRIM(tipo_pago)) IN ('a trato','trato') THEN 0 ELSE 1 END,
                    cuartel_cc, labor
            """,
                params,
            )
            rows = cur.fetchall()

            cur.execute(
                f"SELECT DISTINCT nombre_campo FROM appsheet.tarjas_pagos {where} LIMIT 1",
                params,
            )
            r = cur.fetchone()
            nombre_campo = r[0] if r else (campo or "")
    finally:
        conn.close()

    if not rows:
        raise HTTPException(
            status_code=404, detail="Sin datos para los filtros indicados"
        )

    total_general = sum(float(r[5] or 0) for r in rows)

    # Apply NC redistribution if requested (largest-remainder method)
    if nc_total and nc_total > 0 and total_general > 0:
        scale = nc_total / total_general
        exact = [float(r[5] or 0) * scale for r in rows]
        floored = [int(v) for v in exact]
        remainder = nc_total - sum(floored)
        fracs = sorted(range(len(exact)), key=lambda i: -(exact[i] - floored[i]))
        for i in range(remainder):
            floored[fracs[i]] += 1
        rows = [(*r[:5], floored[i]) for i, r in enumerate(rows)]
        total_general = nc_total

    total_trato = sum(
        float(r[5] or 0)
        for r in rows
        if (r[0] or "").lower().strip() in ("a trato", "trato")
    )
    total_aldia = total_general - total_trato

    d1 = _fmt_date_display(fecha_inicio)
    d2 = _fmt_date_display(fecha_termino)
    glosa = f"SERVICIOS DE LABORES AGRÍCOLAS {d1.upper()} AL {d2.upper()}"
    semana = f"Semana desde {d1} al {d2}"
    empresa_display = (
        f"AGRÍCOLA DONAR — {nombre_campo.upper()}" if nombre_campo else "AGRÍCOLA DONAR"
    )

    rows_html = ""
    for i, (tipo, cc, labor, jornadas, unitario, total) in enumerate(rows):
        is_trato = (tipo or "").lower().strip() in ("trato", "a trato")
        tipo_label = "Trato" if is_trato else "Al día"
        tipo_cls = "badge-trato" if is_trato else "badge-aldia"
        even_cls = "even" if i % 2 == 0 else ""
        rows_html += f"""<tr class="{even_cls}">
          <td><span class="{tipo_cls}">{tipo_label}</span></td>
          <td>{cc or ""}</td>
          <td>{labor or ""}</td>
          <td class="num">{int(jornadas) if jornadas is not None else "–"}</td>
          <td class="num">{_fmt_clp(unitario)}</td>
          <td class="num">{_fmt_clp(total)}</td>
        </tr>"""

    logo = _logo_b64()
    logo_html = (
        f'<img src="data:image/png;base64,{logo}" class="hdr-logo-img" />'
        if logo
        else "EMPRESAS DONAR"
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Nota de Crédito — {contratista}</title>
<style>{_NOTA_PDF_CSS}</style>
</head><body>

<table class="hdr-table">
  <tr>
    <td class="hdr-logo">{logo_html}</td>
    <td class="hdr-mid">
      <p class="co-name">{empresa_display}</p>
      <p class="co-sub">Contratista: {contratista}</p>
      <p class="co-week">{semana}</p>
    </td>
    <td class="hdr-right">
      <p class="dt-row"><span class="dt-label">Fecha Inicio&nbsp;&nbsp;</span>
        <span class="dt-val">{d1}</span></p>
      <p class="dt-row"><span class="dt-label">Fecha Término&nbsp;&nbsp;</span>
        <span class="dt-val">{d2}</span></p>
      <p class="grand-tot">{_fmt_clp(total_general)}</p>
    </td>
  </tr>
</table>

<table class="glosa-table">
  <tr><td colspan="3" class="glosa-title">GLOSA</td></tr>
  <tr><td colspan="3" class="glosa-body">{glosa}</td></tr>
  <tr class="totals-row">
    <td class="tot-cell"><div class="tot-label">Total a Trato</div>
      <div class="tot-value">{_fmt_clp(total_trato)}</div></td>
    <td class="tot-cell"><div class="tot-label">Total Al Día</div>
      <div class="tot-value">{_fmt_clp(total_aldia)}</div></td>
    <td class="tot-cell tot-cell-hl"><div class="tot-label">Total a Pagar</div>
      <div class="tot-value">{_fmt_clp(total_general)}</div></td>
  </tr>
</table>

<table class="detail-table">
  <thead>
    <tr>
      <th>Tipo de Pago</th><th>CC</th><th>Nombre Labor</th>
      <th class="num">Jornadas</th><th class="num">Precio Unitario</th>
      <th class="num">Total a Pagar</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>

</body></html>"""

    buf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=buf)
    buf.seek(0)
    filename = (
        f"nota_{contratista.replace(' ', '_')}_{(campo or 'all').replace(' ', '_')}"
        f"_{fecha_inicio}_{fecha_termino}.pdf"
    )
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


# ===========================================================================
# App → Registros de campo — audit timeline of tarjas_pagos (+ mal-digitado edits)
# ===========================================================================

_REGISTROS_CAMPO_KNOWN_ESTADOS = {"aprobado", "pendiente"}
_REGISTROS_CAMPO_EXCEL_MAX = 10000
_REGISTROS_CAMPO_COLUMNS = [
    "id_Resumen",
    "id_supervisor",
    "fecha",
    "fecha_iso",
    "nombre_campo",
    "cuartel_cc",
    "labor",
    "contratista",
    "trabajador",
    "rut_trabajador",
    "tipo_pago",
    "valor_jornada",
    "valor_trato",
    "base_trato",
    "rendimiento",
    "horas_extras",
    "horas_trabajadas",
    "maquina",
    "total_tractor",
    "total_hora_extra",
    "total_jornada",
    "total_trato",
    "total_trabajado",
    "contratista_jornada",
    "contratista_trato",
    "total_contratista",
    "total_pagar",
    "estado",
    "id_tarja_supervisor",
    "id_labor",
]
_REGISTROS_CAMPO_SELECT = """
SELECT
  "id_Resumen",
  id_supervisor,
  fecha,
  fecha::date AS fecha_iso,
  nombre_campo,
  cuartel_cc,
  labor,
  contratista,
  trabajador,
  rut_trabajador,
  tipo_pago,
  valor_jornada,
  valor_trato,
  base_trato,
  rendimiento,
  horas_extras,
  horas_trabajadas,
  maquina,
  total_tractor,
  total_hora_extra,
  total_jornada,
  total_trato,
  total_trabajado,
  contratista_jornada,
  contratista_trato,
  total_contratista,
  total_pagar,
  estado,
  id_tarja_supervisor,
  id_labor
FROM appsheet.tarjas_pagos
"""


_REGISTROS_CAMPO_EDITABLE = frozenset(
    {
        "trabajador",
        "rut_trabajador",
        "horas_trabajadas",
        "horas_extras",
    }
)
_MAL_DIGITADO_FLAGS = frozenset(
    {
        "bad_rut",
        "double_space_name",
        "name_has_digits",
        "implausible_horas_extra",
        "hours_and_extras",
        "implausible_hours",
    }
)
# Same heuristics as _registros_campo_flags / _is_mal_digitado, for GROUP BY.
_HORAS_NUM_SQL = (
    "CASE WHEN NULLIF(TRIM({col}::text), '') ~ '^[+-]?[0-9]+([.,][0-9]+)?$' "
    "THEN REPLACE(TRIM({col}::text), ',', '.')::numeric END"
)
_REGISTROS_CAMPO_MAL_SQL = (
    "("
    " (NULLIF(TRIM(rut_trabajador), '') IS NOT NULL AND ("
    "   rut_trabajador ~* '[eE][+\\-]' "
    "   OR length(regexp_replace(rut_trabajador, '[^0-9kK]', '', 'g')) "
    "      BETWEEN 1 AND 6"
    " ))"
    " OR (NULLIF(TRIM(trabajador), '') IS NOT NULL AND ("
    "   position('  ' in trabajador) > 0 OR trabajador ~ '[0-9]'"
    " ))"
    f" OR COALESCE({_HORAS_NUM_SQL.format(col='horas_extras')}, 0) > 8"
    f" OR COALESCE({_HORAS_NUM_SQL.format(col='horas_trabajadas')}, 0) > 14"
    f" OR (COALESCE({_HORAS_NUM_SQL.format(col='horas_trabajadas')}, 0) > 0"
    f"     AND COALESCE({_HORAS_NUM_SQL.format(col='horas_extras')}, 0) > 0)"
    ")"
)


def _is_blank(v) -> bool:
    if v is None:
        return True
    return str(v).strip() == ""


def _as_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float, decimal.Decimal)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _rut_is_malformed(v) -> bool:
    """True when a present RUT looks typed/exported wrong, not when it is missing."""
    if _is_blank(v):
        return False
    s = str(v).strip()
    if re.search(r"[eE][+\-]", s):
        return True
    digits = re.sub(r"[^0-9kK]", "", s)
    return 0 < len(digits) < 7


def _registros_campo_flags(row: dict) -> list[str]:
    """Heuristic QA flags for a raw field record. No DB writes."""
    flags: list[str] = []
    if _is_blank(row.get("labor")):
        flags.append("missing_labor")
    if _is_blank(row.get("nombre_campo")):
        flags.append("missing_campo")
    trabajador = row.get("trabajador")
    if _is_blank(trabajador):
        flags.append("missing_trabajador")
    else:
        name = str(trabajador)
        if "  " in name:
            flags.append("double_space_name")
        if re.search(r"\d", name):
            flags.append("name_has_digits")
    if _is_blank(row.get("estado")):
        flags.append("missing_estado")
    elif str(row["estado"]).strip().lower() not in _REGISTROS_CAMPO_KNOWN_ESTADOS:
        flags.append("unexpected_estado")
    if _is_blank(row.get("fecha")) or _is_blank(row.get("fecha_iso")):
        flags.append("invalid_fecha")
    if _is_blank(row.get("id_supervisor")):
        flags.append("missing_supervisor")
    if _rut_is_malformed(row.get("rut_trabajador")):
        flags.append("bad_rut")
    horas = _as_number(row.get("horas_trabajadas"))
    extras = _as_number(row.get("horas_extras"))
    if extras is not None and extras > 8:
        flags.append("implausible_horas_extra")
    if horas is not None and horas > 14:
        flags.append("implausible_hours")
    # 0 regular hours is valid when overtime was entered instead.
    # Entering both horas_trabajadas and horas_extras on the same row is the typo.
    if (horas or 0) > 0 and (extras or 0) > 0:
        flags.append("hours_and_extras")
    return flags


def _is_mal_digitado(flags: list[str]) -> bool:
    return any(f in _MAL_DIGITADO_FLAGS for f in flags)


def _build_registros_campo_where(
    fecha_inicio: str,
    fecha_termino: str,
    empresa: str | None = None,
    labor: str | None = None,
    estado: str | None = None,
    contratista: str | None = None,
    supervisor: str | None = None,
) -> tuple[str, list]:
    filters = ["fecha::date BETWEEN %s AND %s"]
    params: list = [fecha_inicio, fecha_termino]
    if empresa:
        filters.append("nombre_campo = %s")
        params.append(_empresa_to_campo(empresa))
    if labor:
        filters.append("labor = %s")
        params.append(labor)
    if estado:
        filters.append("estado = %s")
        params.append(estado)
    if contratista:
        filters.append("contratista = %s")
        params.append(contratista)
    if supervisor:
        filters.append("id_supervisor = %s")
        params.append(supervisor)
    return "WHERE " + " AND ".join(filters), params


def _annotate_registros_campo(rows: list[dict]) -> list[dict]:
    for row in rows:
        flags = _registros_campo_flags(row)
        row["flags"] = flags
        row["mal_digitado"] = _is_mal_digitado(flags)
    return rows


@router.get("/tarjas/registros-campo", response_class=HTMLResponse)
async def tarjas_registros_campo_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_registros_campo.html")


@router.get("/api/tarjas/registros-campo/filters")
async def get_tarjas_registros_campo_filters():
    """Distinct dropdown values for the field-record timeline."""
    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )
    try:
        with conn.cursor() as cur:
            empresas = _get_empresas(cur, "appsheet.tarjas_pagos")
            cur.execute(
                "SELECT DISTINCT labor FROM appsheet.tarjas_pagos "
                "WHERE labor IS NOT NULL AND TRIM(labor) <> '' ORDER BY labor"
            )
            labores = [r[0] for r in cur.fetchall()]
            cur.execute(
                "SELECT DISTINCT estado FROM appsheet.tarjas_pagos "
                "WHERE estado IS NOT NULL AND TRIM(estado) <> '' ORDER BY estado"
            )
            estados = [r[0] for r in cur.fetchall()]
            cur.execute(
                "SELECT DISTINCT contratista FROM appsheet.tarjas_pagos "
                "WHERE contratista IS NOT NULL AND TRIM(contratista) <> '' "
                "ORDER BY contratista"
            )
            contratistas = [r[0] for r in cur.fetchall()]
            cur.execute(
                "SELECT DISTINCT id_supervisor FROM appsheet.tarjas_pagos "
                "WHERE id_supervisor IS NOT NULL AND TRIM(id_supervisor) <> '' "
                "ORDER BY id_supervisor"
            )
            supervisores = [r[0] for r in cur.fetchall()]
    finally:
        conn.close()

    return {
        "empresas": empresas,
        "labores": labores,
        "estados": estados,
        "contratistas": contratistas,
        "supervisores": supervisores,
    }


@router.get("/api/tarjas/registros-campo")
async def get_tarjas_registros_campo(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    empresa: str = Query(None),
    labor: str = Query(None),
    estado: str = Query(None),
    contratista: str = Query(None),
    supervisor: str = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    """Paginated chronological dump of raw appsheet.tarjas_pagos rows."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    where, params = _build_registros_campo_where(
        fecha_inicio,
        fecha_termino,
        empresa=empresa,
        labor=labor,
        estado=estado,
        contratista=contratista,
        supervisor=supervisor,
    )
    page_params = params + [limit, offset]

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT COUNT(*) FROM appsheet.tarjas_pagos {where}",
                params,
            )
            total = int(cur.fetchone()[0])
            cur.execute(
                f"{_REGISTROS_CAMPO_SELECT} {where} "
                'ORDER BY fecha::date DESC, "id_Resumen" DESC '
                "LIMIT %s OFFSET %s",
                page_params,
            )
            rows = _annotate_registros_campo(_rows_to_dicts(cur))
    finally:
        conn.close()

    return {
        "columns": _REGISTROS_CAMPO_COLUMNS,
        "rows": rows,
        "count": len(rows),
        "total": total,
        "limit": limit,
        "offset": offset,
    }


def _coerce_registro_campo_edit(col: str, raw):
    if col in ("horas_trabajadas", "horas_extras"):
        if raw is None or str(raw).strip() == "":
            return None
        n = _as_number(raw)
        if n is None:
            raise HTTPException(
                status_code=422, detail=f"Valor inválido para {col}"
            )
        if n == int(n):
            return str(int(n))
        return str(n)
    if raw is None:
        return ""
    return str(raw)


@router.patch("/api/tarjas/registros-campo/{id_resumen}")
async def patch_tarjas_registro_campo(
    id_resumen: str,
    payload: dict = Body(...),
):
    """Update allowlisted typo fields on a tarjas_pagos row."""
    if not id_resumen or not str(id_resumen).strip():
        raise HTTPException(status_code=400, detail="id_resumen is required")
    fields = payload.get("fields") if isinstance(payload, dict) else None
    if not isinstance(fields, dict) or not fields:
        raise HTTPException(status_code=422, detail="fields is required")
    unknown = [k for k in fields if k not in _REGISTROS_CAMPO_EDITABLE]
    if unknown:
        raise HTTPException(
            status_code=422,
            detail=f"Campos no editables: {', '.join(unknown)}",
        )
    cols = [k for k in _REGISTROS_CAMPO_EDITABLE if k in fields]
    values = [_coerce_registro_campo_edit(k, fields[k]) for k in cols]

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    assign = psql.SQL(", ").join(
        psql.SQL("{} = %s").format(psql.Identifier(col)) for col in cols
    )
    update_sql = psql.SQL(
        "UPDATE appsheet.tarjas_pagos SET {} WHERE {} = %s RETURNING {}"
    ).format(assign, psql.Identifier("id_Resumen"), psql.Identifier("id_Resumen"))

    rows: list[dict] = []
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(update_sql, values + [id_resumen])
                updated = cur.fetchone()
                if updated is None:
                    raise HTTPException(
                        status_code=404, detail="Registro no encontrado"
                    )
                cur.execute(
                    f'{_REGISTROS_CAMPO_SELECT} WHERE "id_Resumen" = %s',
                    (id_resumen,),
                )
                rows = _annotate_registros_campo(_rows_to_dicts(cur))
    finally:
        conn.close()

    if not rows:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    return {"row": rows[0]}


@router.get("/api/tarjas/registros-campo/download-excel")
async def download_tarjas_registros_campo_excel(
    fecha_inicio: str = Query(...),
    fecha_termino: str = Query(...),
    empresa: str = Query(None),
    labor: str = Query(None),
    estado: str = Query(None),
    contratista: str = Query(None),
    supervisor: str = Query(None),
):
    """Excel dump of the filtered field-record timeline (capped)."""
    if not _DATE_RE.match(fecha_inicio) or not _DATE_RE.match(fecha_termino):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    where, params = _build_registros_campo_where(
        fecha_inicio,
        fecha_termino,
        empresa=empresa,
        labor=labor,
        estado=estado,
        contratista=contratista,
        supervisor=supervisor,
    )
    excel_params = params + [_REGISTROS_CAMPO_EXCEL_MAX]

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"{_REGISTROS_CAMPO_SELECT} {where} "
                'ORDER BY fecha::date DESC, "id_Resumen" DESC '
                "LIMIT %s",
                excel_params,
            )
            rows = _annotate_registros_campo(_rows_to_dicts(cur))
    finally:
        conn.close()

    headers = [
        ("fecha_iso", "Fecha"),
        ("estado", "Estado"),
        ("labor", "Labor"),
        ("id_supervisor", "Supervisor"),
        ("nombre_campo", "Campo"),
        ("cuartel_cc", "Cuartel / CC"),
        ("trabajador", "Trabajador"),
        ("rut_trabajador", "RUT"),
        ("contratista", "Contratista"),
        ("tipo_pago", "Tipo de pago"),
        ("horas_trabajadas", "Horas trabajadas"),
        ("horas_extras", "Horas extras"),
        ("rendimiento", "Rendimiento"),
        ("maquina", "Máquina"),
        ("total_jornada", "Total jornada"),
        ("total_trato", "Total trato"),
        ("total_trabajado", "Total trabajado"),
        ("total_pagar", "Total a pagar"),
        ("total_tractor", "Total tractor"),
        ("id_Resumen", "ID resumen"),
        ("flags", "Alertas"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros de campo"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    warn_fill = PatternFill("solid", fgColor="FDE68A")

    for col, (_, label) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, r in enumerate(rows, 2):
        flags = r.get("flags") or []
        flagged = bool(flags)
        for col, (key, _) in enumerate(headers, 1):
            if key == "flags":
                value = ", ".join(flags)
            elif key == "fecha_iso":
                iso = str(r.get("fecha_iso") or "")[:10]
                try:
                    d = datetime.date.fromisoformat(iso)
                    value = d.strftime("%d/%m/%Y")
                except Exception:
                    value = iso or (r.get("fecha") or "")
            else:
                value = r.get(key)
                if value is None:
                    value = ""
            cell = ws.cell(row=row_num, column=col, value=value)
            if flagged:
                cell.fill = warn_fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["I"].width = 28

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"registros_campo_{fecha_inicio}_{fecha_termino}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ===========================================================================
# App → Calendario — monthly heatmap of tarjas_pagos + tarjas_plan_diario
# ===========================================================================

_PLAN_DIARIO_FROM = """
appsheet.tarjas_plan_diario p
LEFT JOIN appsheet.tarjas_contratistas c
  ON c.id_contratista = p.id_contratista
LEFT JOIN appsheet.tarjas_cc cc
  ON TRIM(cc.id_cc::text) = TRIM(p.id_cc)
LEFT JOIN appsheet.tarjas_campo campo
  ON campo.id_campo::text = TRIM(cc.id_campo::text)
LEFT JOIN appsheet.tarjas_trato t
  ON t.id_trato = p.id_trato
LEFT JOIN LATERAL (
  SELECT nombre
  FROM appsheet.tarjas_labor
  WHERE TRIM(id_labor::text) = TRIM(t.id_labor)
  ORDER BY nombre
  LIMIT 1
) l ON TRUE
"""

_PLAN_DIARIO_DATE_OK = (
    r"p.fecha_inicio ~ '^\d{4}-\d{2}-\d{2}' "
    r"AND (p.fecha_fin IS NULL OR TRIM(p.fecha_fin) = '' "
    r"OR p.fecha_fin ~ '^\d{4}-\d{2}-\d{2}')"
)

_PLAN_DIARIO_END = (
    "COALESCE(NULLIF(TRIM(p.fecha_fin), '')::date, p.fecha_inicio::date)"
)


def _month_to_date_range(mes: str) -> tuple[str, str]:
    year, month = int(mes[:4]), int(mes[5:7])
    start = datetime.date(year, month, 1)
    if month == 12:
        end = datetime.date(year, 12, 31)
    else:
        end = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
    return start.isoformat(), end.isoformat()


def _month_grid_range(mes: str) -> tuple[str, str]:
    """Monday of the first week through Sunday of the last week."""
    month_start, month_end = _month_to_date_range(mes)
    start = datetime.date.fromisoformat(month_start)
    end = datetime.date.fromisoformat(month_end)
    grid_start = start - datetime.timedelta(days=start.weekday())
    grid_end = end + datetime.timedelta(days=(6 - end.weekday()))
    return grid_start.isoformat(), grid_end.isoformat()


@router.get("/tarjas/calendario", response_class=HTMLResponse)
async def tarjas_calendario_page(request: Request):
    return _templates.TemplateResponse(request, "tarjas_calendario.html")


def _build_plan_diario_where(
    fecha_inicio: str,
    fecha_termino: str,
    empresa: str | None = None,
    labor: str | None = None,
    contratista: str | None = None,
) -> tuple[str, list]:
    """Range overlap: plan paints every day from fecha_inicio to fecha_fin.

    estado / supervisor apply only to tarjas_pagos, not to planificación.
    """
    filters = [
        _PLAN_DIARIO_DATE_OK,
        "p.fecha_inicio::date <= %s",
        f"{_PLAN_DIARIO_END} >= %s",
    ]
    params: list = [fecha_termino, fecha_inicio]
    if empresa:
        filters.append("campo.nombre = %s")
        params.append(_empresa_to_campo(empresa))
    if labor:
        filters.append(
            "LOWER(COALESCE(NULLIF(TRIM(p.labor_especifica), ''), "
            "l.nombre, '')) LIKE LOWER(%s)"
        )
        params.append(f"%{labor}%")
    if contratista:
        filters.append("c.nombre = %s")
        params.append(contratista)
    return "WHERE " + " AND ".join(filters), params


def _empty_calendar_day(fecha: str) -> dict:
    return {
        "fecha": fecha,
        "total": 0,
        "aprobado": 0,
        "pendiente": 0,
        "sospechosos": 0,
        "planes": 0,
    }


@router.get("/api/tarjas/calendario")
async def get_tarjas_calendario(
    mes: str = Query(..., description="YYYY-MM"),
    empresa: str = Query(None),
    labor: str = Query(None),
    estado: str = Query(None),
    contratista: str = Query(None),
    supervisor: str = Query(None),
):
    """Per-day registro counts plus overlapping planificación for the month."""
    if not _MONTH_RE.match(mes):
        raise HTTPException(status_code=400, detail="mes must be YYYY-MM")

    month_start, month_end = _month_to_date_range(mes)
    grid_start, grid_end = _month_grid_range(mes)

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    where, params = _build_registros_campo_where(
        grid_start,
        grid_end,
        empresa=empresa,
        labor=labor,
        estado=estado,
        contratista=contratista,
        supervisor=supervisor,
    )
    plan_where, plan_params = _build_plan_diario_where(
        grid_start,
        grid_end,
        empresa=empresa,
        labor=labor,
        contratista=contratista,
    )
    plan_where_month, plan_params_month = _build_plan_diario_where(
        month_start,
        month_end,
        empresa=empresa,
        labor=labor,
        contratista=contratista,
    )

    by_fecha: dict[str, dict] = {}
    total = aprobado = pendiente = max_count = 0
    sospechosos_total = 0
    planes_total = 0
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                  fecha::date::text AS fecha,
                  COUNT(*) AS total,
                  COUNT(*) FILTER (
                    WHERE LOWER(TRIM(estado)) = 'aprobado'
                  ) AS aprobado,
                  COUNT(*) FILTER (
                    WHERE LOWER(TRIM(estado)) = 'pendiente'
                  ) AS pendiente,
                  COUNT(*) FILTER (
                    WHERE {_REGISTROS_CAMPO_MAL_SQL}
                  ) AS sospechosos
                FROM appsheet.tarjas_pagos
                {where}
                GROUP BY fecha::date
                ORDER BY fecha::date
                """,
                params,
            )
            for row in cur.fetchall():
                fecha = row[0]
                if not fecha:
                    continue
                day = _empty_calendar_day(fecha)
                day["total"] = int(row[1] or 0)
                day["aprobado"] = int(row[2] or 0)
                day["pendiente"] = int(row[3] or 0)
                day["sospechosos"] = int(row[4] or 0)
                by_fecha[day["fecha"]] = day
                if month_start <= fecha <= month_end:
                    total += day["total"]
                    aprobado += day["aprobado"]
                    pendiente += day["pendiente"]
                    sospechosos_total += day["sospechosos"]
                    if day["total"] > max_count:
                        max_count = day["total"]

            try:
                cur.execute(
                    f"SELECT COUNT(DISTINCT p.id_plan) FROM {_PLAN_DIARIO_FROM} {plan_where_month}",
                    plan_params_month,
                )
                planes_total = int(cur.fetchone()[0] or 0)
                series_params = [grid_start, grid_end] + plan_params
                cur.execute(
                    f"""
                    SELECT gs.d::date::text AS fecha, COUNT(DISTINCT p.id_plan) AS planes
                    FROM {_PLAN_DIARIO_FROM}
                    CROSS JOIN LATERAL generate_series(
                      GREATEST(p.fecha_inicio::date, %s::date),
                      LEAST({_PLAN_DIARIO_END}, %s::date),
                      interval '1 day'
                    ) AS gs(d)
                    {plan_where}
                    GROUP BY gs.d
                    """,
                    series_params,
                )
                for row in cur.fetchall():
                    fecha = row[0]
                    if not fecha:
                        continue
                    if fecha not in by_fecha:
                        by_fecha[fecha] = _empty_calendar_day(fecha)
                    by_fecha[fecha]["planes"] = int(row[1] or 0)
            except Exception:
                logger.exception("Calendar plan overlay failed for mes=%s", mes)
                planes_total = 0
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Calendar month query failed for mes=%s", mes)
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    finally:
        conn.close()

    days = sorted(by_fecha.values(), key=lambda d: d["fecha"])
    return {
        "mes": mes,
        "fecha_inicio": month_start,
        "fecha_termino": month_end,
        "grid_inicio": grid_start,
        "grid_termino": grid_end,
        "days": days,
        "total": total,
        "aprobado": aprobado,
        "pendiente": pendiente,
        "sospechosos": sospechosos_total,
        "planes": planes_total,
        "max": max_count,
    }


@router.get("/api/tarjas/calendario/planes")
async def get_tarjas_calendario_planes(
    fecha: str = Query(..., description="YYYY-MM-DD"),
    empresa: str = Query(None),
    labor: str = Query(None),
    contratista: str = Query(None),
):
    """Planificación overlapping a calendar day (AppSheet paints the full span)."""
    if not _DATE_RE.match(fecha):
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(
            status_code=503, detail="Error de conexión a la base de datos"
        )

    where, params = _build_plan_diario_where(
        fecha,
        fecha,
        empresa=empresa,
        labor=labor,
        contratista=contratista,
    )

    try:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                  p.id_plan,
                  COALESCE(
                    NULLIF(TRIM(p.labor_especifica), ''),
                    l.nombre
                  ) AS labor,
                  p.comentario,
                  p.numero_personas,
                  p.id_cc,
                  cc.cultivo,
                  campo.nombre AS nombre_campo,
                  c.nombre AS contratista,
                  p.fecha_inicio::date::text AS fecha_inicio,
                  {_PLAN_DIARIO_END}::text AS fecha_fin,
                  ({_PLAN_DIARIO_END} - p.fecha_inicio::date) + 1 AS dias
                FROM {_PLAN_DIARIO_FROM}
                {where}
                ORDER BY
                  COALESCE(NULLIF(TRIM(p.labor_especifica), ''), l.nombre, ''),
                  c.nombre NULLS LAST,
                  p.id_cc
                """,
                params,
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    for row in rows:
        dias = int(row.get("dias") or 1)
        row["dias"] = dias
        row["largo"] = dias >= 7
        for key in ("fecha_inicio", "fecha_fin"):
            raw = row.get(key)
            row[key] = str(raw)[:10] if raw else None

    return {"fecha": fecha, "rows": rows, "count": len(rows)}

